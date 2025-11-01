#!/usr/bin/env python3
"""
Excel Report Generator untuk Laporan Kinerja Kerani, Mandor, dan Asisten
Mengimplementasikan logika lengkap yang sama dengan gui_multi_estate_ffb_analysis.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from datetime import datetime, date
from firebird_connector import FirebirdConnector
import os

class ExcelReportGenerator:
    def __init__(self):
        self.template_name = "Laporan_Kinerja_Kerani_Mandor_Asisten_Complete.xlsx"
        self.worksheet_name = "Laporan Kinerja"

    def generate_complete_report(self, config_file="config.json", start_date=None, end_date=None):
        """
        Generate laporan Excel lengkap dengan logika yang sama persis dengan GUI
        """
        # Load configuration
        with open(config_file, 'r') as f:
            estates = json.load(f)

        # Set default dates if not provided
        if start_date is None:
            start_date = date(2025, 5, 1)
        if end_date is None:
            end_date = date(2025, 5, 31)

        print(f"Generating report for period: {start_date} to {end_date}")

        # Process all estates
        all_results = []
        for estate_name, db_path in estates.items():
            try:
                print(f"Processing {estate_name}...")
                estate_results = self.analyze_estate_complete(estate_name, db_path, start_date, end_date)
                if estate_results:
                    all_results.extend(estate_results)
                    print(f"{estate_name}: {len(estate_results)} divisions")
            except Exception as e:
                print(f"Error processing {estate_name}: {e}")

        if all_results:
            print("Creating Excel report...")
            excel_path = self.create_excel_report(all_results, start_date, end_date)
            print(f"Excel report created: {excel_path}")
            return excel_path
        else:
            print("No data found")
            return None

    def analyze_estate_complete(self, estate_name, db_path, start_date, end_date):
        """
        Analisis lengkap per estate - sama persis dengan original code
        """
        # Handle path that is a folder
        if os.path.isdir(db_path):
            for file in os.listdir(db_path):
                if file.upper().endswith('.FDB'):
                    db_path = os.path.join(db_path, file)
                    break

        if not os.path.exists(db_path):
            print(f"Database not found: {db_path}")
            return None

        try:
            connector = FirebirdConnector(db_path)
            if not connector.test_connection():
                return None

            employee_mapping = self.get_employee_mapping(connector)
            divisions, month_tables = self.get_divisions(connector, start_date, end_date)

            # Special filter logic
            use_status_704_filter = (start_date.month == 5 or end_date.month == 5)

            if use_status_704_filter:
                print(f"  *** FILTER TRANSSTATUS 704 AKTIF untuk {estate_name} ***")

            estate_employee_totals = {}
            estate_results = []

            for div_id, div_name in divisions.items():
                result = self.analyze_division_complete(
                    connector, estate_name, div_id, div_name,
                    start_date, end_date, employee_mapping,
                    use_status_704_filter, month_tables
                )
                if result:
                    # Accumulate per employee
                    for emp_id, emp_data in result['employee_details'].items():
                        if emp_id not in estate_employee_totals:
                            estate_employee_totals[emp_id] = {
                                'name': emp_data['name'],
                                'kerani': 0,
                                'kerani_verified': 0,
                                'kerani_differences': 0,
                                'mandor': 0,
                                'asisten': 0
                            }

                        estate_employee_totals[emp_id]['kerani'] += emp_data['kerani']
                        estate_employee_totals[emp_id]['kerani_verified'] += emp_data['kerani_verified']
                        estate_employee_totals[emp_id]['kerani_differences'] += emp_data['kerani_differences']
                        estate_employee_totals[emp_id]['mandor'] += emp_data['mandor']
                        estate_employee_totals[emp_id]['asisten'] += emp_data['asisten']

                    estate_results.append(result)

            return estate_results

        except Exception as e:
            print(f"Error analyzing estate {estate_name}: {e}")
            return None

    def get_employee_mapping(self, connector):
        """Sama dengan original code"""
        query = "SELECT ID, NAME FROM EMP"
        try:
            result = connector.execute_query(query)
            df = connector.to_pandas(result)
            mapping = {}
            if not df.empty:
                for _, row in df.iterrows():
                    emp_id = str(row.iloc[0]).strip()
                    emp_name = str(row.iloc[1]).strip()
                    mapping[emp_id] = emp_name
            return mapping
        except:
            return {}

    def get_divisions(self, connector, start_date, end_date):
        """Sama dengan original code"""
        month_tables = []
        current_date = start_date
        while current_date <= end_date:
            month_tables.append(f"FFBSCANNERDATA{current_date.month:02d}")
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1, day=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1, day=1)

        month_tables = list(set(month_tables))
        print(f"  Tables to query: {', '.join(month_tables)}")

        all_divisions = {}
        for ffb_table in month_tables:
            query = f"""
            SELECT DISTINCT b.DIVID, c.DIVNAME
            FROM {ffb_table} a
            JOIN OCFIELD b ON a.FIELDID = b.ID
            LEFT JOIN CRDIVISION c ON b.DIVID = c.ID
            WHERE b.DIVID IS NOT NULL AND c.DIVNAME IS NOT NULL
            """

            try:
                result = connector.execute_query(query)
                df = connector.to_pandas(result)
                if not df.empty:
                    for _, row in df.iterrows():
                        div_id = str(row.iloc[0]).strip()
                        div_name = str(row.iloc[1]).strip()
                        if div_id not in all_divisions:
                            all_divisions[div_id] = div_name
            except Exception as e:
                print(f"Warning getting divisions from {ffb_table}: {e}")
                continue

        return all_divisions, month_tables

    def analyze_division_complete(self, connector, estate_name, div_id, div_name,
                                start_date, end_date, employee_mapping,
                                use_status_704_filter, month_tables):
        """Sama persis dengan original analyze_division method"""
        start_str = start_date.strftime('%Y-%m-%d')
        end_str = end_date.strftime('%Y-%m-%d')

        all_data_df = pd.DataFrame()

        for ffb_table in month_tables:
            query = f"""
            SELECT a.ID, a.SCANUSERID, a.OCID, a.WORKERID, a.CARRIERID, a.FIELDID, a.TASKNO,
                   a.RIPEBCH, a.UNRIPEBCH, a.BLACKBCH, a.ROTTENBCH, a.LONGSTALKBCH, a.RATDMGBCH,
                   a.LOOSEFRUIT, a.TRANSNO, a.TRANSDATE, a.TRANSTIME, a.UPLOADDATETIME,
                   a.RECORDTAG, a.TRANSSTATUS, a.TRANSTYPE, a.LASTUSER, a.LASTUPDATED,
                   a.OVERRIPEBCH, a.UNDERRIPEBCH, a.ABNORMALBCH, a.LOOSEFRUIT2
            FROM {ffb_table} a
            JOIN OCFIELD b ON a.FIELDID = b.ID
            WHERE b.DIVID = '{div_id}'
                AND a.TRANSDATE >= '{start_str}'
                AND a.TRANSDATE <= '{end_str}'
            """
            try:
                result = connector.execute_query(query)
                df_monthly = connector.to_pandas(result)
                if not df_monthly.empty:
                    all_data_df = pd.concat([all_data_df, df_monthly], ignore_index=True)
            except Exception as e:
                print(f"Warning getting data from {ffb_table}: {e}")
                continue

        df = all_data_df
        if df.empty:
            return None

        # Remove duplicates
        df.drop_duplicates(subset=['ID'], inplace=True)

        # Find duplicates (same logic as original)
        duplicated_rows = df[df.duplicated(subset=['TRANSNO'], keep=False)]
        verified_transnos = set(duplicated_rows['TRANSNO'].tolist())

        employee_details = {}

        # Initialize employee structure
        all_user_ids = df['SCANUSERID'].unique()
        for user_id in all_user_ids:
            user_id_str = str(user_id).strip()
            employee_details[user_id_str] = {
                'name': employee_mapping.get(user_id_str, f"EMP-{user_id_str}"),
                'kerani': 0,
                'kerani_verified': 0,
                'kerani_differences': 0,
                'mandor': 0,
                'asisten': 0
            }

        # Process Kerani data (RECORDTAG = 'PM')
        kerani_df = df[df['RECORDTAG'] == 'PM']
        if not kerani_df.empty:
            for user_id, group in kerani_df.groupby('SCANUSERID'):
                user_id_str = str(user_id).strip()
                total_created = len(group)

                # Count differences
                differences_count = 0
                for _, kerani_row in group.iterrows():
                    if kerani_row['TRANSNO'] in verified_transnos:
                        matching_transactions = df[(df['TRANSNO'] == kerani_row['TRANSNO']) &
                                                  (df['RECORDTAG'] != 'PM')]

                        if use_status_704_filter:
                            matching_transactions = matching_transactions[matching_transactions['TRANSSTATUS'] == '704']

                        if not matching_transactions.empty:
                            p1_records = matching_transactions[matching_transactions['RECORDTAG'] == 'P1']
                            p5_records = matching_transactions[matching_transactions['RECORDTAG'] == 'P5']

                            if not p1_records.empty:
                                other_row = p1_records.iloc[0]
                            elif not p5_records.empty:
                                other_row = p5_records.iloc[0]
                            else:
                                continue

                            # Compare fields
                            fields_to_compare = ['RIPEBCH', 'UNRIPEBCH', 'BLACKBCH', 'ROTTENBCH',
                                               'LONGSTALKBCH', 'RATDMGBCH', 'LOOSEFRUIT']

                            has_difference = False
                            for field in fields_to_compare:
                                try:
                                    kerani_val = float(kerani_row[field]) if kerani_row[field] else 0
                                    other_val = float(other_row[field]) if other_row[field] else 0
                                    if kerani_val != other_val:
                                        has_difference = True
                                        break
                                except (ValueError, TypeError):
                                    continue

                            if has_difference:
                                differences_count += 1

                verified_count = len(group[group['TRANSNO'].isin(verified_transnos)])

                if user_id_str in employee_details:
                    employee_details[user_id_str]['kerani'] = total_created
                    employee_details[user_id_str]['kerani_verified'] = verified_count
                    employee_details[user_id_str]['kerani_differences'] = differences_count

        # Process Mandor data (RECORDTAG = 'P1')
        mandor_df = df[df['RECORDTAG'] == 'P1']
        if not mandor_df.empty:
            mandor_counts = mandor_df.groupby('SCANUSERID').size()
            for user_id, count in mandor_counts.items():
                user_id_str = str(user_id).strip()
                if user_id_str in employee_details:
                    employee_details[user_id_str]['mandor'] = count

        # Process Asisten data (RECORDTAG = 'P5')
        asisten_df = df[df['RECORDTAG'] == 'P5']
        if not asisten_df.empty:
            asisten_counts = asisten_df.groupby('SCANUSERID').size()
            for user_id, count in asisten_counts.items():
                user_id_str = str(user_id).strip()
                if user_id_str in employee_details:
                    employee_details[user_id_str]['asisten'] = count

        # Calculate division totals
        kerani_total = sum(d['kerani'] for d in employee_details.values())
        mandor_total = sum(d['mandor'] for d in employee_details.values())
        asisten_total = sum(d['asisten'] for d in employee_details.values())

        div_kerani_verified_total = sum(d['kerani_verified'] for d in employee_details.values())
        verification_rate = (div_kerani_verified_total / kerani_total * 100) if kerani_total > 0 else 0

        return {
            'estate': estate_name,
            'division': div_name,
            'kerani_total': kerani_total,
            'mandor_total': mandor_total,
            'asisten_total': asisten_total,
            'verifikasi_total': div_kerani_verified_total,
            'verification_rate': verification_rate,
            'employee_details': employee_details
        }

    def create_excel_report(self, all_results, start_date, end_date):
        """Create Excel report with same format as PDF"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.worksheet_name

        # Set column widths
        column_widths = {
            'A': 15, 'B': 20, 'C': 25, 'D': 12, 'E': 18, 'F': 25, 'G': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Create styles
        self.create_styles()

        # Add headers
        headers = [
            "ESTATE", "DIVISI", "KARYAWAN", "ROLE",
            "JUMLAH TRANSAKSI", "PERSENTASE TERVERIFIKASI", "KETERANGAN PERBEDAAN"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        row_num = 2
        grand_kerani = 0
        grand_mandor = 0
        grand_asisten = 0
        grand_kerani_verified = 0

        # Process results (same logic as PDF)
        for result in all_results:
            estate = result['estate']
            division = result['division']
            kerani_total = result['kerani_total']
            mandor_total = result['mandor_total']
            asisten_total = result['asisten_total']
            verifikasi_total = result['verifikasi_total']
            verification_rate = result['verification_rate']
            employee_details = result['employee_details']

            # Division summary row
            total_kerani_only = kerani_total
            total_verified_kerani = verifikasi_total
            division_verification_rate = (total_verified_kerani / total_kerani_only * 100) if total_kerani_only > 0 else 0

            summary_data = [
                estate, division, f"== {division} TOTAL ==", "SUMMARY",
                total_kerani_only, f"{division_verification_rate:.2f}% ({total_verified_kerani})", ""
            ]

            for col_idx, value in enumerate(summary_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = self.summary_fill
                cell.font = self.summary_font
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

            row_num += 1

            # Employee rows
            for emp_id, emp_data in employee_details.items():
                # Kerani rows
                if emp_data['kerani'] > 0:
                    kerani_verification_rate = (emp_data.get('kerani_verified', 0) / emp_data['kerani'] * 100) if emp_data['kerani'] > 0 else 0
                    verified_count = emp_data.get('kerani_verified', 0)
                    differences_count = emp_data.get('kerani_differences', 0)
                    percentage_text = f"{kerani_verification_rate:.2f}% ({verified_count})"

                    difference_percentage = (differences_count / verified_count * 100) if verified_count > 0 else 0
                    keterangan_text = f"{differences_count} perbedaan ({difference_percentage:.1f}%)"

                    kerani_data = [
                        estate, division, emp_data['name'], "KERANI",
                        emp_data['kerani'], percentage_text, keterangan_text
                    ]

                    for col_idx, value in enumerate(kerani_data, 1):
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.fill = self.kerani_fill
                        cell.font = self.default_font
                        cell.alignment = self.center_alignment
                        cell.border = self.thin_border

                        if col_idx == 6:  # Percentage column
                            cell.font = self.kerani_percentage_font
                        elif col_idx == 7:  # Keterangan column
                            cell.font = self.keterangan_font

                    row_num += 1

                # Mandor rows
                if emp_data['mandor'] > 0:
                    mandor_percentage = (emp_data['mandor'] / kerani_total * 100) if kerani_total > 0 else 0
                    mandor_data = [
                        estate, division, emp_data['name'], "MANDOR",
                        emp_data['mandor'], f"{mandor_percentage:.2f}%", ""
                    ]

                    for col_idx, value in enumerate(mandor_data, 1):
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.fill = self.mandor_fill
                        cell.font = self.default_font
                        cell.alignment = self.center_alignment
                        cell.border = self.thin_border

                        if col_idx == 6:  # Percentage column
                            cell.font = self.mandor_percentage_font

                    row_num += 1

                # Asisten rows
                if emp_data['asisten'] > 0:
                    asisten_percentage = (emp_data['asisten'] / kerani_total * 100) if kerani_total > 0 else 0
                    asisten_data = [
                        estate, division, emp_data['name'], "ASISTEN",
                        emp_data['asisten'], f"{asisten_percentage:.2f}%", ""
                    ]

                    for col_idx, value in enumerate(asisten_data, 1):
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.fill = self.asisten_fill
                        cell.font = self.default_font
                        cell.alignment = self.center_alignment
                        cell.border = self.thin_border

                        if col_idx == 6:  # Percentage column
                            cell.font = self.asisten_percentage_font

                    row_num += 1

            # Add separator
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_num, column=col_idx, value="")
                cell.border = self.thin_border
            row_num += 1

            # Add to grand totals
            grand_kerani += kerani_total
            grand_mandor += mandor_total
            grand_asisten += asisten_total
            grand_kerani_verified += verifikasi_total

        # Grand total row
        grand_total_kerani_only = grand_kerani
        grand_total_verified_kerani = grand_kerani_verified
        grand_verification_rate = (grand_total_verified_kerani / grand_total_kerani_only * 100) if grand_total_kerani_only > 0 else 0

        grand_total_data = [
            "=== GRAND TOTAL ===", "", "", "",
            grand_total_kerani_only,
            f"{grand_verification_rate:.2f}% ({grand_total_verified_kerani})",
            ""
        ]

        for col_idx, value in enumerate(grand_total_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = self.grand_total_fill
            cell.font = self.grand_total_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        # Save file
        output_dir = "reports"
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        period = f"{start_date.strftime('%B_%Y')}"
        filename = f"Laporan_Kinerja_Excel_{period}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb.save(filepath)
        return filepath

    def create_styles(self):
        """Create all styles for the Excel report"""
        # Header styles
        self.header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        self.header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Summary styles
        self.summary_fill = PatternFill(start_color="4299E1", end_color="4299E1", fill_type="solid")
        self.summary_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

        # Kerani styles
        self.kerani_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
        self.kerani_percentage_font = Font(name="Arial", size=8, color="E53E3E")
        self.keterangan_font = Font(name="Arial", size=8, color="C53030", bold=True)

        # Mandor styles
        self.mandor_fill = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
        self.mandor_percentage_font = Font(name="Arial", size=8, color="38A169")

        # Asisten styles
        self.asisten_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
        self.asisten_percentage_font = Font(name="Arial", size=8, color="3182CE")

        # Grand total styles
        self.grand_total_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        self.grand_total_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        # Default styles
        self.default_font = Font(name="Arial", size=8)
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def main():
    """Main function to test the Excel generator"""
    generator = ExcelReportGenerator()

    print("=== Excel Report Generator for Laporan Kinerja ===")
    print("This will generate the same data as the PDF report")

    # Generate report
    excel_path = generator.generate_complete_report()

    if excel_path:
        print(f"\n✅ Excel report successfully generated:")
        print(f"   File: {excel_path}")
        print(f"   The report contains the same data and logic as the PDF report")
    else:
        print("\n❌ Failed to generate Excel report")

if __name__ == "__main__":
    main()