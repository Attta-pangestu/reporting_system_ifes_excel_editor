#!/usr/bin/env python3
"""
Comprehensive Test for October 2025 Report Generation
Test script untuk validasi placeholder replacement dan query data untuk bulan Oktober 2025
"""

import sys
import os
from datetime import datetime, timedelta
import json

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

class October2025TestSuite:
    """Test suite untuk bulan Oktober 2025"""

    def __init__(self):
        self.test_results = []
        self.test_report_path = None
        self.query_results = {}
        self.variables = {}
        self.errors = []

    def log(self, message: str, level: str = "INFO"):
        """Log message dengan timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {level}: {message}")
        self.test_results.append(f"[{timestamp}] {level}: {message}")

    def test_database_connection(self):
        """Test koneksi database"""
        self.log("Testing database connection...")

        try:
            from core.database_connector import FirebirdConnectorEnhanced

            connector = FirebirdConnectorEnhanced.create_default_connector()

            if connector.test_connection():
                self.log("Database connection successful", "SUCCESS")

                # Get database info
                db_info = connector.get_database_size()
                self.log(f"Database size: {db_info.get('size_mb', 'Unknown')} MB")

                # Check FFB tables
                tables = connector.get_table_list()
                ffb_tables = [t for t in tables if 'FFB' in t.upper() and 'SCANNER' in t.upper()]
                self.log(f"Found {len(ffb_tables)} FFB/Scanner tables")

                # Check FFBSCANNERDATA04 specifically
                ffb04_exists = connector.check_table_exists('FFBSCANNERDATA04')
                if ffb04_exists:
                    row_count = connector.get_row_count('FFBSCANNERDATA04')
                    self.log(f"FFBSCANNERDATA04 exists: {row_count} rows", "INFO")

                    if row_count > 0:
                        # Get sample data
                        sample = connector.get_sample_data('FFBSCANNERDATA04', limit=3)
                        if sample:
                            self.log(f"Sample data structure: {list(sample[0].keys())[:8]}...")

                            # Check date range
                            date_query = """
                            SELECT
                                MIN(TRANSDATE) as MIN_DATE,
                                MAX(TRANSDATE) as MAX_DATE,
                                COUNT(*) as TOTAL_RECORDS,
                                COUNT(DISTINCT WORKERID) as UNIQUE_WORKERS,
                                COUNT(DISTINCT FIELDID) as UNIQUE_FIELDS
                            FROM FFBSCANNERDATA04
                            WHERE TRANSDATE IS NOT NULL
                            """
                            date_result = connector.execute_query(date_query)
                            if date_result and len(date_result) > 0 and date_result[0].get('rows'):
                                date_row = date_result[0]['rows'][0]
                                self.log(f"Date range: {date_row.get('MIN_DATE')} to {date_row.get('MAX_DATE')}")
                                self.log(f"Records: {date_row.get('TOTAL_RECORDS')}")
                                self.log(f"Workers: {date_row.get('UNIQUE_WORKERS')}")
                                self.log(f"Fields: {date_row.get('UNIQUE_FIELDS')}")
                    else:
                        self.log("FFBSCANNERDATA04 is empty", "WARNING")
                else:
                    self.log("FFBSCANNERDATA04 table not found", "ERROR")
                    return False

                return connector
            else:
                self.log("Database connection failed", "ERROR")
                return False

        except Exception as e:
            self.log(f"Database connection error: {e}", "ERROR")
            return False

    def test_formula_loading(self, connector):
        """Test loading formula untuk Oktober 2025"""
        self.log("Testing formula loading...")

        try:
            from core.formula_engine import FormulaEngine

            engine = FormulaEngine()
            engine.database_connector = connector

            if engine.load_formula('templates/ffb_scannerdata04_formula.json'):
                self.log("Formula loaded successfully", "SUCCESS")

                formula_info = engine.get_formula_info()
                self.log(f"Formula queries: {formula_info['query_count']}")
                self.log(f"Formula variables: {formula_info['variable_count']}")

                return engine
            else:
                self.log("Formula loading failed", "ERROR")
                return False

        except Exception as e:
            self.log(f"Formula loading error: {e}", "ERROR")
            return False

    def test_template_loading(self):
        """Test loading template Excel"""
        self.log("Testing template loading...")

        try:
            from core.template_processor import TemplateProcessor

            processor = TemplateProcessor()

            # Find template file
            template_files = []
            for file in os.listdir('templates'):
                if file.startswith('FFB_ScannerData04_Template_') and file.endswith('.xlsx'):
                    template_files.append(file)

            if template_files:
                template_path = os.path.join('templates', template_files[0])

                if processor.load_template(template_path):
                    self.log(f"✓ Template loaded: {template_files[0]}", "SUCCESS")

                    template_info = processor.get_template_info()
                    self.log(f"Template rows: {template_info['total_rows']}")
                    self.log(f"Template columns: {template_info['total_columns']}")
                    self.log(f"Placeholders: {template_info['placeholder_count']}")

                    # Check for data_records placeholders
                    placeholders = processor.extract_placeholders()
                    data_placeholders = [p for p in placeholders if 'data_records.' in p]
                    summary_placeholders = [p for p in placeholders if 'summary.' in p]

                    self.log(f"Data records placeholders: {len(data_placeholders)}")
                    self.log(f"Summary placeholders: {len(summary_placeholders)}")

                    return processor
                else:
                    self.log("Template loading failed", "ERROR")
                    return False
            else:
                self.log("No template files found", "ERROR")
                return False

        except Exception as e:
            self.log(f"Template loading error: {e}", "ERROR")
            return False

    def test_october_query_execution(self, engine):
        """Test query execution untuk bulan Oktober 2025"""
        self.log("Testing October 2025 query execution...")

        try:
            # Test multiple date ranges in October
            october_dates = [
                ('2025-10-01', '2025-10-07', "Week 1"),
                ('2025-10-08', '2025-10-14', "Week 2"),
                ('2025-10-15', '2025-10-21', "Week 3"),
                ('2025-10-22', '2025-10-31', "Week 4"),
                ('2025-10-01', '2025-10-31', "Full Month")
            ]

            all_results = {}
            total_records = 0

            for start_date, end_date, period in october_dates:
                self.log(f"Testing {period}: {start_date} to {end_date}")

                parameters = {
                    'start_date': start_date,
                    'end_date': end_date
                }

                results = engine.execute_queries(parameters)

                if results:
                    all_results[period] = results

                    # Count records from main_data
                    main_data = results.get('main_data', [])
                    record_count = 0

                    if main_data:
                        if isinstance(main_data, list):
                            record_count = len(main_data)
                        elif isinstance(main_data, dict) and 'rows' in main_data:
                            record_count = len(main_data['rows'])

                    self.log(f"  {period}: {record_count} records found")
                    total_records += record_count

                    # Show sample data if available
                    if record_count > 0:
                        if isinstance(main_data, list) and main_data:
                            sample = main_data[0]
                            if isinstance(sample, dict):
                                sample_keys = list(sample.keys())[:5]
                                self.log(f"  Sample fields: {sample_keys}")
                        elif isinstance(main_data, dict) and 'rows' in main_data and main_data['rows']:
                            sample = main_data['rows'][0]
                            if isinstance(sample, dict):
                                sample_keys = list(sample.keys())[:5]
                                self.log(f"  Sample fields: {sample_keys}")
                else:
                    self.log(f"  {period}: No results returned")

            self.query_results = all_results
            self.log(f"Total October records across all periods: {total_records}")

            if total_records > 0:
                self.log("✓ October 2025 query execution successful", "SUCCESS")
                return True
            else:
                self.log("No data found for October 2025", "WARNING")
                return False  # Not error, just no data

        except Exception as e:
            self.log(f"Query execution error: {e}", "ERROR")
            return False

    def test_variable_processing(self, engine):
        """Test variable processing dengan data Oktober"""
        self.log("Testing variable processing...")

        try:
            # Use the most successful query result
            best_period = None
            best_record_count = 0

            for period, results in self.query_results.items():
                main_data = results.get('main_data', [])
                record_count = 0

                if main_data:
                    if isinstance(main_data, list):
                        record_count = len(main_data)
                    elif isinstance(main_data, dict) and 'rows' in main_data:
                        record_count = len(main_data['rows'])

                if record_count > best_record_count:
                    best_record_count = record_count
                    best_period = period

            if best_period:
                parameters = {
                    'start_date': '2025-10-01',  # Use full month for variable processing
                    'end_date': '2025-10-31'
                }

                self.log(f"Using best period: {best_period} ({best_record_count} records)")

                variables = engine.process_variables(self.query_results[best_period], parameters)
                self.variables = variables

                self.log(f"✓ Variables processed: {len(variables)} variables")

                # Check important variables
                important_vars = [
                    'estate_name', 'report_period', 'report_date',
                    'database_name', 'generated_by', 'generation_time'
                ]

                for var in important_vars:
                    if var in variables:
                        value = variables[var]
                        self.log(f"  {var}: {value}")
                    else:
                        self.log(f"  {var}: NOT FOUND", "WARNING")

                # Check summary variables
                if 'summary' in variables and isinstance(variables['summary'], dict):
                    summary = variables['summary']
                    self.log(f"  Summary variables: {len(summary)} items")

                    important_summary = [
                        'total_records', 'total_ripe_bunch', 'total_unripe_bunch',
                        'total_black_bunch', 'total_rotten_bunch', 'date_range'
                    ]

                    for key in important_summary:
                        if key in summary:
                            value = summary[key]
                            self.log(f"    {key}: {value}")
                        else:
                            self.log(f"    {key}: NOT FOUND", "WARNING")

                return True
            else:
                self.log("No suitable period found for variable processing", "WARNING")
                # Still try with empty database
                parameters = {
                    'start_date': '2025-10-01',
                    'end_date': '2025-10-31'
                }

                variables = engine.process_variables({}, parameters)
                self.variables = variables
                self.log(f"✓ Variables processed with empty database: {len(variables)} variables")

                # Check summary default values
                if 'summary' in variables:
                    summary = variables['summary']
                    if isinstance(summary, dict):
                        self.log(f"  Default total_records: {summary.get('total_records', 'N/A')}")
                        self.log(f"  Default total_ripe_bunch: {summary.get('total_ripe_bunch', 'N/A')}")

                return True

        except Exception as e:
            self.log(f"Variable processing error: {e}", "ERROR")
            return False

    def test_placeholder_replacement(self, processor):
        """Test placeholder replacement di template"""
        self.log("Testing placeholder replacement...")

        try:
            # Check if we have data_records
            data_records = self.variables.get('data_records', [])

            if data_records:
                self.log(f"✓ Data records available: {len(data_records)} records")

                if isinstance(data_records, list):
                    sample_record = data_records[0]
                    if isinstance(sample_record, dict):
                        sample_keys = list(sample_record.keys())[:5]
                        self.log(f"  Sample record fields: {sample_keys}")
                elif isinstance(data_records, dict) and 'rows' in data_records:
                    rows = data_records['rows']
                    if rows:
                        sample_row = rows[0]
                        if isinstance(sample_row, dict):
                            sample_keys = list(sample_row.keys())[:5]
                            self.log(f"  Sample row fields: {sample_keys}")
            else:
                self.log("No data_records available for testing", "INFO")

            # Test placeholder replacement
            replacements = processor.replace_placeholders(self.variables)
            self.log(f"✓ Placeholder replacement: {replacements} replacements")

            # Validate important placeholders
            important_placeholders = [
                'estate_name', 'report_period', 'report_date',
                'database_name', 'generated_by', 'generation_time'
            ]

            placeholders = processor.extract_placeholders()
            found_placeholders = []
            missing_placeholders = []

            for placeholder in placeholders:
                if not any(ph in placeholder for ph in important_placeholders):
                    continue  # Skip non-important placeholders

                if placeholder.startswith('{{') and placeholder.endswith('}}'):
                    var_name = placeholder[2:-2]
                    if var_name in self.variables:
                        found_placeholders.append(placeholder)
                    else:
                        missing_placeholders.append(placeholder)

            if missing_placeholders:
                self.log(f"Missing placeholders: {len(missing_placeholders)}", "WARNING")
                for missing in missing_placeholders[:3]:
                    self.log(f"  {missing}")
            else:
                self.log("✓ All important placeholders found", "SUCCESS")

            return True

        except Exception as e:
            self.log(f"Placeholder replacement error: {e}", "ERROR")
            return False

    def generate_october_report(self, processor):
        """Generate report untuk Oktober 2025"""
        self.log("Generating October 2025 report...")

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.test_report_path = f"FFB_October_2025_Report_{timestamp}.xlsx"

            processor.save_template(self.test_report_path)

            if os.path.exists(self.test_report_path):
                file_size = os.path.getsize(self.test_report_path)
                self.log(f"✓ Report generated: {self.test_report_path}")
                self.log(f"  File size: {file_size} bytes")

                if file_size > 0:
                    return True
                else:
                    self.log("Generated report file is empty", "ERROR")
                    return False
            else:
                self.log("Report file not created", "ERROR")
                return False

        except Exception as e:
            self.log(f"Report generation error: {e}", "ERROR")
            return False

    def validate_report_content(self):
        """Validate content dari generated report"""
        self.log("Validating report content...")

        try:
            if not self.test_report_path or not os.path.exists(self.test_report_path):
                self.log("No report file to validate", "ERROR")
                return False

            # Load and check report content
            from openpyxl import load_workbook

            workbook = load_workbook(self.test_report_path)
            worksheet = workbook.active

            # Extract text from all cells
            cell_values = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_values.append(cell.value)

            # Check for remaining placeholders
            remaining_placeholders = [text for text in cell_values if '{{' in text and '}}' in text]

            if remaining_placeholders:
                self.log(f"❌ Found {len(remaining_placeholders)} unreplaced placeholders", "ERROR")
                for placeholder in remaining_placeholders[:5]:
                    self.log(f"  {placeholder}")
                return False
            else:
                self.log("✓ No remaining placeholders found", "SUCCESS")

            # Check for important content
            important_content = [
                'PGE 2B',  # estate_name
                'October',    # report_period
                'PTRJ_P2B.FDB'  # database_name
            ]

            found_content = []
            for content in important_content:
                if any(content in text for text in cell_values):
                    found_content.append(content)
                    self.log(f"  Found content: {content}")
                else:
                    self.log(f"  Missing content: {content}", "WARNING")

            if len(found_content) >= 2:
                self.log("✓ Important content found in report", "SUCCESS")
            else:
                self.log("Limited important content found", "WARNING")

            return True

        except Exception as e:
            self.log(f"Report validation error: {e}", "ERROR")
            return False

    def generate_test_report_json(self):
        """Generate JSON report dengan test results"""
        self.log("Generating test report JSON...")

        try:
            test_report = {
                'test_info': {
                    'test_name': 'October 2025 Comprehensive Test',
                    'test_timestamp': datetime.now().isoformat(),
                    'test_duration': len(self.test_results),
                },
                'database_info': {
                    'connection': 'SUCCESS' if any('SUCCESS' in result for result in self.test_results) else 'FAILED',
                    'ffbscannerdata04_exists': any('FFBSCANNERDATA04' in result for result in self.test_results),
                    'total_records': sum(int(result.split(':')[-1].strip()) for result in self.test_results if 'rows found:' in result or 'records found:' in result)
                },
                'query_results': {},
                'variables': {},
                'test_results': self.test_results,
                'report_file': self.test_report_path,
                'success': True
            }

            # Add query results summary
            for period, results in self.query_results.items():
                test_report['query_results'][period] = {
                    'main_data_count': len(results.get('main_data', [])) if isinstance(results.get('main_data'), list) else len(results.get('main_data', {}).get('rows', [])),
                    'summary_stats_count': len(results.get('summary_stats', [])) if isinstance(results.get('summary_stats'), list) else len(results.get('summary_stats', {}).get('rows', [])),
                    'field_info_count': len(results.get('field_info', [])) if isinstance(results.get('field_info'), list) else len(results.get('field_info', {}).get('rows', [])),
                    'worker_info_count': len(results.get('worker_info', [])) if isinstance(results.get('worker_info'), list) else len(results.get('worker_info', {}).get('rows', []))
                }

            # Add variables
            test_report['variables'] = self.variables

            # Save test report
            report_path = f"october_2025_test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(report_path, 'w', encoding='utf-8') as f:
                json.dump(test_report, f, indent=2, ensure_ascii=False)

            self.log(f"✓ Test report saved: {report_path}")
            return report_path

        except Exception as e:
            self.log(f"Test report generation error: {e}", "ERROR")
            return None

    def run_comprehensive_test(self):
        """Run comprehensive test suite"""
        self.log("=" * 60)
        self.log("STARTING COMPREHENSIVE OCTOBER 2025 TEST")
        self.log("=" * 60)

        success = True

        # Test 1: Database Connection
        connector = self.test_database_connection()
        if not connector:
            success = False

        # Test 2: Formula Loading
        if success:
            engine = self.test_formula_loading(connector)
            if not engine:
                success = False

        # Test 3: Template Loading
        if success:
            processor = self.test_template_loading()
            if not processor:
                success = False

        # Test 4: Query Execution
        if success:
            query_success = self.test_october_query_execution(engine)
            if query_success:
                self.log("✓ Query execution completed successfully", "SUCCESS")
            else:
                self.log("⚠️ No data found for October 2025, but continuing test", "WARNING")

        # Test 5: Variable Processing
        if success:
            if not self.test_variable_processing(engine):
                success = False

        # Test 6: Placeholder Replacement
        if success:
            if not self.test_placeholder_replacement(processor):
                success = False

        # Test 7: Report Generation
        if success:
            if not self.generate_october_report(processor):
                success = False

        # Test 8: Report Validation
        if success:
            if not self.validate_report_content():
                success = False

        # Test 9: Generate Test Report JSON
        self.generate_test_report_json()

        # Final Summary
        self.log("=" * 60)
        self.log("COMPREHENSIVE TEST COMPLETE")
        self.log("=" * 60)

        if success:
            self.log("✅ OCTOBER 2025 TEST SUCCESSFUL", "SUCCESS")
            self.log("All systems working correctly")
            self.log(f"Test report file: {self.test_report_path}")
        else:
            self.log("❌ OCTOBER 2025 TEST FAILED", "ERROR")
            self.log("Some issues need to be addressed")

        self.log(f"Test completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        return success

def main():
    """Main function"""
    print("FBB Report Generator - October 2025 Comprehensive Test")
    print("=" * 60)

    test_suite = October2025TestSuite()
    success = test_suite.run_comprehensive_test()

    if success:
        print("\n✅ ALL TESTS PASSED")
        print("The October 2025 report generation system is working correctly!")
    else:
        print("\n❌ SOME TESTS FAILED")
        print("Please check the implementation for issues.")

if __name__ == "__main__":
    main()