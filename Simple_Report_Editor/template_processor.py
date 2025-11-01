#!/usr/bin/env python3
"""
Template Processor - Excel Template Processing Engine
====================================================

Menangani pemrosesan template Excel, placeholder mapping, dan rendering data.
Mendukung penggantian placeholder, section berulang, dan formatting.

Author: Claude AI Assistant
Version: 1.0.0
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy

class TemplateProcessor:
    """
    Processor untuk template Excel dengan placeholder dan dynamic sections.

    Features:
    - Placeholder substitution {{variable}}
    - Repeating sections dengan data expansion
    - Conditional formatting
    - Formula Excel preservation
    - Styling dan formatting preservation
    """

    def __init__(self):
        """Initialize Template Processor"""
        self.logger = logging.getLogger(__name__)
        self.workbook = None
        self.template_workbook = None
        self.placeholders = {}
        self.repeating_sections = {}
        self.processed_data = {}

    def load_template(self, template_path: str) -> bool:
        """
        Load template Excel

        Args:
            template_path: Path ke file template Excel

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            if not Path(template_path).exists():
                self.logger.error(f"Template file not found: {template_path}")
                return False

            self.template_workbook = load_workbook(template_path)
            self.logger.info(f"Template loaded: {template_path}")

            # Scan template untuk placeholders dan sections
            self.scan_template()

            return True

        except Exception as e:
            self.logger.error(f"Error loading template {template_path}: {e}")
            return False

    def scan_template(self) -> None:
        """
        Scan template untuk menemukan placeholders dan repeating sections.
        """
        if not self.template_workbook:
            return

        self.placeholders = {}
        self.repeating_sections = {}

        for sheet_name in self.template_workbook.sheetnames:
            sheet = self.template_workbook[sheet_name]

            # Scan untuk placeholders
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        placeholders = self.extract_placeholders(cell.value)
                        if placeholders:
                            if sheet_name not in self.placeholders:
                                self.placeholders[sheet_name] = {}
                            self.placeholders[sheet_name][cell.coordinate] = placeholders

        self.logger.info(f"Found {sum(len(sheet_places) for sheet_places in self.placeholders.values())} placeholders")

    def extract_placeholders(self, text: str) -> List[str]:
        """
        Extract placeholders dari text

        Args:
            text: Text yang mengandung placeholders

        Returns:
            List placeholder names
        """
        if not isinstance(text, str):
            return []

        # Pattern untuk {{variable_name}}
        pattern = r'\{\{([^}]+)\}\}'
        matches = re.findall(pattern, text)

        return matches

    def process_template(self, data: Dict, repeating_sections_config: Dict = None) -> bool:
        """
        Process template dengan data

        Args:
            data: Data untuk substitusi
            repeating_sections_config: Configuration untuk repeating sections

        Returns:
            True jika berhasil, False jika gagal
        """
        if not self.template_workbook:
            self.logger.error("No template loaded")
            return False

        try:
            # Create working copy
            self.workbook = copy.deepcopy(self.template_workbook)
            self.processed_data = data

            # Process repeating sections first
            if repeating_sections_config:
                self.process_repeating_sections(repeating_sections_config)

            # Process placeholders
            self.process_placeholders()

            self.logger.info("Template processing completed")
            return True

        except Exception as e:
            self.logger.error(f"Error processing template: {e}")
            return False

    def process_repeating_sections(self, sections_config: Dict) -> None:
        """
        Process repeating sections untuk data expansion

        Args:
            sections_config: Configuration untuk repeating sections
        """
        for sheet_name, sheet_sections in sections_config.items():
            if sheet_name not in self.workbook.sheetnames:
                self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                continue

            sheet = self.workbook[sheet_name]

            for section_name, section_config in sheet_sections.items():
                self.process_repeating_section(sheet, section_config)

    def process_repeating_section(self, sheet, section_config: Dict) -> None:
        """
        Process individual repeating section

        Args:
            sheet: Excel sheet object
            section_config: Section configuration
        """
        try:
            data_source = section_config.get('data_source', '')
            start_row = section_config.get('start_row', 1)
            end_row = section_config.get('end_row', None)
            columns = section_config.get('columns', {})
            auto_expand = section_config.get('auto_expand', True)
            preserve_formatting = section_config.get('preserve_formatting', True)

            # Get data untuk section
            section_data = self.get_section_data(data_source)
            if not section_data or len(section_data) == 0:
                self.logger.warning(f"No data found for section: {data_source}")
                return

            # Determine end row
            if end_row is None:
                # Find last row with content in template
                end_row = self.find_template_end_row(sheet, start_row, columns)

            # Process data rows
            self.expand_data_rows(sheet, section_data, start_row, end_row, columns, preserve_formatting)

            self.logger.info(f"Processed {len(section_data)} rows for section: {data_source}")

        except Exception as e:
            self.logger.error(f"Error processing repeating section: {e}")

    def get_section_data(self, data_source: str) -> List[Dict]:
        """
        Get data untuk repeating section

        Args:
            data_source: Data source key

        Returns:
            List data rows
        """
        if data_source in self.processed_data:
            data = self.processed_data[data_source]

            # Convert to list of dictionaries if needed
            if isinstance(data, dict):
                # Check if it's a summary stats with single row
                if 'total_records' in data:
                    return [data]
                else:
                    return list(data.values()) if data else []
            elif isinstance(data, list):
                return data
            else:
                return []

        return []

    def find_template_end_row(self, sheet, start_row: int, columns: Dict) -> int:
        """
        Find end row dari template section

        Args:
            sheet: Excel sheet object
            start_row: Start row number
            columns: Column mapping

        Returns:
            End row number
        """
        if not columns:
            return start_row

        # Find the maximum column that has data
        max_col = 0
        for col_letter in columns.keys():
            try:
                col_num = openpyxl.utils.column_index_from_string(col_letter)
                max_col = max(max_col, col_num)
            except:
                continue

        if max_col == 0:
            return start_row

        # Look for empty row after start_row
        for row in range(start_row + 1, sheet.max_row + 2):
            has_content = False
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != '':
                    has_content = True
                    break

            if not has_content:
                return row - 1

        return sheet.max_row

    def expand_data_rows(self, sheet, data: List[Dict], start_row: int, end_row: int,
                        columns: Dict, preserve_formatting: bool) -> None:
        """
        Expand data rows dari template

        Args:
            sheet: Excel sheet object
            data: Data rows
            start_row: Template start row
            end_row: Template end row
            columns: Column mapping
            preserve_formatting: Whether to preserve formatting
        """
        if not data or len(data) == 0:
            return

        # Get template row formatting
        template_row_styles = {}
        if preserve_formatting:
            template_row_styles = self.get_row_styles(sheet, start_row)

        # Calculate how many rows to insert
        template_rows = end_row - start_row + 1
        data_rows_needed = len(data)
        rows_to_insert = max(0, data_rows_needed - template_rows)

        if rows_to_insert > 0:
            # Insert new rows
            sheet.insert_rows(end_row + 1, rows_to_insert)

        # Process each data row
        for i, row_data in enumerate(data):
            target_row = start_row + i

            # Apply formatting if preserving
            if preserve_formatting and i == 0:  # Apply template formatting to first row
                self.apply_row_styles(sheet, target_row, template_row_styles)

            # Map data to columns
            for col_letter, field_name in columns.items():
                try:
                    col_num = openpyxl.utils.column_index_from_string(col_letter)
                    cell = sheet.cell(row=target_row, column=col_num)

                    # Get value from data
                    value = self.get_nested_value(row_data, field_name)
                    cell.value = value

                    # Apply formatting for subsequent rows
                    if preserve_formatting and i > 0:
                        self.copy_cell_formatting(sheet, start_row, col_num, target_row, col_num)

                except Exception as e:
                    self.logger.warning(f"Error mapping column {col_letter} to field {field_name}: {e}")

    def get_nested_value(self, data: Dict, field_path: str) -> Any:
        """
        Get nested value dari dictionary menggunakan dot notation

        Args:
            data: Data dictionary
            field_path: Field path dengan dot notation

        Returns:
            Value atau None jika tidak ditemukan
        """
        if not field_path:
            return None

        keys = field_path.split('.')
        value = data

        try:
            for key in keys:
                if isinstance(value, dict) and key in value:
                    value = value[key]
                elif isinstance(value, list) and key.isdigit():
                    index = int(key)
                    if 0 <= index < len(value):
                        value = value[index]
                    else:
                        return None
                else:
                    return None

            return value

        except (KeyError, IndexError, TypeError):
            return None

    def get_row_styles(self, sheet, row_num: int) -> Dict:
        """
        Extract formatting styles dari row

        Args:
            sheet: Excel sheet object
            row_num: Row number

        Returns:
            Dictionary styles per column
        """
        styles = {}

        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_num, column=col)
            styles[col] = {
                'font': copy.copy(cell.font),
                'fill': copy.copy(cell.fill),
                'border': copy.copy(cell.border),
                'alignment': copy.copy(cell.alignment),
                'number_format': cell.number_format
            }

        return styles

    def apply_row_styles(self, sheet, row_num: int, styles: Dict) -> None:
        """
        Apply formatting styles ke row

        Args:
            sheet: Excel sheet object
            row_num: Target row number
            styles: Styles dictionary
        """
        for col, style in styles.items():
            cell = sheet.cell(row=row_num, column=col)

            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'border' in style:
                cell.border = style['border']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'number_format' in style:
                cell.number_format = style['number_format']

    def copy_cell_formatting(self, sheet, source_row: int, source_col: int,
                           target_row: int, target_col: int) -> None:
        """
        Copy formatting dari source cell ke target cell

        Args:
            sheet: Excel sheet object
            source_row: Source row number
            source_col: Source column number
            target_row: Target row number
            target_col: Target column number
        """
        source_cell = sheet.cell(row=source_row, column=source_col)
        target_cell = sheet.cell(row=target_row, column=target_col)

        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format

    def process_placeholders(self) -> None:
        """
        Process semua placeholders di workbook
        """
        for sheet_name, sheet_placeholders in self.placeholders.items():
            if sheet_name not in self.workbook.sheetnames:
                continue

            sheet = self.workbook[sheet_name]

            for cell_coord, placeholders in sheet_placeholders.items():
                self.process_cell_placeholders(sheet, cell_coord, placeholders)

    def process_cell_placeholders(self, sheet, cell_coord: str, placeholders: List[str]) -> None:
        """
        Process placeholders di cell tertentu

        Args:
            sheet: Excel sheet object
            cell_coord: Cell coordinate (A1, B2, etc)
            placeholders: List placeholder names
        """
        try:
            cell = sheet[cell_coord]
            original_value = cell.value

            if not original_value or not isinstance(original_value, str):
                return

            # Replace each placeholder
            new_value = original_value
            for placeholder in placeholders:
                value = self.get_placeholder_value(placeholder)
                if value is not None:
                    new_value = new_value.replace(f'{{{{{placeholder}}}}}', str(value))

            cell.value = new_value

        except Exception as e:
            self.logger.error(f"Error processing placeholders in {cell_coord}: {e}")

    def get_placeholder_value(self, placeholder: str) -> Any:
        """
        Get value untuk placeholder

        Args:
            placeholder: Placeholder name

        Returns:
            Value atau None
        """
        # Handle complex placeholders dengan format
        if '|' in placeholder:
            name, format_spec = placeholder.split('|', 1)
            value = self.get_data_value(name)
            return self.format_value(value, format_spec)
        else:
            return self.get_data_value(placeholder)

    def get_data_value(self, key: str) -> Any:
        """
        Get value dari processed data

        Args:
            key: Data key

        Returns:
            Value atau None
        """
        if key in self.processed_data:
            return self.processed_data[key]

        # Handle nested keys dengan dot notation
        if '.' in key:
            return self.get_nested_value(self.processed_data, key)

        return None

    def format_value(self, value: Any, format_spec: str) -> str:
        """
        Format value sesuai specification

        Args:
            value: Value to format
            format_spec: Format specification

        Returns:
            Formatted string
        """
        if value is None:
            return ''

        try:
            if format_spec.startswith('format:'):
                # Date formatting
                format_string = format_spec[7:]  # Remove 'format:'
                if hasattr(value, 'strftime'):
                    return value.strftime(format_string)
                else:
                    # Try to parse as date
                    if isinstance(value, str):
                        try:
                            date_obj = datetime.strptime(value, '%Y-%m-%d')
                            return date_obj.strftime(format_string)
                        except ValueError:
                            return str(value)
                    else:
                        return str(value)

            elif format_spec.startswith('number:'):
                # Number formatting
                format_string = format_spec[7:]  # Remove 'number:'
                try:
                    if '.' in format_string:
                        decimal_places = int(format_string.split('.')[1])
                        return f"{float(value):.{decimal_places}f}"
                    else:
                        return f"{int(value)}"
                except (ValueError, TypeError):
                    return str(value)

            else:
                return str(value)

        except Exception as e:
            self.logger.warning(f"Error formatting value {value} with {format_spec}: {e}")
            return str(value)

    def save_workbook(self, output_path: str) -> bool:
        """
        Save processed workbook

        Args:
            output_path: Output file path

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            if not self.workbook:
                self.logger.error("No workbook to save")
                return False

            # Create output directory if needed
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # Save workbook
            self.workbook.save(output_path)
            self.logger.info(f"Workbook saved: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error saving workbook: {e}")
            return False

    def get_template_info(self) -> Dict:
        """
        Get informasi tentang template

        Returns:
            Dictionary template information
        """
        if not self.template_workbook:
            return {}

        info = {
            'sheet_names': self.template_workbook.sheetnames,
            'total_placeholders': sum(len(sheet_places) for sheet_places in self.placeholders.values()),
            'placeholders_per_sheet': {sheet: len(places) for sheet, places in self.placeholders.items()},
            'max_row': max(sheet.max_row for sheet in self.template_workbook.worksheets),
            'max_col': max(sheet.max_column for sheet in self.template_workbook.worksheets)
        }

        return info

    def validate_template(self) -> List[str]:
        """
        Validate template dan return list dari issues

        Returns:
            List validation issues
        """
        issues = []

        if not self.template_workbook:
            issues.append("No template loaded")
            return issues

        # Check for sheets
        if len(self.template_workbook.sheetnames) == 0:
            issues.append("No sheets found in template")

        # Check for placeholders
        total_placeholders = sum(len(sheet_places) for sheet_places in self.placeholders.values())
        if total_placeholders == 0:
            issues.append("No placeholders found in template")

        # Check for data sections in sheets
        for sheet_name in self.template_workbook.sheetnames:
            sheet = self.template_workbook[sheet_name]
            if sheet.max_row < 2:
                issues.append(f"Sheet '{sheet_name}' has insufficient rows for data")

        return issues

    def create_sample_template(self, output_path: str, template_config: Dict) -> bool:
        """
        Create sample template dari configuration

        Args:
            output_path: Output file path
            template_config: Template configuration

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            # Create new workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets based on config
            for sheet_name, sheet_config in template_config.get('sheets', {}).items():
                ws = wb.create_sheet(title=sheet_name)

                # Add headers
                headers = sheet_config.get('headers', [])
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # Add placeholders
                placeholders = sheet_config.get('placeholders', {})
                for cell_coord, placeholder_text in placeholders.items():
                    ws[cell_coord] = placeholder_text

                # Add formatting
                self.apply_sample_formatting(ws, sheet_config.get('formatting', {}))

            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Sample template created: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error creating sample template: {e}")
            return False

    def apply_sample_formatting(self, sheet, formatting_config: Dict) -> None:
        """
        Apply formatting ke sample template

        Args:
            sheet: Excel sheet object
            formatting_config: Formatting configuration
        """
        try:
            # Header formatting
            if 'header' in formatting_config:
                header_config = formatting_config['header']
                header_font = Font(
                    bold=header_config.get('bold', True),
                    size=header_config.get('size', 12)
                )
                header_fill = PatternFill(
                    start_color=header_config.get('bg_color', 'CCCCCC'),
                    end_color=header_config.get('bg_color', 'CCCCCC'),
                    fill_type='solid'
                )

                for cell in sheet[1]:  # First row
                    cell.font = header_font
                    cell.fill = header_fill

            # Column widths
            if 'column_widths' in formatting_config:
                for col_letter, width in formatting_config['column_widths'].items():
                    sheet.column_dimensions[col_letter].width = width

        except Exception as e:
            self.logger.warning(f"Error applying sample formatting: {e}")