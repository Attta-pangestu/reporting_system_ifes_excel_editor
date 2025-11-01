"""
Excel Template Generator for FFBSCANNERDATA04
Membuat template Excel dengan placeholder variables untuk laporan FFB
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class FFBTemplateGenerator:
    """Generator template Excel untuk laporan FFB"""

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def create_ffb_scannerdata04_template(self, output_path: str = None):
        """
        Membuat template Excel untuk FFBSCANNERDATA04

        Args:
            output_path: Path untuk menyimpan template file
        """
        # Create workbook and worksheet
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "FFB Scanner Data 04"

        # Define styles
        self._define_styles()

        # Create header section
        self._create_header_section()

        # Create data table headers
        self._create_table_headers()

        # Create summary section
        self._create_summary_section()

        # Set column widths
        self._set_column_widths()

        # Save template
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"FFB_ScannerData04_Template_{timestamp}.xlsx"

        self.workbook.save(output_path)
        return output_path

    def _define_styles(self):
        """Define reusable styles"""
        # Title style
        self.title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        self.title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        self.title_alignment = Alignment(horizontal='center', vertical='center')

        # Header style
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')

        # Label style
        self.label_font = Font(name='Arial', size=11, bold=True)
        self.label_alignment = Alignment(horizontal='right', vertical='center')

        # Value style
        self.value_font = Font(name='Arial', size=11)
        self.value_alignment = Alignment(horizontal='left', vertical='center')

        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.thin_border = thin_border

    def _create_header_section(self):
        """Membuat header section dengan informasi laporan"""
        # Main title
        self.worksheet.merge_cells('A1:V1')
        title_cell = self.worksheet['A1']
        title_cell.value = "LAPORAN FFB SCANNER DATA 04"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        title_cell.border = self.thin_border

        # Report information
        info_data = [
            ('Estate', '{{estate_name}}'),
            ('Periode', '{{report_period}}'),
            ('Tanggal Laporan', '{{report_date}}'),
            ('Database', '{{database_name}}'),
            ('User Generate', '{{generated_by}}'),
            ('Generate Time', '{{generation_time}}')
        ]

        row = 3
        for label, value in info_data:
            self.worksheet[f'A{row}'] = label
            self.worksheet[f'B{row}'] = value
            self.worksheet[f'A{row}'].font = self.label_font
            self.worksheet[f'A{row}'].alignment = self.label_alignment
            self.worksheet[f'B{row}'].font = self.value_font
            self.worksheet[f'B{row}'].alignment = self.value_alignment

            # Add light borders
            self.worksheet[f'A{row}'].border = self.thin_border
            self.worksheet[f'B{row}'].border = self.thin_border

            row += 1

    def _create_table_headers(self):
        """Membuat header tabel data"""
        # Table title
        start_row = 10
        self.worksheet.merge_cells(f'A{start_row}:V{start_row}')
        table_title = self.worksheet[f'A{start_row}']
        table_title.value = "DATA TRANSAKSI FFB"
        table_title.font = self.header_font
        table_title.fill = self.header_fill
        table_title.alignment = self.title_alignment
        table_title.border = self.thin_border

        # Column headers
        headers = [
            'ID', 'Scan User ID', 'OCID', 'Worker ID', 'Carrier ID', 'Field ID',
            'Task No', 'Ripe Bunch', 'Unripe Bunch', 'Black Bunch', 'Rotten Bunch',
            'Long Stalk Bunch', 'Rat Damage Bunch', 'Loose Fruit', 'Trans No',
            'Trans Date', 'Trans Time', 'Upload DateTime', 'Record Tag',
            'Trans Status', 'Trans Type', 'Last User', 'Last Updated',
            'Underripe Bunch', 'Overripe Bunch', 'Abnormal Bunch', 'Loose Fruit 2'
        ]

        header_row = start_row + 1
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        # Data template rows (showing placeholders for repeating data)
        data_start_row = header_row + 1
        for row in range(data_start_row, data_start_row + 5):  # Create 5 template rows
            for col in range(1, len(headers) + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.border = self.thin_border

                # Add placeholder variables for the first row
                if row == data_start_row:
                    col_letter = get_column_letter(col)
                    if col_letter == 'A':
                        cell.value = '{{data_records.0.ID}}'
                    elif col_letter == 'B':
                        cell.value = '{{data_records.0.SCANUSERID}}'
                    elif col_letter == 'C':
                        cell.value = '{{data_records.0.OCID}}'
                    elif col_letter == 'D':
                        cell.value = '{{data_records.0.WORKERID}}'
                    elif col_letter == 'E':
                        cell.value = '{{data_records.0.CARRIERID}}'
                    elif col_letter == 'F':
                        cell.value = '{{data_records.0.FIELDID}}'
                    elif col_letter == 'G':
                        cell.value = '{{data_records.0.TASKNO}}'
                    elif col_letter == 'H':
                        cell.value = '{{data_records.0.RIPEBCH}}'
                    elif col_letter == 'I':
                        cell.value = '{{data_records.0.UNRIPEBCH}}'
                    elif col_letter == 'J':
                        cell.value = '{{data_records.0.BLACKBCH}}'
                    elif col_letter == 'K':
                        cell.value = '{{data_records.0.ROTTENBCH}}'
                    elif col_letter == 'L':
                        cell.value = '{{data_records.0.LONGSTALKBCH}}'
                    elif col_letter == 'M':
                        cell.value = '{{data_records.0.RATDMGBCH}}'
                    elif col_letter == 'N':
                        cell.value = '{{data_records.0.LOOSEFRUIT}}'
                    elif col_letter == 'O':
                        cell.value = '{{data_records.0.TRANSNO}}'
                    elif col_letter == 'P':
                        cell.value = '{{data_records.0.TRANSDATE}}'
                    elif col_letter == 'Q':
                        cell.value = '{{data_records.0.TRANSTIME}}'
                    elif col_letter == 'R':
                        cell.value = '{{data_records.0.UPLOADDATETIME}}'
                    elif col_letter == 'S':
                        cell.value = '{{data_records.0.RECORDTAG}}'
                    elif col_letter == 'T':
                        cell.value = '{{data_records.0.TRANSSTATUS}}'
                    elif col_letter == 'U':
                        cell.value = '{{data_records.0.TRANSTYPE}}'
                    elif col_letter == 'V':
                        cell.value = '{{data_records.0.LASTUSER}}'

    def _create_summary_section(self):
        """Membuat summary section"""
        summary_start_row = 20

        # Summary title
        self.worksheet.merge_cells(f'A{summary_start_row}:V{summary_start_row}')
        summary_title = self.worksheet[f'A{summary_start_row}']
        summary_title.value = "RINGKASAN DATA"
        summary_title.font = self.header_font
        summary_title.fill = self.header_fill
        summary_title.alignment = self.title_alignment
        summary_title.border = self.thin_border

        # Summary data
        summary_data = [
            ('Total Records', '{{summary.total_records}}'),
            ('Total Ripe Bunch', '{{summary.total_ripe_bunch}}'),
            ('Total Unripe Bunch', '{{summary.total_unripe_bunch}}'),
            ('Total Black Bunch', '{{summary.total_black_bunch}}'),
            ('Total Rotten Bunch', '{{summary.total_rotten_bunch}}'),
            ('Total Long Stalk Bunch', '{{summary.total_long_stalk_bunch}}'),
            ('Total Rat Damage Bunch', '{{summary.total_rat_damage_bunch}}'),
            ('Total Loose Fruit', '{{summary.total_loose_fruit}}'),
            ('Total Loose Fruit 2', '{{summary.total_loose_fruit_2}}'),
            ('Total Underripe Bunch', '{{summary.total_underripe_bunch}}'),
            ('Total Overripe Bunch', '{{summary.total_overripe_bunch}}'),
            ('Total Abnormal Bunch', '{{summary.total_abnormal_bunch}}'),
            ('Date Range', '{{summary.date_range}}'),
            ('Unique Workers', '{{summary.unique_workers}}'),
            ('Unique Fields', '{{summary.unique_fields}}'),
        ]

        row = summary_start_row + 2
        for i, (label, value) in enumerate(summary_data):
            if i % 2 == 0:  # Start new row every 2 items
                col_a = 'A'
                col_b = 'B'
                col_c = 'E'
                col_d = 'F'

                self.worksheet[f'{col_a}{row}'] = label
                self.worksheet[f'{col_b}{row}'] = value
                self.worksheet[f'{col_a}{row}'].font = self.label_font
                self.worksheet[f'{col_a}{row}'].alignment = self.label_alignment
                self.worksheet[f'{col_b}{row}'].font = self.value_font
                self.worksheet[f'{col_b}{row}'].alignment = self.value_alignment

                self.worksheet[f'{col_a}{row}'].border = self.thin_border
                self.worksheet[f'{col_b}{row}'].border = self.thin_border
            else:  # Same row, different columns
                col_c = 'E'
                col_d = 'F'

                self.worksheet[f'{col_c}{row}'] = label
                self.worksheet[f'{col_d}{row}'] = value
                self.worksheet[f'{col_c}{row}'].font = self.label_font
                self.worksheet[f'{col_c}{row}'].alignment = self.label_alignment
                self.worksheet[f'{col_d}{row}'].font = self.value_font
                self.worksheet[f'{col_d}{row}'].alignment = self.value_alignment

                self.worksheet[f'{col_c}{row}'].border = self.thin_border
                self.worksheet[f'{col_d}{row}'].border = self.thin_border

                row += 1  # Move to next row after 2 items

        # Footer
        footer_row = summary_start_row + 12
        self.worksheet.merge_cells(f'A{footer_row}:V{footer_row}')
        footer_cell = self.worksheet[f'A{footer_row}']
        footer_cell.value = "Catatan: Laporan ini di-generate otomatis dari database FFBSCANNERDATA04"
        footer_cell.font = Font(name='Arial', size=10, italic=True)
        footer_cell.alignment = Alignment(horizontal='center', vertical='center')

    def _set_column_widths(self):
        """Set column widths untuk better formatting"""
        column_widths = {
            'A': 10,  # ID
            'B': 12,  # Scan User ID
            'C': 10,  # OCID
            'D': 12,  # Worker ID
            'E': 12,  # Carrier ID
            'F': 10,  # Field ID
            'G': 10,  # Task No
            'H': 12,  # Ripe Bunch
            'I': 12,  # Unripe Bunch
            'J': 12,  # Black Bunch
            'K': 12,  # Rotten Bunch
            'L': 15,  # Long Stalk Bunch
            'M': 15,  # Rat Damage Bunch
            'N': 12,  # Loose Fruit
            'O': 12,  # Trans No
            'P': 12,  # Trans Date
            'Q': 12,  # Trans Time
            'R': 15,  # Upload DateTime
            'S': 12,  # Record Tag
            'T': 12,  # Trans Status
            'U': 12,  # Trans Type
            'V': 12,  # Last User
        }

        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width

def main():
    """Main function untuk testing"""
    generator = FFBTemplateGenerator()

    print("Creating FFB SCANNERDATA04 Excel Template...")
    template_path = generator.create_ffb_scannerdata04_template()

    print(f"Template created successfully: {template_path}")
    print(f"Template file location: {os.path.abspath(template_path)}")

    return template_path

if __name__ == "__main__":
    main()