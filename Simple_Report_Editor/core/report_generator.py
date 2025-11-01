"""
FFB Scanner Data 04 Report Generator
Modul untuk mengenerate laporan FFB dari template Excel dan formula JSON
"""

import os
import json
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from firebird_connector_enhanced import FirebirdConnectorEnhanced

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ffb_scannerdata04_report.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class FFBReportGenerator:
    """Generator laporan FFB berbasis template dan formula"""

    def __init__(self, template_path: str = None, formula_path: str = None):
        """
        Inisialisasi FFB Report Generator

        Args:
            template_path: Path ke file Excel template
            formula_path: Path ke file JSON formula
        """
        self.template_path = template_path
        self.formula_path = formula_path
        self.formula_config = None
        self.workbook = None
        self.worksheet = None
        self.connector = None

        # Load formula configuration
        if formula_path:
            self.load_formula(formula_path)

    def load_formula(self, formula_path: str):
        """Load konfigurasi formula dari JSON file"""
        try:
            with open(formula_path, 'r', encoding='utf-8') as f:
                self.formula_config = json.load(f)
            logger.info(f"Formula loaded successfully from {formula_path}")
        except Exception as e:
            logger.error(f"Error loading formula: {e}")
            raise

    def connect_to_database(self) -> bool:
        """Connect ke database Firebird"""
        try:
            if not self.formula_config:
                raise ValueError("Formula config not loaded")

            db_config = self.formula_config.get('database_config', {})
            self.connector = FirebirdConnectorEnhanced(
                db_path=db_config.get('connection_string'),
                username=db_config.get('username', 'SYSDBA'),
                password=db_config.get('password', 'masterkey'),
                charset=db_config.get('charset', 'UTF8')
            )

            if self.connector.test_connection():
                logger.info("Database connection successful")
                return True
            else:
                logger.error("Database connection failed")
                return False

        except Exception as e:
            logger.error(f"Error connecting to database: {e}")
            return False

    def execute_queries(self, parameters: Dict[str, Any] = None) -> Dict[str, Any]:
        """Execute semua queries dari formula configuration"""
        if not self.connector or not self.formula_config:
            raise ValueError("Database connection or formula config not available")

        results = {}
        queries = self.formula_config.get('queries', {})

        for query_name, query_config in queries.items():
            try:
                logger.info(f"Executing query: {query_name}")
                sql = query_config['sql']
                return_format = query_config.get('return_format', 'dict')

                # Default parameters
                default_params = {
                    'start_date': None,
                    'end_date': None,
                    'field_id': None,
                    'worker_id': None,
                    'record_tag': None
                }

                # Merge with provided parameters
                if parameters:
                    default_params.update(parameters)

                result = self.connector.execute_query(sql, default_params, return_format)
                results[query_name] = result
                logger.info(f"Query {query_name} executed successfully, returned {len(result) if result else 0} rows")

            except Exception as e:
                logger.error(f"Error executing query {query_name}: {e}")
                results[query_name] = None

        return results

    def process_variables(self, query_results: Dict[str, Any], parameters: Dict[str, Any] = None) -> Dict[str, Any]:
        """Process variables dari query results dan formula configuration"""
        variables = {}
        variables_config = self.formula_config.get('variables', {})

        for var_name, var_config in variables_config.items():
            try:
                var_type = var_config.get('type')

                if var_type == 'constant':
                    variables[var_name] = var_config.get('value')

                elif var_type == 'direct':
                    source = var_config.get('source')
                    if source in query_results:
                        variables[var_name] = query_results[source]

                elif var_type == 'formatting':
                    source = var_config.get('source')
                    format_str = var_config.get('format', '')

                    if source == 'TODAY()':
                        value = date.today()
                    elif source == 'NOW()':
                        value = datetime.now()
                    else:
                        value = source

                    if format_str:
                        variables[var_name] = value.strftime(format_str.replace('MMMM', '%B').replace('dddd', '%A').replace('yyyy', '%Y').replace('MM', '%m').replace('dd', '%d').replace('HH', '%H').replace('mm', '%M').replace('ss', '%S'))
                    else:
                        variables[var_name] = str(value)

                elif var_type == 'calculation':
                    formula = var_config.get('formula', '')
                    if formula.startswith('IF('):
                        # Simple IF formula processing
                        variables[var_name] = self._process_if_formula(formula, query_results, parameters)
                    else:
                        variables[var_name] = formula

            except Exception as e:
                logger.warning(f"Error processing variable {var_name}: {e}")
                variables[var_name] = f"Error: {str(e)}"

        return variables

    def _process_if_formula(self, formula: str, query_results: Dict[str, Any], parameters: Dict[str, Any]) -> str:
        """Process simple IF formula"""
        try:
            # Simple IF(ISBLANK(start_date), "value1", "value2")
            if 'ISBLANK(start_date)' in formula:
                if parameters and parameters.get('start_date'):
                    # Return second value
                    parts = formula.split(',')
                    if len(parts) >= 3:
                        return parts[2].strip().rstrip(')').replace('"', '').replace("'", "")
                else:
                    # Return first value
                    parts = formula.split(',')
                    if len(parts) >= 2:
                        return parts[1].strip().replace('"', '').replace("'", "")
        except:
            pass
        return ""

    def load_template(self, template_path: str = None):
        """Load Excel template"""
        template_file = template_path or self.template_path
        if not template_file:
            raise ValueError("Template path not provided")

        try:
            self.workbook = load_workbook(template_file)
            self.worksheet = self.workbook.active
            logger.info(f"Template loaded from {template_file}")
        except Exception as e:
            logger.error(f"Error loading template: {e}")
            raise

    def replace_placeholders(self, variables: Dict[str, Any]):
        """Replace placeholder variables di worksheet"""
        if not self.worksheet:
            raise ValueError("Worksheet not loaded")

        replacements = 0

        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{{' in cell.value and '}}' in cell.value:
                    original_value = cell.value
                    new_value = original_value

                    # Replace all placeholders
                    for var_name, var_value in variables.items():
                        placeholder = f'{{{{{var_name}}}}}'
                        if placeholder in new_value:
                            if isinstance(var_value, (list, dict)):
                                # Handle complex data types
                                if isinstance(var_value, list) and var_name == 'data_records':
                                    # Skip data_records placeholder as it will be handled by repeating sections
                                    continue
                                else:
                                    new_value = new_value.replace(placeholder, str(var_value))
                            else:
                                new_value = new_value.replace(placeholder, str(var_value))

                    if new_value != original_value:
                        cell.value = new_value
                        replacements += 1

        logger.info(f"Replaced {replacements} placeholders")

    def process_repeating_sections(self, query_results: Dict[str, Any]):
        """Process repeating sections untuk data tabel"""
        repeating_sections = self.formula_config.get('repeating_sections', {})

        for sheet_name, sections in repeating_sections.items():
            for section_name, section_config in sections.items():
                try:
                    data_source = section_config.get('data_source')
                    if data_source not in query_results:
                        continue

                    data = query_results[data_source]
                    if not data or not isinstance(data, list) or len(data) == 0:
                        continue

                    start_row = section_config.get('start_row', 1)
                    columns = section_config.get('columns', {})

                    # Clear existing template rows
                    end_row = self.worksheet.max_row
                    if end_row > start_row:
                        self.worksheet.delete_rows(start_row + 1, end_row - start_row)

                    # Insert data rows
                    for row_idx, record in enumerate(data):
                        current_row = start_row + row_idx + 1

                        for col_letter, field_name in columns.items():
                            if field_name in record:
                                cell = self.worksheet[f'{col_letter}{current_row}']
                                cell.value = record[field_name]

                    logger.info(f"Processed repeating section {section_name} with {len(data)} records")

                except Exception as e:
                    logger.error(f"Error processing repeating section {section_name}: {e}")

    def generate_report(self,
                       template_path: str = None,
                       output_path: str = None,
                       parameters: Dict[str, Any] = None) -> str:
        """
        Generate laporan lengkap

        Args:
            template_path: Path ke template file
            output_path: Path untuk output file
            parameters: Parameters untuk query

        Returns:
            Path ke file output yang di-generate
        """
        try:
            # Load template
            self.load_template(template_path)

            # Connect to database
            if not self.connect_to_database():
                raise Exception("Failed to connect to database")

            # Execute queries
            query_results = self.execute_queries(parameters)

            # Process variables
            variables = self.process_variables(query_results, parameters)

            # Replace placeholders
            self.replace_placeholders(variables)

            # Process repeating sections
            self.process_repeating_sections(query_results)

            # Generate output filename
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"FFB_ScannerData04_Report_{timestamp}.xlsx"

            # Save report
            self.workbook.save(output_path)
            logger.info(f"Report generated successfully: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error generating report: {e}")
            raise

    def generate_pdf_report(self,
                           template_path: str = None,
                           output_path: str = None,
                           parameters: Dict[str, Any] = None) -> str:
        """
        Generate laporan dalam format PDF

        Args:
            template_path: Path ke template file
            output_path: Path untuk output PDF file
            parameters: Parameters untuk query

        Returns:
            Path ke file PDF yang di-generate
        """
        try:
            # Generate Excel report first
            excel_path = self.generate_report(template_path, None, parameters)

            # Convert to PDF
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"FFB_ScannerData04_Report_{timestamp}.pdf"

            # PDF conversion would require additional library like win32com or reportlab
            # For now, just return the Excel path
            logger.info(f"PDF generation not yet implemented. Excel report saved at: {excel_path}")
            return excel_path

        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            raise

def main():
    """Main function untuk testing"""
    try:
        # Initialize generator
        generator = FFBReportGenerator()

        # Find generated template and formula files
        template_file = None
        for file in os.listdir('.'):
            if file.startswith('FFB_ScannerData04_Template_') and file.endswith('.xlsx'):
                template_file = file
                break

        formula_file = 'ffb_scannerdata04_formula.json'

        if not template_file:
            print("Template file not found!")
            return

        if not os.path.exists(formula_file):
            print("Formula file not found!")
            return

        print(f"Using template: {template_file}")
        print(f"Using formula: {formula_file}")

        # Load configuration
        generator.load_formula(formula_file)

        # Generate sample report with default parameters
        parameters = {
            'start_date': None,  # Use default
            'end_date': None,    # Use default
            'field_id': None,
            'worker_id': None,
            'record_tag': None
        }

        print("Generating report...")
        output_path = generator.generate_report(
            template_path=template_file,
            parameters=parameters
        )

        print(f"Report generated successfully: {output_path}")
        print(f"File location: {os.path.abspath(output_path)}")

    except Exception as e:
        print(f"Error: {e}")
        logger.error(f"Main function error: {e}")

if __name__ == "__main__":
    main()