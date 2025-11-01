"""
Template Processor Module
Memproses file Excel template dengan placeholder variables
"""

import os
import re
import logging
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class TemplateProcessor:
    """Processor untuk file Excel template dengan placeholder variables"""

    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.template_path = None

    def load_template(self, template_path: str) -> bool:
        """
        Load template Excel file

        Args:
            template_path: Path ke file template

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            if not os.path.exists(template_path):
                logger.error(f"Template file not found: {template_path}")
                return False

            self.workbook = load_workbook(template_path)
            self.worksheet = self.workbook.active
            self.template_path = template_path

            logger.info(f"Template loaded: {template_path}")
            return True

        except Exception as e:
            logger.error(f"Error loading template: {e}")
            return False

    def extract_placeholders(self) -> List[str]:
        """
        Extract semua placeholder variables dari template

        Returns:
            List nama placeholder variables
        """
        if not self.worksheet:
            logger.error("No template loaded")
            return []

        placeholders = set()
        placeholder_pattern = r'\{\{([^}]+)\}\}'

        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    matches = re.findall(placeholder_pattern, cell.value)
                    placeholders.update(matches)

        return list(placeholders)

    def validate_placeholders(self, variables: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """
        Validasi apakah semua placeholder memiliki nilai

        Args:
            variables: Dictionary variables untuk replacement

        Returns:
            Tuple (is_valid, missing_placeholders)
        """
        placeholders = self.extract_placeholders()
        missing = []

        for placeholder in placeholders:
            # Handle nested placeholders like data_records.0.FIELD
            if '.' in placeholder:
                parts = placeholder.split('.')
                if parts[0] not in variables:
                    missing.append(placeholder)
            else:
                if placeholder not in variables:
                    missing.append(placeholder)

        is_valid = len(missing) == 0
        return is_valid, missing

    def replace_placeholders(self, variables: Dict[str, Any]) -> int:
        """
        Replace placeholder variables dengan nilai

        Args:
            variables: Dictionary variables untuk replacement

        Returns:
            Jumlah replacement yang dilakukan
        """
        if not self.worksheet:
            logger.error("No template loaded")
            return 0

        replacements = 0

        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                    original_value = cell.value
                    new_value = original_value

                    # Handle data_records placeholders like {{data_records.0.FIELD_NAME}}
                    new_value = self._process_data_records_placeholders(new_value, variables)

                    # Replace all other placeholders
                    for var_name, var_value in variables.items():
                        if var_name == 'data_records':
                            continue  # Skip data_records, handled separately

                        placeholder = f'{{{{{var_name}}}}}'
                        if placeholder in new_value:
                            # Handle different data types
                            if isinstance(var_value, (list, dict)):
                                new_value = new_value.replace(placeholder, str(var_value))
                            else:
                                new_value = new_value.replace(placeholder, str(var_value))

                    if new_value != original_value:
                        cell.value = new_value
                        replacements += 1

        logger.info(f"Replaced {replacements} placeholders")
        return replacements

    def _process_data_records_placeholders(self, text: str, variables: Dict[str, Any]) -> str:
        """Process data_records placeholders like {{data_records.0.FIELD_NAME}}"""
        import re

        # Pattern to match data_records placeholders
        pattern = r'\{\{data_records\.(\d+)\.([^}]+)\}\}'
        matches = re.findall(pattern, text)

        if not matches:
            return text

        data_records = variables.get('data_records', [])
        if not data_records:
            return text

        for match in matches:
            row_index, field_name = int(match[0]), match[1]
            placeholder = f'{{{{data_records.{row_index}.{field_name}}}}}'

            # Get the value from data_records
            value = ""
            if row_index < len(data_records):
                record = data_records[row_index]
                if isinstance(record, dict):
                    value = record.get(field_name, "")
                elif hasattr(record, field_name):
                    value = getattr(record, field_name, "")

            # Replace the placeholder
            text = text.replace(placeholder, str(value))

        return text

    def find_repeating_sections(self) -> Dict[str, Dict]:
        """
        Cari repeating sections berdasarkan placeholder pattern

        Returns:
            Dictionary dengan informasi repeating sections
        """
        if not self.worksheet:
            logger.error("No template loaded")
            return {}

        sections = {}
        current_section = None
        section_data = {}

        for row_idx, row in enumerate(self.worksheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    # Look for data_records placeholders
                    if 'data_records.' in cell.value:
                        # Extract field name
                        match = re.search(r'data_records\.\d+\.([^}]+)', cell.value)
                        if match:
                            field_name = match.group(1)
                            col_letter = get_column_letter(col_idx)

                            if current_section is None:
                                current_section = "default_section"
                                section_data = {
                                    'start_row': row_idx,
                                    'columns': {},
                                    'data_field': 'data_records'
                                }

                            section_data['columns'][col_letter] = field_name

        if current_section:
            sections[current_section] = section_data

        return sections

    def process_repeating_sections(self, data: List[Dict], section_config: Dict = None):
        """
        Process repeating sections untuk data tabel

        Args:
            data: List data records
            section_config: Konfigurasi section (optional)
        """
        if not self.worksheet or not data:
            return

        # Find or use provided section config
        if section_config is None:
            sections = self.find_repeating_sections()
            if not sections:
                logger.warning("No repeating sections found")
                return
            section_config = list(sections.values())[0]

        start_row = section_config.get('start_row', 1)
        columns = section_config.get('columns', {})

        if not columns:
            logger.warning("No columns defined for repeating section")
            return

        # Clear existing template rows (keep header, remove template data rows)
        end_row = self.worksheet.max_row
        if end_row > start_row:
            self.worksheet.delete_rows(start_row + 1, end_row - start_row)

        # Insert data rows
        for row_idx, record in enumerate(data):
            current_row = start_row + row_idx + 1

            for col_letter, field_name in columns.items():
                if field_name in record:
                    cell = self.worksheet[f'{col_letter}{current_row}']
                    cell.value = record[field_name]

                    # Copy formatting from the row above (template row)
                    if row_idx == 0:
                        template_cell = self.worksheet[f'{col_letter}{start_row}']
                        if template_cell.has_style:
                            cell.font = template_cell.font
                            cell.border = template_cell.border
                            cell.fill = template_cell.fill
                            cell.number_format = template_cell.number_format
                            cell.alignment = template_cell.alignment

        logger.info(f"Processed {len(data)} records in repeating section")

    def get_template_info(self) -> Dict[str, Any]:
        """
        Get informasi tentang template

        Returns:
            Dictionary dengan informasi template
        """
        if not self.worksheet:
            return {}

        placeholders = self.extract_placeholders()
        repeating_sections = self.find_repeating_sections()

        return {
            'template_path': self.template_path,
            'worksheet_name': self.worksheet.title,
            'total_rows': self.worksheet.max_row,
            'total_columns': self.worksheet.max_column,
            'placeholders': placeholders,
            'placeholder_count': len(placeholders),
            'repeating_sections': repeating_sections,
            'repeating_section_count': len(repeating_sections)
        }

    def save_template(self, output_path: str):
        """
        Save template ke file

        Args:
            output_path: Path untuk output file
        """
        if not self.workbook:
            raise ValueError("No workbook to save")

        try:
            self.workbook.save(output_path)
            logger.info(f"Template saved: {output_path}")
        except Exception as e:
            logger.error(f"Error saving template: {e}")
            raise

    def create_copy(self) -> 'TemplateProcessor':
        """
        Create copy dari template processor

        Returns:
            TemplateProcessor instance baru dengan copy workbook
        """
        if not self.workbook:
            raise ValueError("No template loaded")

        new_processor = TemplateProcessor()
        new_processor.workbook = load_workbook(self.template_path)
        new_processor.worksheet = new_processor.workbook.active
        new_processor.template_path = self.template_path

        return new_processor

    def __del__(self):
        """Cleanup"""
        if hasattr(self, 'workbook') and self.workbook:
            self.workbook.close()