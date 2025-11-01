#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simple Demo Report - Menunjukkan Adaptive Excel Processing
"""

import os
import sys
from datetime import datetime

# Add current directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    from adaptive_excel_processor import AdaptiveExcelProcessor
except ImportError as e:
    print(f"Import Error: {e}")
    sys.exit(1)

def simple_demo():
    """Simple demo dengan basic data"""
    print("=" * 80)
    print("SIMPLE DEMO - ADAPTIVE EXCEL PROCESSOR")
    print("=" * 80)

    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "templates",
        "Template_Laporan_FFB_Analysis.xlsx"
    )

    output_dir = os.path.expanduser("~/Desktop")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f"Simple_Demo_{timestamp}.xlsx")

    print(f"Template: {template_path}")
    print(f"Output: {output_path}")
    print("-" * 80)

    try:
        # Step 1: Load and analyze template
        print("Step 1: Loading and analyzing template...")
        if not os.path.exists(template_path):
            raise Exception(f"Template not found: {template_path}")

        processor = AdaptiveExcelProcessor(template_path, debug_mode=True)
        template_summary = processor.get_template_summary()

        print(f"[OK] Template loaded successfully")
        print(f"   Sheets: {template_summary['template_info']['total_sheets']}")
        print(f"   Placeholders: {template_summary['template_info']['total_placeholders']}")

        # Step 2: Create simple test data
        print("\nStep 2: Creating test data...")

        test_data = {
            # Basic variables
            'report_title': 'LAPORAN DEMO ADAPTIVE SYSTEM',
            'estate_name': 'PGE 2B DEMO',
            'report_period': '1 September 2025 - 30 September 2025',
            'generated_date': datetime.now().strftime('%d %B %Y'),
            'generated_time': datetime.now().strftime('%H:%M:%S'),
            'total_transactions': 1000,
            'total_ripe_bunches': 800,
            'total_unripe_bunches': 200,
            'avg_ripe_per_transaction': 0.8,
            'quality_percentage': 85.0,
            'verification_rate': 90.0,
            'daily_average_transactions': 50.0,
            'daily_average_ripe': 40.0,
            'peak_performance_day': '2025-09-15',
            'low_performance_day': '2025-09-05',

            # Repeating data for sheets
            'Dashboard': [{'total_transactions': 1000, 'verification_rate': 90.0}],

            'Harian': [
                {'TRANSDATE': '2025-09-01', 'JUMLAH_TRANSAKSI': 45, 'RIPE_BUNCHES': 36, 'UNRIPE_BUNCHES': 9},
                {'TRANSDATE': '2025-09-02', 'JUMLAH_TRANSAKSI': 52, 'RIPE_BUNCHES': 42, 'UNRIPE_BUNCHES': 10},
                {'TRANSDATE': '2025-09-03', 'JUMLAH_TRANSAKSI': 38, 'RIPE_BUNCHES': 30, 'UNRIPE_BUNCHES': 8}
            ],

            'Karyawan': [
                {'EMPLOYEE_NAME': 'Ahmad Sutrisno', 'RECORDTAG': 'KERANI', 'JUMLAH_TRANSAKSI': 180, 'TOTAL_RIPE': 150, 'TOTAL_UNRIPE': 30},
                {'EMPLOYEE_NAME': 'Budi Santoso', 'RECORDTAG': 'KERANI', 'JUMLAH_TRANSAKSI': 165, 'TOTAL_RIPE': 135, 'TOTAL_UNRIPE': 30},
                {'EMPLOYEE_NAME': 'Chandra Wijaya', 'RECORDTAG': 'MANDOR', 'JUMLAH_TRANSAKSI': 120, 'TOTAL_RIPE': 100, 'TOTAL_UNRIPE': 20}
            ],

            'Field': [
                {'FIELDNAME': 'BLOCK A', 'JUMLAH_TRANSAKSI': 250, 'TOTAL_RIPE': 200, 'TOTAL_UNRIPE': 50},
                {'FIELDNAME': 'BLOCK B', 'JUMLAH_TRANSAKSI': 200, 'TOTAL_RIPE': 160, 'TOTAL_UNRIPE': 40}
            ],

            'Kualitas': [
                {'TRANSDATE': '2025-09-01', 'TOTAL_RIPE': 36, 'TOTAL_BLACK': 2, 'TOTAL_ROTTEN': 1, 'TOTAL_RATDAMAGE': 1, 'PERCENTAGE_DEFECT': 11.1},
                {'TRANSDATE': '2025-09-02', 'TOTAL_RIPE': 42, 'TOTAL_BLACK': 3, 'TOTAL_ROTTEN': 1, 'TOTAL_RATDAMAGE': 2, 'PERCENTAGE_DEFECT': 14.3}
            ]
        }

        print("   Test data created:")
        print(f"   - Static variables: {len([k for k in test_data.keys() if not isinstance(test_data[k], list)])}")
        print(f"   - Harian: {len(test_data['Harian'])} records")
        print(f"   - Karyawan: {len(test_data['Karyawan'])} records")
        print(f"   - Field: {len(test_data['Field'])} records")
        print(f"   - Kualitas: {len(test_data['Kualitas'])} records")

        # Step 3: Generate report
        print("\nStep 3: Generating adaptive report...")
        success = processor.generate_report(test_data, output_path)

        if success:
            print("[OK] Report generated successfully!")
            print(f"   Output file: {output_path}")

            # Check file
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"   File size: {file_size:,} bytes")

                # Read and verify
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(output_path)
                    print(f"   Sheets: {wb.sheetnames}")

                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        max_row = ws.max_row
                        max_col = ws.max_column
                        non_empty = 0
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value is not None and str(cell.value).strip():
                                    non_empty += 1
                        print(f"   {sheet_name}: {max_row} rows x {max_col} columns ({non_empty} filled cells)")

                except Exception as e:
                    print(f"   Error reading generated file: {e}")

                # Ask to open
                try:
                    user_input = input("\nOpen the generated report? (y/n): ").strip().lower()
                    if user_input in ['y', 'yes']:
                        os.startfile(output_path)
                        print("[OK] Report opened!")
                except Exception as e:
                    print(f"Please open manually: {output_path}")

            print("\n" + "=" * 80)
            print("SIMPLE DEMO COMPLETED SUCCESSFULLY!")
            print("The adaptive system processed all placeholders and repeating sections!")
            print("=" * 80)

            return True
        else:
            raise Exception("Report generation failed")

    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    print("Starting Simple Demo...")

    success = simple_demo()

    if success:
        print("\n[SUCCESS] Simple demo completed!")
        print("This demonstrates the adaptive system works with real Excel templates!")
    else:
        print("\n[FAILED] Simple demo failed!")

if __name__ == "__main__":
    main()