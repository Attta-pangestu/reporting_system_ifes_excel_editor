#!/usr/bin/env python3
"""
Generate new report and immediately check its content
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from excel_report_generator import ExcelReportGenerator
from firebird_connector_enhanced import FirebirdConnectorEnhanced

def main():
    print('=== GENERATING NEW REPORT TO CHECK CONTENT ===')

    try:
        # Initialize (correct signature)
        report_generator = ExcelReportGenerator(
            template_path='templates/Template_Laporan_FFB_Analysis.xlsx',
            formula_path='laporan_ffb_analysis_formula.json'
        )

        # Generate report
        success, output_file = report_generator.generate_report(
            '2020-10-01',
            '2020-10-10',
            ['PGE 2B'],
            '.'
        )

        print(f'Success: {success}')
        print(f'Output: {output_file}')

        if success and output_file and os.path.exists(output_file[0]):
            # Read and display content
            try:
                from openpyxl import load_workbook

                print(f'File size: {os.path.getsize(output_file[0])} bytes')

                wb = load_workbook(output_file[0])
                print(f'Sheets: {wb.sheetnames}')

                # Check Dashboard
                if 'Dashboard' in wb.sheetnames:
                    ws = wb['Dashboard']
                    print('\nDashboard content (first 15 rows):')
                    for row in range(1, 16):
                        row_data = []
                        for col in range(1, 4):
                            val = ws.cell(row=row, column=col).value
                            if val is not None:
                                row_data.append(str(val))
                        if row_data:
                            print(f'  Row {row}: {" | ".join(row_data)}')

                # Check Harian
                if 'Harian' in wb.sheetnames:
                    ws = wb['Harian']
                    print('\nHarian content (first 10 rows):')
                    for row in range(1, 11):
                        row_data = []
                        for col in range(1, 6):
                            val = ws.cell(row=row, column=col).value
                            if val is not None:
                                row_data.append(str(val))
                        if row_data:
                            print(f'  Row {row}: {" | ".join(row_data)}')

            except Exception as e:
                print(f'Error reading Excel: {e}')
                import traceback
                traceback.print_exc()
        else:
            print('No valid output file generated')
            return False

        return True

    except Exception as e:
        print(f'Error: {e}')
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)