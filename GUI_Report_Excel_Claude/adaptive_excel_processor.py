#!/usr/bin/env python3
"""
Adaptive Excel Processor - Sistem yang menyesuaikan dengan template Excel secara dinamis
Bisa meng-handle perubahan format template tanpa perlu mengubah kode
"""

import os
import json
import logging
import re
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple, Set
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class AdaptiveExcelProcessor:
    """
    Processor yang adaptif terhadap perubahan template Excel
    """

    def __init__(self, template_path: str, debug_mode: bool = True):
        """
        Inisialisasi adaptive processor

        Args:
            template_path: Path ke template Excel
            debug_mode: Enable debug logging
        """
        self.template_path = template_path
        self.debug_mode = debug_mode
        self.logger = logging.getLogger(__name__)

        if debug_mode:
            logging.basicConfig(
                level=logging.DEBUG,
                format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )

        # Template properties
        self.workbook = None
        self.template_info = {}
        self.placeholder_map = {}  # Map placeholder -> (sheet, cell, original_value)
        self.data_sections = {}   # Dynamic data sections
        self.format_presets = {}  # Preserved formatting

        # Load and analyze template
        self._load_template()
        self._analyze_template()

    def _load_template(self):
        """Load template Excel file"""
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")

            self.workbook = load_workbook(self.template_path, data_only=False)
            self.logger.info(f"Template loaded: {self.template_path}")
            self.logger.info(f"  Sheets: {self.workbook.sheetnames}")

        except Exception as e:
            self.logger.error(f"Error loading template: {e}")
            raise

    def _analyze_template(self):
        """Analisis struktur template untuk mengidentifikasi pattern"""
        self.logger.info("Analyzing template structure...")

        self.template_info = {
            'total_sheets': len(self.workbook.sheetnames),
            'sheet_names': self.workbook.sheetnames,
            'total_placeholders': 0,
            'data_sections': {},
            'format_analysis': {}
        }

        # Analyze each sheet
        for sheet_name in self.workbook.sheetnames:
            sheet_info = self._analyze_sheet(sheet_name)
            self.template_info['data_sections'][sheet_name] = sheet_info
            self.template_info['total_placeholders'] += sheet_info['placeholder_count']

        self.logger.info(f"Template analysis complete:")
        self.logger.info(f"  Total sheets: {self.template_info['total_sheets']}")
        self.logger.info(f"  Total placeholders: {self.template_info['total_placeholders']}")

    def _analyze_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Analisis individual sheet"""
        sheet = self.workbook[sheet_name]
        sheet_info = {
            'sheet_name': sheet_name,
            'used_range': self._get_used_range(sheet),
            'placeholders': [],
            'placeholder_count': 0,
            'data_patterns': [],
            'formatting': {},
            'tables': []
        }

        # Find all placeholders
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    placeholders = self._extract_placeholders(cell.value)
                    if placeholders:
                        for placeholder in placeholders:
                            placeholder_info = {
                                'placeholder': placeholder,
                                'cell': cell.coordinate,
                                'sheet': sheet_name,
                                'original_value': cell.value,
                                'row': cell.row,
                                'column': cell.column_letter,
                                'context': self._get_cell_context(sheet, cell)
                            }
                            sheet_info['placeholders'].append(placeholder_info)
                            self.placeholder_map[placeholder] = (sheet_name, cell.coordinate, cell.value)

        sheet_info['placeholder_count'] = len(sheet_info['placeholders'])

        # Detect data patterns (repeating sections)
        sheet_info['data_patterns'] = self._detect_data_patterns(sheet, sheet_info['placeholders'])

        # Detect tables
        sheet_info['tables'] = self._detect_tables(sheet)

        # Store formatting info
        sheet_info['formatting'] = self._analyze_formatting(sheet)

        return sheet_info

    def _extract_placeholders(self, text: str) -> List[str]:
        """Extract all placeholders from text"""
        # Support multiple placeholder formats
        patterns = [
            r'\{\{([^}]+)\}\}',      # {{variable}}
            r'\{\$([^}]+)\$\}',      # {$variable$}
            r'\{([^}]+)\}',          # {variable}
            r'\[([^\]]+)\]',         # [variable]
        ]

        all_placeholders = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            all_placeholders.extend(matches)

        return list(set(all_placeholders))  # Remove duplicates

    def _get_cell_context(self, sheet, cell, context_size: int = 2) -> Dict[str, Any]:
        """Get context around a cell to understand its purpose"""
        context = {
            'header_cells': [],
            'left_cells': [],
            'above_cells': [],
            'data_pattern': None
        }

        # Get cells to the left (potential headers)
        for col_offset in range(1, context_size + 1):
            left_col = cell.column - col_offset
            if left_col >= 1:
                left_cell = sheet.cell(row=cell.row, column=left_col)
                if left_cell.value:
                    context['left_cells'].append({
                        'cell': left_cell.coordinate,
                        'value': str(left_cell.value),
                        'offset': col_offset
                    })

        # Get cells above (potential headers)
        for row_offset in range(1, context_size + 1):
            above_row = cell.row - row_offset
            if above_row >= 1:
                above_cell = sheet.cell(row=above_row, column=cell.column)
                if above_cell.value:
                    context['above_cells'].append({
                        'cell': above_cell.coordinate,
                        'value': str(above_cell.value),
                        'offset': row_offset
                    })

        # Detect data pattern
        context['data_pattern'] = self._detect_cell_pattern(sheet, cell)

        return context

    def _detect_cell_pattern(self, sheet, cell) -> Optional[str]:
        """Detect pattern type based on cell context"""
        # Check if this looks like a header row
        row_values = []
        for col in sheet[cell.row]:
            if col.value:
                row_values.append(str(col.value).strip())

        # Check if row contains placeholder-like patterns
        placeholder_count = sum(1 for val in row_values if self._extract_placeholders(val))

        if placeholder_count > 0:
            if placeholder_count == len(row_values):
                return 'data_template_row'
            elif placeholder_count > len(row_values) / 2:
                return 'mixed_data_row'
            else:
                return 'partial_data_row'

        # Check if this looks like a summary cell
        if 'total' in str(cell.value).lower() or 'jumlah' in str(cell.value).lower():
            return 'summary_cell'

        # Check if this looks like a title/header
        if cell.row <= 3 and any(keyword in str(cell.value).lower() for keyword in ['laporan', 'report', 'judul', 'title']):
            return 'title_cell'

        return 'unknown'

    def _detect_data_patterns(self, sheet, placeholders: List[Dict]) -> List[Dict]:
        """Detect repeating data patterns"""
        patterns = []

        # Group placeholders by row
        row_groups = {}
        for ph in placeholders:
            row = ph['row']
            if row not in row_groups:
                row_groups[row] = []
            row_groups[row].append(ph)

        # Look for template rows (rows with multiple placeholders)
        for row_num, row_placeholders in row_groups.items():
            if len(row_placeholders) >= 3:  # Minimum 3 placeholders to consider as template row
                pattern = {
                    'type': 'template_row',
                    'row': row_num,
                    'placeholders': row_placeholders,
                    'start_col': min(ph['column'] for ph in row_placeholders),
                    'end_col': max(ph['column'] for ph in row_placeholders),
                    'col_count': len(set(ph['column'] for ph in row_placeholders))
                }
                patterns.append(pattern)

        return patterns

    def _detect_tables(self, sheet) -> List[Dict]:
        """Detect table structures in the sheet"""
        tables = []

        # Look for header rows followed by data rows
        for row_num in range(1, sheet.max_row - 1):
            # Check if current row looks like a header
            current_row = sheet[row_num]
            header_values = [str(cell.value) if cell.value else '' for cell in current_row]

            # Check if next row has placeholders (template row)
            next_row = sheet[row_num + 1]
            next_values = [str(cell.value) if cell.value else '' for cell in next_row]

            has_placeholders = any(self._extract_placeholders(val) for val in next_values)
            has_headers = any(val.strip() for val in header_values)

            if has_placeholders and has_headers:
                table = {
                    'type': 'dynamic_table',
                    'header_row': row_num,
                    'template_row': row_num + 1,
                    'headers': header_values,
                    'template_values': next_values,
                    'start_col': 1,
                    'end_col': len(header_values),
                    'col_count': len(header_values)
                }
                tables.append(table)

        return tables

    def _analyze_formatting(self, sheet) -> Dict[str, Any]:
        """Analyze formatting patterns in the sheet"""
        formatting = {
            'header_format': {},
            'data_format': {},
            'summary_format': {}
        }

        # Sample formatting from different cell types
        for row_num in range(1, min(10, sheet.max_row)):
            for col_num in range(1, min(10, sheet.max_column)):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value:
                    cell_format = {
                        'font': {
                            'name': cell.font.name,
                            'size': cell.font.size,
                            'bold': cell.font.bold,
                            'italic': cell.font.italic,
                            'color': str(cell.font.color) if cell.font.color else None
                        },
                        'fill': {
                            'pattern_type': cell.fill.patternType,
                            'start_color': str(cell.fill.start_color) if cell.fill.start_color else None,
                            'end_color': str(cell.fill.end_color) if cell.fill.end_color else None
                        },
                        'alignment': {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical,
                            'wrap_text': cell.alignment.wrap_text
                        },
                        'border': {
                            'left': str(cell.border.left.style) if cell.border.left else None,
                            'right': str(cell.border.right.style) if cell.border.right else None,
                            'top': str(cell.border.top.style) if cell.border.top else None,
                            'bottom': str(cell.border.bottom.style) if cell.border.bottom else None
                        }
                    }

                    # Categorize formatting
                    if row_num == 1 or any(keyword in str(cell.value).lower() for keyword in ['judul', 'title', 'laporan', 'report']):
                        formatting['header_format'][f"{cell.coordinate}"] = cell_format
                    elif 'total' in str(cell.value).lower() or 'jumlah' in str(cell.value).lower():
                        formatting['summary_format'][f"{cell.coordinate}"] = cell_format
                    else:
                        formatting['data_format'][f"{cell.coordinate}"] = cell_format

        return formatting

    def _get_used_range(self, sheet) -> Dict[str, int]:
        """Get the used range of the sheet"""
        return {
            'min_row': sheet.min_row,
            'max_row': sheet.max_row,
            'min_col': sheet.min_column,
            'max_col': sheet.max_column
        }

    def generate_report(self, data: Dict[str, Any], output_path: str) -> bool:
        """
        Generate report dengan data yang diberikan, menyesuaikan dengan template

        Args:
            data: Dictionary dengan data yang akan diisi
            output_path: Path untuk output file

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            self.logger.info("Starting adaptive report generation...")

            # Create new workbook from template
            new_workbook = self._create_workbook_from_template()

            # Process each sheet
            for sheet_name in new_workbook.sheetnames:
                if sheet_name in self.template_info['data_sections']:
                    success = self._process_sheet(new_workbook[sheet_name], data, sheet_name)
                    if not success:
                        self.logger.error(f"Failed to process sheet: {sheet_name}")
                        return False

            # Save the workbook
            new_workbook.save(output_path)
            self.logger.info(f"Report saved to: {output_path}")

            return True

        except Exception as e:
            self.logger.error(f"Error generating report: {e}")
            return False

    def _create_workbook_from_template(self) -> Workbook:
        """Create new workbook from template preserving formatting"""
        # Copy the template workbook
        new_workbook = Workbook()

        # Remove default sheet
        new_workbook.remove(new_workbook.active)

        # Copy all sheets from template
        for sheet_name in self.workbook.sheetnames:
            template_sheet = self.workbook[sheet_name]
            new_sheet = new_workbook.create_sheet(title=sheet_name)

            # Copy all cells with values and formatting
            for row in template_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]

                    # Copy value
                    new_cell.value = cell.value

                    # Copy formatting
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()

            # Copy column dimensions
            for col_letter, col_dim in template_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = col_dim.width
                new_sheet.column_dimensions[col_letter].hidden = col_dim.hidden

            # Copy row dimensions
            for row_num, row_dim in template_sheet.row_dimensions.items():
                new_sheet.row_dimensions[row_num].height = row_dim.height
                new_sheet.row_dimensions[row_num].hidden = row_dim.hidden

            # Copy sheet properties
            new_sheet.sheet_properties.outlinePr = template_sheet.sheet_properties.outlinePr
            new_sheet.page_setup = template_sheet.page_setup
            new_sheet.page_margins = template_sheet.page_margins

        return new_workbook

    def _process_sheet(self, sheet, data: Dict[str, Any], sheet_name: str) -> bool:
        """Process individual sheet with adaptive logic"""
        try:
            self.logger.info(f"Processing sheet: {sheet_name}")

            sheet_info = self.template_info['data_sections'][sheet_name]

            # Process single value placeholders
            self._process_single_placeholders(sheet, data, sheet_info['placeholders'])

            # Process dynamic tables
            for table in sheet_info['tables']:
                success = self._process_dynamic_table(sheet, data, table, sheet_name)
                if not success:
                    return False

            # Process template rows (repeating sections)
            for pattern in sheet_info['data_patterns']:
                if pattern['type'] == 'template_row':
                    success = self._process_template_row(sheet, data, pattern, sheet_name)
                    if not success:
                        return False

            return True

        except Exception as e:
            self.logger.error(f"Error processing sheet {sheet_name}: {e}")
            return False

    def _process_single_placeholders(self, sheet, data: Dict[str, Any], placeholders: List[Dict]):
        """Process single value placeholders"""
        for placeholder_info in placeholders:
            placeholder = placeholder_info['placeholder']
            cell = sheet[placeholder_info['cell']]

            # Try to find value in data
            value = self._find_placeholder_value(placeholder, data)

            if value is not None:
                # Replace placeholder with actual value with better type handling
                try:
                    if isinstance(value, (datetime, date)):
                        cell.value = value.strftime('%d %B %Y') if isinstance(value, date) else str(value)
                    elif isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str):
                        cell.value = value
                    elif value is None:
                        cell.value = ""
                    else:
                        # Try to convert to string
                        cell.value = str(value)
                except Exception as e:
                    self.logger.warning(f"Error setting cell value for placeholder: {e}, value: {value}")
                    cell.value = str(value) if value is not None else ""

                self.logger.debug(f"Replaced {placeholder} with {value} at {cell.coordinate}")
            else:
                self.logger.warning(f"No value found for placeholder: {placeholder} at {cell.coordinate}")

    def _find_placeholder_value(self, placeholder: str, data: Dict[str, Any]) -> Any:
        """Find value for placeholder in data using various lookup strategies"""
        # Direct lookup
        if placeholder in data:
            value = data[placeholder]
            # Safety check - don't return complex objects for single placeholders
            if isinstance(value, (list, dict, set, tuple)):
                self.logger.warning(f"Value for {placeholder} is complex, converting to string")
                return str(value) if value else ""
            return value

        # Case-insensitive lookup
        for key, value in data.items():
            if key.lower() == placeholder.lower():
                # Safety check
                if isinstance(value, (list, dict, set, tuple)):
                    self.logger.warning(f"Value for {placeholder} (case-insensitive) is complex, converting to string")
                    return str(value) if value else ""
                return value

        # Partial match lookup
        for key, value in data.items():
            if placeholder.lower() in key.lower() or key.lower() in placeholder.lower():
                # Safety check
                if isinstance(value, (list, dict, set, tuple)):
                    self.logger.warning(f"Value for {placeholder} (partial match) is complex, converting to string")
                    return str(value) if value else ""
                return value

        # Nested lookup
        if isinstance(data, dict):
            for nested_key, nested_value in data.items():
                if isinstance(nested_value, dict) and placeholder in nested_value:
                    value = nested_value[placeholder]
                    # Safety check
                    if isinstance(value, (list, dict, set, tuple)):
                        self.logger.warning(f"Value for {placeholder} (nested) is complex, converting to string")
                        return str(value) if value else ""
                    return value

        # Default values for common placeholders
        defaults = {
            'current_date': datetime.now().strftime('%d %B %Y'),
            'current_time': datetime.now().strftime('%H:%M:%S'),
            'generated_date': datetime.now().strftime('%d %B %Y'),
            'generated_time': datetime.now().strftime('%H:%M:%S'),
            'report_date': datetime.now().strftime('%d %B %Y'),
        }

        if placeholder.lower() in defaults:
            return defaults[placeholder.lower()]

        return None

    def _process_dynamic_table(self, sheet, data: Dict[str, Any], table: Dict, sheet_name: str) -> bool:
        """Process dynamic table by expanding template rows"""
        try:
            # Find data source for this table
            table_data = self._find_table_data(table, data, sheet_name)

            if not table_data:
                self.logger.warning(f"No data found for table in sheet: {sheet_name}")
                return True  # Not an error, just no data

            # Get template row
            template_row_num = table['template_row']
            template_row = sheet[template_row_num]

            # Get placeholder mapping for this table
            placeholder_mapping = {}
            for col_idx, template_value in enumerate(table['template_values'], 1):
                placeholders = self._extract_placeholders(template_value)
                if placeholders:
                    placeholder_mapping[col_idx] = placeholders[0]  # Use first placeholder

            # Insert new rows for data
            data_start_row = template_row_num + 1
            rows_to_insert = len(table_data)

            if rows_to_insert > 1:
                sheet.insert_rows(data_start_row, rows_to_insert - 1)

            # Fill data rows
            for row_idx, row_data in enumerate(table_data):
                current_row = data_start_row + row_idx

                for col_idx, placeholder in placeholder_mapping.items():
                    cell = sheet.cell(row=current_row, column=col_idx)

                    # Get value for this placeholder
                    value = self._get_value_from_row_data(placeholder, row_data)

                    if value is not None:
                        # Apply formatting from template row
                        template_cell = sheet.cell(row=template_row_num, column=col_idx)
                        self._copy_cell_formatting(template_cell, cell)

                        # Set value with better type handling
                        try:
                            if isinstance(value, (datetime, date)):
                                cell.value = value.strftime('%d %B %Y') if isinstance(value, date) else str(value)
                            elif isinstance(value, (int, float)):
                                cell.value = value
                            elif isinstance(value, str):
                                cell.value = value
                            elif value is None:
                                cell.value = ""
                            else:
                                # Try to convert to string
                                cell.value = str(value)
                        except Exception as e:
                            self.logger.warning(f"Error setting cell value: {e}, value: {value}")
                            cell.value = str(value) if value is not None else ""

            self.logger.info(f"Processed table with {len(table_data)} rows in sheet: {sheet_name}")
            return True

        except Exception as e:
            self.logger.error(f"Error processing table in sheet {sheet_name}: {e}")
            return False

    def _process_template_row(self, sheet, data: Dict[str, Any], pattern: Dict, sheet_name: str) -> bool:
        """Process template row pattern (alternative to table detection)"""
        try:
            # Find data source for this pattern
            pattern_data = self._find_pattern_data(pattern, data, sheet_name)

            if not pattern_data:
                self.logger.warning(f"No data found for pattern in sheet: {sheet_name}")
                return True

            template_row_num = pattern['row']
            rows_to_insert = len(pattern_data)

            if rows_to_insert > 1:
                sheet.insert_rows(template_row_num + 1, rows_to_insert - 1)

            # Fill data rows
            for row_idx, row_data in enumerate(pattern_data):
                current_row = template_row_num + row_idx

                for placeholder_info in pattern['placeholders']:
                    col_num = column_index_from_string(placeholder_info['column'])
                    cell = sheet.cell(row=current_row, column=col_num)

                    # Get value for this placeholder
                    value = self._get_value_from_row_data(placeholder_info['placeholder'], row_data)

                    if value is not None:
                        # Copy formatting from template
                        template_cell = sheet.cell(row=template_row_num, column=col_num)
                        self._copy_cell_formatting(template_cell, cell)

                        # Set value with better type handling
                        try:
                            if isinstance(value, (datetime, date)):
                                cell.value = value.strftime('%d %B %Y') if isinstance(value, date) else str(value)
                            elif isinstance(value, (int, float)):
                                cell.value = value
                            elif isinstance(value, str):
                                cell.value = value
                            elif value is None:
                                cell.value = ""
                            else:
                                # Try to convert to string
                                cell.value = str(value)
                        except Exception as e:
                            self.logger.warning(f"Error setting cell value in pattern: {e}, value: {value}")
                            cell.value = str(value) if value is not None else ""

            self.logger.info(f"Processed pattern with {len(pattern_data)} rows in sheet: {sheet_name}")
            return True

        except Exception as e:
            self.logger.error(f"Error processing pattern in sheet {sheet_name}: {e}")
            return False

    def _find_table_data(self, table: Dict, data: Dict[str, Any], sheet_name: str) -> Optional[List[Dict]]:
        """Find data source for a table"""
        # Try various data source strategies

        # 1. Look for data keyed by sheet name
        if sheet_name in data and isinstance(data[sheet_name], list):
            return data[sheet_name]

        # 2. Look for common table data keys
        table_keys = [
            'data', 'rows', 'records', 'items',
            sheet_name.lower().replace(' ', '_'),
            sheet_name.lower()
        ]

        for key in table_keys:
            if key in data and isinstance(data[key], list):
                return data[key]

        # 3. Look for data based on headers
        headers = [h.strip() for h in table['headers'] if h.strip()]
        if headers:
            for data_key, data_value in data.items():
                if isinstance(data_value, list) and data_value:
                    # Check if first row has matching keys
                    first_row = data_value[0]
                    if isinstance(first_row, dict):
                        matching_keys = sum(1 for header in headers if any(header.lower() in str(k).lower() for k in first_row.keys()))
                        if matching_keys >= len(headers) / 2:  # At least 50% match
                            return data_value

        return None

    def _find_pattern_data(self, pattern: Dict, data: Dict[str, Any], sheet_name: str) -> Optional[List[Dict]]:
        """Find data source for a pattern"""
        # Similar to table data finding but more flexible
        return self._find_table_data({'headers': [ph['placeholder'] for ph in pattern['placeholders']]}, data, sheet_name)

    def _get_value_from_row_data(self, placeholder: str, row_data: Dict) -> Any:
        """Get value from row data for a placeholder"""
        if not isinstance(row_data, dict):
            self.logger.warning(f"Row data is not a dictionary: {type(row_data)} - {row_data}")
            return None

        # Direct lookup
        if placeholder in row_data:
            value = row_data[placeholder]
            # Additional safety check - don't return complex objects
            if isinstance(value, (list, dict, set, tuple)):
                self.logger.warning(f"Value for {placeholder} is a complex type: {type(value)}")
                return str(value) if value else ""
            return value

        # Case-insensitive lookup
        for key, value in row_data.items():
            if key.lower() == placeholder.lower():
                # Additional safety check
                if isinstance(value, (list, dict, set, tuple)):
                    self.logger.warning(f"Value for {placeholder} (case-insensitive) is a complex type: {type(value)}")
                    return str(value) if value else ""
                return value

        # Partial match lookup
        for key, value in row_data.items():
            if placeholder.lower() in key.lower() or key.lower() in placeholder.lower():
                # Additional safety check
                if isinstance(value, (list, dict, set, tuple)):
                    self.logger.warning(f"Value for {placeholder} (partial match) is a complex type: {type(value)}")
                    return str(value) if value else ""
                return value

        return None

    def _copy_cell_formatting(self, source_cell, target_cell):
        """Copy formatting from source cell to target cell"""
        try:
            if source_cell and target_cell:
                if hasattr(source_cell, 'has_style') and source_cell.has_style:
                    if hasattr(source_cell, 'font') and source_cell.font:
                        target_cell.font = source_cell.font.copy()
                    if hasattr(source_cell, 'border') and source_cell.border:
                        target_cell.border = source_cell.border.copy()
                    if hasattr(source_cell, 'fill') and source_cell.fill:
                        target_cell.fill = source_cell.fill.copy()
                    if hasattr(source_cell, 'number_format'):
                        target_cell.number_format = source_cell.number_format
                    if hasattr(source_cell, 'alignment') and source_cell.alignment:
                        target_cell.alignment = source_cell.alignment.copy()
        except Exception as e:
            self.logger.warning(f"Error copying cell formatting: {e}")
            # Continue without formatting if copying fails

    def get_template_summary(self) -> Dict[str, Any]:
        """Get comprehensive template summary"""
        summary = {
            'template_path': self.template_path,
            'analysis_date': datetime.now().isoformat(),
            'template_info': self.template_info,
            'placeholder_map': dict(list(self.placeholder_map.items())[:10]),  # First 10 for brevity
            'data_sections_count': len([s for s in self.template_info['data_sections'].values() if s['tables'] or s['data_patterns']]),
            'capabilities': {
                'supports_dynamic_tables': len([t for s in self.template_info['data_sections'].values() for t in s['tables']]) > 0,
                'supports_template_rows': len([p for s in self.template_info['data_sections'].values() for p in s['data_patterns']]) > 0,
                'total_placeholders': self.template_info['total_placeholders'],
                'sheets_with_placeholders': len([s for s in self.template_info['data_sections'].values() if s['placeholder_count'] > 0])
            }
        }

        return summary


def main():
    """Test function"""
    template_path = "templates/Template_Laporan_FFB_Analysis.xlsx"

    if not os.path.exists(template_path):
        print(f"Template not found: {template_path}")
        return

    processor = AdaptiveExcelProcessor(template_path)

    # Print template summary
    summary = processor.get_template_summary()
    print("Template Summary:")
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()