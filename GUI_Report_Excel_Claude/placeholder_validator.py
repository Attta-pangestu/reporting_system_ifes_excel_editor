#!/usr/bin/env python3
"""
Placeholder Validator untuk Template Processing
Memastikan semua placeholder dalam template memiliki nilai sebelum report generation
"""

import os
import json
import logging
from typing import Dict, List, Set, Tuple, Any
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

class PlaceholderValidator:
    """
    Validator untuk memastikan semua placeholder dalam template terisi dengan benar
    """

    def __init__(self, template_path: str, formula_path: str):
        """
        Inisialisasi validator

        Args:
            template_path: Path ke file template Excel
            formula_path: Path ke file formula JSON
        """
        self.template_path = template_path
        self.formula_path = formula_path
        self.logger = logging.getLogger(__name__)

        # Setup logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )

        # Load template and formula
        self.workbook = None
        self.formula_data = None
        self.placeholders = {}
        self.validation_results = {}

        self._load_template()
        self._load_formula()
        self._extract_placeholders()

    def _load_template(self):
        """Load template Excel file"""
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")

            self.workbook = load_workbook(self.template_path, data_only=True)
            self.logger.info(f"Template loaded: {self.template_path}")
            self.logger.info(f"  Sheets: {self.workbook.sheetnames}")

        except Exception as e:
            self.logger.error(f"Error loading template: {e}")
            raise

    def _load_formula(self):
        """Load formula JSON file"""
        try:
            if not os.path.exists(self.formula_path):
                raise FileNotFoundError(f"Formula file not found: {self.formula_path}")

            with open(self.formula_path, 'r', encoding='utf-8') as f:
                self.formula_data = json.load(f)

            self.logger.info(f"Formula loaded: {self.formula_path}")

        except Exception as e:
            self.logger.error(f"Error loading formula: {e}")
            raise

    def _extract_placeholders(self):
        """Extract all placeholders from template"""
        self.placeholders = {}

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_placeholders = []

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        placeholders = self._find_placeholders_in_text(cell.value)
                        if placeholders:
                            sheet_placeholders.extend([
                                {
                                    'placeholder': ph,
                                    'cell': cell.coordinate,
                                    'value': cell.value,
                                    'sheet': sheet_name
                                }
                                for ph in placeholders
                            ])

            if sheet_placeholders:
                self.placeholders[sheet_name] = sheet_placeholders

        self.logger.info(f"Found placeholders in {len(self.placeholders)} sheets")
        for sheet_name, ph_list in self.placeholders.items():
            self.logger.info(f"  {sheet_name}: {len(ph_list)} placeholders")

    def _find_placeholders_in_text(self, text: str) -> List[str]:
        """Find all {{placeholder}} patterns in text"""
        import re
        pattern = r'\{\{([^}]+)\}\}'
        return re.findall(pattern, text)

    def validate_against_formula(self) -> Dict[str, Any]:
        """
        Validate placeholders against formula definitions

        Returns:
            Dictionary dengan validation results
        """
        validation_results = {
            'total_placeholders': 0,
            'unique_placeholders': set(),
            'missing_variables': [],
            'repeating_sections': {},
            'sheet_validation': {},
            'validation_passed': True,
            'validation_timestamp': datetime.now().isoformat()
        }

        # Get all placeholders
        all_placeholders = []
        for sheet_name, ph_list in self.placeholders.items():
            for ph_info in ph_list:
                all_placeholders.append(ph_info['placeholder'])
                validation_results['unique_placeholders'].add(ph_info['placeholder'])

        validation_results['total_placeholders'] = len(all_placeholders)

        # Check formula variables
        formula_variables = self._extract_formula_variables()

        # Find missing variables
        found_variables = set(validation_results['unique_placeholders'])
        defined_variables = set(formula_variables.keys())

        validation_results['missing_variables'] = list(found_variables - defined_variables)

        # Validate repeating sections
        validation_results['repeating_sections'] = self._validate_repeating_sections()

        # Per-sheet validation
        for sheet_name in self.workbook.sheetnames:
            sheet_result = self._validate_sheet(sheet_name)
            validation_results['sheet_validation'][sheet_name] = sheet_result

        # Overall validation status
        if (validation_results['missing_variables'] or
            any(not result['is_valid'] for result in validation_results['sheet_validation'].values())):
            validation_results['validation_passed'] = False

        return validation_results

    def _extract_formula_variables(self) -> Dict[str, Any]:
        """Extract all variables from formula JSON"""
        variables = {}

        if 'variables' in self.formula_data:
            for category, var_dict in self.formula_data['variables'].items():
                if isinstance(var_dict, dict):
                    for var_name, var_info in var_dict.items():
                        variables[var_name] = {
                            'category': category,
                            'type': var_info.get('type', 'unknown'),
                            'source': var_info.get('source', 'unknown')
                        }

        return variables

    def _validate_repeating_sections(self) -> Dict[str, Any]:
        """Validate repeating sections in template"""
        repeating_sections = {}

        # Look for repeating sections pattern (row 8)
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            row_8_placeholders = []

            for cell in sheet[8]:  # Row 8
                if cell.value and isinstance(cell.value, str):
                    placeholders = self._find_placeholders_in_text(cell.value)
                    if placeholders:
                        row_8_placeholders.extend(placeholders)

            if row_8_placeholders:
                repeating_sections[sheet_name] = {
                    'template_row': 8,
                    'placeholders': row_8_placeholders,
                    'has_repeating_data': True
                }
            else:
                repeating_sections[sheet_name] = {
                    'template_row': None,
                    'placeholders': [],
                    'has_repeating_data': False
                }

        return repeating_sections

    def _validate_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Validate individual sheet"""
        sheet = self.workbook[sheet_name]
        sheet_placeholders = self.placeholders.get(sheet_name, [])

        validation_result = {
            'is_valid': True,
            'placeholder_count': len(sheet_placeholders),
            'missing_placeholders': [],
            'issues': []
        }

        # Check for empty placeholders
        for ph_info in sheet_placeholders:
            if not ph_info['placeholder'].strip():
                validation_result['issues'].append({
                    'type': 'empty_placeholder',
                    'cell': ph_info['cell'],
                    'message': 'Empty placeholder found'
                })
                validation_result['is_valid'] = False

        # Check for duplicate placeholders in same sheet
        placeholder_counts = {}
        for ph_info in sheet_placeholders:
            ph = ph_info['placeholder']
            placeholder_counts[ph] = placeholder_counts.get(ph, 0) + 1

        duplicates = {ph: count for ph, count in placeholder_counts.items() if count > 1}
        if duplicates:
            validation_result['issues'].append({
                'type': 'duplicate_placeholders',
                'duplicates': duplicates,
                'message': f'Duplicate placeholders found: {list(duplicates.keys())}'
            })

        return validation_result

    def validate_data_completeness(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Validate that all placeholders have corresponding data

        Args:
            data: Dictionary dengan data yang akan diisi ke template

        Returns:
            Dictionary dengan validation results
        """
        completeness_results = {
            'is_complete': True,
            'missing_data': [],
            'data_sources': {},
            'validation_timestamp': datetime.now().isoformat()
        }

        # Flatten all data keys
        all_data_keys = set()
        for key, value in data.items():
            if isinstance(value, dict):
                all_data_keys.update(value.keys())
            else:
                all_data_keys.add(key)

        # Check each placeholder
        for sheet_name, ph_list in self.placeholders.items():
            for ph_info in ph_list:
                placeholder = ph_info['placeholder']

                if placeholder not in all_data_keys:
                    completeness_results['missing_data'].append({
                        'placeholder': placeholder,
                        'sheet': sheet_name,
                        'cell': ph_info['cell'],
                        'message': f"No data found for placeholder: {placeholder}"
                    })
                    completeness_results['is_complete'] = False

        return completeness_results

    def generate_validation_report(self) -> str:
        """Generate comprehensive validation report"""
        validation_results = self.validate_against_formula()

        report_lines = [
            "=" * 80,
            "PLACEHOLDER VALIDATION REPORT",
            "=" * 80,
            f"Template: {self.template_path}",
            f"Formula: {self.formula_path}",
            f"Validation Date: {validation_results['validation_timestamp']}",
            "",
            "SUMMARY:",
            f"  Total Placeholders: {validation_results['total_placeholders']}",
            f"  Unique Placeholders: {len(validation_results['unique_placeholders'])}",
            f"  Validation Status: {'PASSED' if validation_results['validation_passed'] else 'FAILED'}",
            ""
        ]

        # Missing variables
        if validation_results['missing_variables']:
            report_lines.extend([
                "❌ MISSING VARIABLES:",
                *[f"  - {var}" for var in validation_results['missing_variables']],
                ""
            ])

        # Sheet validation
        report_lines.append("SHEET VALIDATION:")
        for sheet_name, result in validation_results['sheet_validation'].items():
            status = "✅" if result['is_valid'] else "❌"
            report_lines.append(f"  {status} {sheet_name}: {result['placeholder_count']} placeholders")

            if result['issues']:
                for issue in result['issues']:
                    report_lines.append(f"     Issue: {issue['message']}")

        # Repeating sections
        report_lines.extend([
            "",
            "REPEATING SECTIONS:",
        ])

        for sheet_name, section in validation_results['repeating_sections'].items():
            if section['has_repeating_data']:
                report_lines.append(f"  📊 {sheet_name}: Row {section['template_row']} ({len(section['placeholders'])} placeholders)")
            else:
                report_lines.append(f"  📄 {sheet_name}: No repeating data")

        report_lines.extend([
            "",
            "=" * 80,
            "END OF REPORT",
            "=" * 80
        ])

        return "\n".join(report_lines)

    def save_validation_report(self, output_path: str = None):
        """Save validation report to file"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"validation_report_{timestamp}.txt"

        report = self.generate_validation_report()

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report)

        self.logger.info(f"Validation report saved: {output_path}")
        return output_path


def main():
    """Test function"""
    template_path = "templates/Template_Laporan_FFB_Analysis.xlsx"
    formula_path = "laporan_ffb_analysis_formula.json"

    if not os.path.exists(template_path):
        print(f"Template not found: {template_path}")
        return

    if not os.path.exists(formula_path):
        print(f"Formula not found: {formula_path}")
        return

    validator = PlaceholderValidator(template_path, formula_path)

    # Generate and save report
    report_path = validator.save_validation_report()
    print(f"Validation report: {report_path}")

    # Print summary
    validation_results = validator.validate_against_formula()
    print(f"\nValidation Status: {'PASSED' if validation_results['validation_passed'] else 'FAILED'}")

    if validation_results['missing_variables']:
        print(f"Missing Variables: {len(validation_results['missing_variables'])}")
        for var in validation_results['missing_variables'][:5]:
            print(f"  - {var}")
        if len(validation_results['missing_variables']) > 5:
            print(f"  ... and {len(validation_results['missing_variables']) - 5} more")


if __name__ == "__main__":
    main()