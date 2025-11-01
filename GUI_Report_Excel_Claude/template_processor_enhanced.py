"""
Enhanced Template Processor Engine
Menangani pembacaan template Excel dan pemrosesan placeholder variables dengan debug logging
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
import json
import os
from typing import Dict, List, Any, Tuple, Optional
import pandas as pd
from datetime import datetime, date
import logging

class TemplateProcessorEnhanced:
    """
    Enhanced template processor dengan debug logging dan perbaikan placeholder processing
    """

    def __init__(self, template_path: str, formula_path: str):
        """
        Inisialisasi enhanced template processor

        Args:
            template_path: Path ke file template Excel
            formula_path: Path ke file formula definition (JSON)
        """
        self.template_path = template_path
        self.formula_path = formula_path
        self.workbook = None
        self.formulas = {}
        self.placeholders = {}
        self.repeating_sections = {}

        # Setup enhanced logging
        self._setup_enhanced_logging()

        # Load template dan formulas
        self._load_template()
        self._load_formulas()
        self._scan_placeholders()
        self._parse_repeating_sections()

    def _setup_enhanced_logging(self):
        """Setup enhanced logging dengan debug detail"""
        self.logger = logging.getLogger(f"{__name__}_enhanced")
        self.logger.setLevel(logging.DEBUG)

        # Create console handler if not exists
        if not self.logger.handlers:
            console_handler = logging.StreamHandler()
            console_handler.setLevel(logging.DEBUG)

            # Create detailed formatter
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s'
            )
            console_handler.setFormatter(formatter)
            self.logger.addHandler(console_handler)
            self.logger.propagate = False

    def _load_template(self):
        """Load template Excel dengan debug logging"""
        try:
            self.logger.info(f"Loading template from: {self.template_path}")

            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file tidak ditemukan: {self.template_path}")

            self.workbook = openpyxl.load_workbook(self.template_path)
            sheet_names = self.workbook.sheetnames

            self.logger.info(f"Template loaded successfully")
            self.logger.info(f"Available sheets: {sheet_names}")
            self.logger.info(f"Total sheets: {len(sheet_names)}")

            # Log sheet details
            for sheet_name in sheet_names:
                sheet = self.workbook[sheet_name]
                max_row = sheet.max_row
                max_col = sheet.max_column
                self.logger.debug(f"Sheet '{sheet_name}': {max_row} rows x {max_col} columns")

        except Exception as e:
            self.logger.error(f"Error loading template: {e}", exc_info=True)
            raise

    def _load_formulas(self):
        """Load formula definitions dari file JSON dengan debug"""
        try:
            self.logger.info(f"Loading formulas from: {self.formula_path}")

            if not os.path.exists(self.formula_path):
                raise FileNotFoundError(f"Formula file tidak ditemukan: {self.formula_path}")

            with open(self.formula_path, 'r', encoding='utf-8') as f:
                self.formulas = json.load(f)

            self.logger.info(f"Formulas loaded successfully")

            # Log formula structure
            queries_count = len(self.formulas.get('queries', {}))
            variables_count = sum(len(vars) for vars in self.formulas.get('variables', {}).values())
            repeating_sections_count = len(self.formulas.get('repeating_sections', {}))

            self.logger.info(f"Formula structure: {queries_count} queries, {variables_count} variables, {repeating_sections_count} repeating sections")

        except Exception as e:
            self.logger.error(f"Error loading formulas: {e}", exc_info=True)
            raise

    def _scan_placeholders(self):
        """Scan template untuk menemukan semua placeholder variables dengan debug detail"""
        placeholder_pattern = r'\{\{([^}]+)\}\}'

        self.logger.info("=== SCANNING PLACEHOLDERS ===")

        total_placeholders = 0

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_placeholders = []

            self.logger.debug(f"Scanning sheet: {sheet_name}")

            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value and isinstance(cell.value, str):
                        matches = re.findall(placeholder_pattern, cell.value)
                        if matches:
                            for match in matches:
                                placeholder_info = {
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'row': row_idx,
                                    'column': col_idx,
                                    'placeholder': match,
                                    'full_placeholder': f'{{{{{match}}}}}',
                                    'original_value': cell.value
                                }
                                sheet_placeholders.append(placeholder_info)
                                total_placeholders += 1

                                self.logger.debug(f"Found placeholder '{match}' at {sheet_name}!{cell.coordinate} (row {row_idx}, col {col_idx})")
                                self.logger.debug(f"  Full text: {cell.value}")

            if sheet_placeholders:
                self.placeholders[sheet_name] = sheet_placeholders
                self.logger.info(f"Sheet '{sheet_name}': {len(sheet_placeholders)} placeholders")
            else:
                self.logger.info(f"Sheet '{sheet_name}': No placeholders found")

        self.logger.info(f"=== PLACEHOLDER SCAN COMPLETED ===")
        self.logger.info(f"Total placeholders found: {total_placeholders}")

        # Log all placeholders by category
        for sheet_name, placeholders in self.placeholders.items():
            self.logger.debug(f"Placeholders in sheet '{sheet_name}':")
            for ph in placeholders:
                self.logger.debug(f"  {ph['placeholder']} at {ph['cell']}")

    def _parse_repeating_sections(self):
        """Parse repeating sections dari formula definitions dengan debug"""
        repeating_config = self.formulas.get('repeating_sections', {})

        self.logger.info(f"=== PARSING REPEATING SECTIONS ===")
        self.logger.info(f"Found {len(repeating_config)} repeating sections in formulas")

        for section_name, section_config in repeating_config.items():
            sheet_name = section_config.get('sheet_name')
            if not sheet_name:
                self.logger.warning(f"Repeating section '{section_name}' has no sheet_name, skipping")
                continue

            self.repeating_sections[sheet_name] = {
                'name': section_name,
                'data_source': section_config.get('data_source'),
                'start_row': section_config.get('start_row', 1),
                'template_rows': section_config.get('template_rows', 1),
                'columns': section_config.get('columns', {}),
                'group_by': section_config.get('group_by')
            }

            self.logger.info(f"Repeating section '{section_name}' -> sheet '{sheet_name}'")
            self.logger.debug(f"  Data source: {section_config.get('data_source')}")
            self.logger.debug(f"  Start row: {section_config.get('start_row', 1)}")
            self.logger.debug(f"  Template rows: {section_config.get('template_rows', 1)}")
            self.logger.debug(f"  Columns: {len(section_config.get('columns', {}))} defined")

        self.logger.info(f"=== REPEATING SECTIONS PARSED ===")

    def get_sheet_placeholders(self, sheet_name: str) -> List[Dict]:
        """Get placeholders untuk sheet tertentu"""
        placeholders = self.placeholders.get(sheet_name, [])
        self.logger.debug(f"Retrieved {len(placeholders)} placeholders for sheet '{sheet_name}'")
        return placeholders

    def get_placeholders(self) -> Dict[str, List[Dict]]:
        """Get all placeholders from all sheets"""
        total = sum(len(p) for p in self.placeholders.values())
        self.logger.debug(f"Retrieved all placeholders: {total} total across {len(self.placeholders)} sheets")
        return self.placeholders

    def get_repeating_section(self, sheet_name: str) -> Optional[Dict]:
        """Get repeating section configuration untuk sheet tertentu"""
        section = self.repeating_sections.get(sheet_name)
        if section:
            self.logger.debug(f"Found repeating section for sheet '{sheet_name}': {section['name']}")
        else:
            self.logger.debug(f"No repeating section found for sheet '{sheet_name}'")
        return section

    def process_sheet_placeholders(self, sheet_name: str, data: Dict[str, Any]) -> bool:
        """
        Process placeholders pada sheet tertentu dengan enhanced debugging

        Args:
            sheet_name: Nama sheet yang akan diproses
            data: Data untuk substitusi placeholder

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            self.logger.info(f"=== PROCESSING PLACEHOLDERS FOR SHEET: {sheet_name} ===")

            if sheet_name not in self.workbook.sheetnames:
                self.logger.error(f"Sheet {sheet_name} tidak ditemukan in workbook")
                return False

            sheet = self.workbook[sheet_name]
            placeholders = self.get_sheet_placeholders(sheet_name)

            self.logger.info(f"Found {len(placeholders)} placeholders to process in sheet '{sheet_name}'")
            self.logger.debug(f"Available data keys: {list(data.keys())}")

            processed_count = 0
            skipped_count = 0

            for placeholder_info in placeholders:
                placeholder = placeholder_info['placeholder']
                cell_coord = placeholder_info['cell']
                cell = sheet[cell_coord]

                self.logger.debug(f"Processing placeholder '{placeholder}' at {cell_coord}")

                # Get value untuk placeholder
                value = self._get_placeholder_value(placeholder, data)

                if value is not None:
                    # Apply formatting jika diperlukan
                    formatted_value = self._apply_formatting(placeholder, value, data)

                    # Set cell value
                    original_value = cell.value
                    cell.value = formatted_value

                    # Apply cell styling
                    self._apply_cell_style(cell, placeholder, formatted_value)

                    processed_count += 1

                    self.logger.debug(f"  SUCCESS: '{placeholder}' = {formatted_value} (type: {type(formatted_value)})")
                    self.logger.debug(f"  Changed from '{original_value}' to '{formatted_value}'")
                else:
                    skipped_count += 1
                    self.logger.warning(f"  SKIPPED: No value found for placeholder '{placeholder}'")
                    self.logger.debug(f"  Available data: {list(data.keys())}")

            self.logger.info(f"=== PLACEHOLDER PROCESSING COMPLETED FOR {sheet_name} ===")
            self.logger.info(f"Processed: {processed_count}, Skipped: {skipped_count}")

            return True

        except Exception as e:
            self.logger.error(f"Error processing sheet {sheet_name}: {e}", exc_info=True)
            return False

    def _get_placeholder_value(self, placeholder: str, data: Dict[str, Any]) -> Any:
        """Get value untuk placeholder dari data dengan enhanced lookup"""
        try:
            # Handle complex placeholders (formatting, calculations)
            if '|' in placeholder:
                base_placeholder = placeholder.split('|')[0]
            else:
                base_placeholder = placeholder

            self.logger.debug(f"Looking up placeholder '{base_placeholder}' in data")

            # Direct lookup
            if base_placeholder in data:
                value = data[base_placeholder]
                self.logger.debug(f"  Direct lookup found: {value} (type: {type(value)})")
                return value

            # Nested lookup (dot notation)
            if '.' in base_placeholder:
                keys = base_placeholder.split('.')
                value = data
                path_traversed = []

                for key in keys:
                    path_traversed.append(key)
                    current_path = '.'.join(path_traversed)

                    if isinstance(value, dict) and key in value:
                        value = value[key]
                        self.logger.debug(f"  Nested lookup at '{current_path}': {value} (type: {type(value)})")
                    else:
                        self.logger.debug(f"  Nested lookup failed at '{current_path}', key '{key}' not found")
                        return None

                return value

            # Check in nested structures (query results)
            for key, val in data.items():
                if isinstance(val, dict) and base_placeholder in val:
                    value = val[base_placeholder]
                    self.logger.debug(f"  Found in nested dict '{key}': {value} (type: {type(value)})")
                    return value

            # Default values
            defaults = {
                'current_date': datetime.now().strftime('%d %B %Y'),
                'current_time': datetime.now().strftime('%H:%M:%S'),
                'report_title': 'LAPORAN ANALISIS FFB',
                'report_period': 'Periode Laporan',
                'estate_name': 'PGE 2B',
                'total_transactions': 0,
                'total_ripe_bunches': 0,
                'verification_rate': 0
            }

            if base_placeholder in defaults:
                value = defaults[base_placeholder]
                self.logger.debug(f"  Using default value for '{base_placeholder}': {value}")
                return value

            self.logger.debug(f"  No value found for placeholder '{base_placeholder}'")
            return None

        except Exception as e:
            self.logger.error(f"Error getting placeholder value for '{placeholder}': {e}")
            return None

    def _apply_formatting(self, placeholder: str, value: Any, data: Dict[str, Any]) -> Any:
        """Apply formatting ke value dengan debug"""
        try:
            if '|' in placeholder:
                format_spec = placeholder.split('|', 1)[1]
                self.logger.debug(f"Applying formatting '{format_spec}' to value {value}")

                if format_spec.startswith('format:'):
                    format_string = format_spec[7:]  # Remove 'format:'
                    if isinstance(value, (datetime, date)):
                        formatted = value.strftime(format_string)
                        self.logger.debug(f"  Date formatting: {value} -> {formatted}")
                        return formatted
                    elif isinstance(value, (int, float)):
                        try:
                            formatted = format(value, format_string)
                            self.logger.debug(f"  Number formatting: {value} -> {formatted}")
                            return formatted
                        except ValueError as e:
                            self.logger.warning(f"  Number formatting failed: {e}")
                            return value

            return value

        except Exception as e:
            self.logger.error(f"Error applying formatting to '{placeholder}': {e}")
            return value

    def _apply_cell_style(self, cell, placeholder: str, value: Any):
        """Apply styling ke cell berdasarkan value"""
        try:
            # Remove placeholder styling
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == 'FFFF5F5F':
                cell.fill = PatternFill(fill_type=None)

            if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF6B6B':
                cell.font = cell.font.copy(color='000000')

            if cell.font and cell.font.italic:
                cell.font = cell.font.copy(italic=False)

            # Apply numeric formatting
            if isinstance(value, (int, float)):
                if '.' in str(value):
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'

        except Exception as e:
            self.logger.error(f"Error applying cell style for placeholder '{placeholder}': {e}")

    def process_repeating_section(self, sheet_name: str, data: List[Dict[str, Any]]) -> bool:
        """
        Process repeating section pada sheet dengan data array dan debug

        Args:
            sheet_name: Nama sheet
            data: Array data untuk repeating section

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            self.logger.info(f"=== PROCESSING REPEATING SECTION FOR SHEET: {sheet_name} ===")

            if sheet_name not in self.workbook.sheetnames:
                self.logger.error(f"Sheet {sheet_name} tidak ditemukan")
                return False

            repeating_section = self.get_repeating_section(sheet_name)
            if not repeating_section:
                self.logger.warning(f"No repeating section found for sheet {sheet_name}")
                return False

            self.logger.info(f"Processing {len(data)} rows for repeating section '{repeating_section['name']}'")

            sheet = self.workbook[sheet_name]
            start_row = repeating_section['start_row']
            template_rows = repeating_section['template_rows']
            columns = repeating_section['columns']

            self.logger.debug(f"Start row: {start_row}, Template rows: {template_rows}")
            self.logger.debug(f"Columns to process: {list(columns.keys())}")

            # Delete existing template rows (keep one for template)
            current_max_row = sheet.max_row
            if current_max_row > start_row + template_rows - 1:
                rows_to_delete = current_max_row - (start_row + template_rows - 1)
                sheet.delete_rows(start_row + template_rows, rows_to_delete)
                self.logger.debug(f"Deleted {rows_to_delete} existing rows")

            # Process each data row
            for i, row_data in enumerate(data):
                target_row = start_row + i

                self.logger.debug(f"Processing data row {i+1}/{len(data)} at target row {target_row}")

                # Copy template row if needed
                if i > 0:
                    self._copy_template_row(sheet, start_row, target_row, template_rows)

                # Fill data in columns
                for col_letter, column_config in columns.items():
                    field_name = column_config.get('field')
                    if field_name and field_name in row_data:
                        cell = sheet[f"{col_letter}{target_row}"]
                        value = row_data[field_name]

                        # Apply formatting
                        formatted_value = self._apply_column_formatting(value, column_config)

                        # Set value
                        cell.value = formatted_value

                        # Apply styling
                        self._apply_column_style(cell, column_config)

                        self.logger.debug(f"  Set {col_letter}{target_row} = {formatted_value} (field: {field_name})")
                    else:
                        self.logger.debug(f"  Field '{field_name}' not found in row data")

            self.logger.info(f"=== REPEATING SECTION PROCESSING COMPLETED FOR {sheet_name} ===")
            return True

        except Exception as e:
            self.logger.error(f"Error processing repeating section in sheet {sheet_name}: {e}", exc_info=True)
            return False

    def _apply_column_formatting(self, value: Any, column_config: Dict) -> Any:
        """Apply column formatting ke value dengan debug"""
        try:
            format_type = column_config.get('format', 'text')

            if format_type == 'date' and value:
                if isinstance(value, str):
                    try:
                        value = datetime.strptime(value, '%Y-%m-%d')
                    except:
                        pass

                if isinstance(value, (datetime, date)):
                    format_string = column_config.get('format_string', 'dd/mm/yyyy')
                    formatted = value.strftime(format_string)
                    self.logger.debug(f"  Date formatting: {value} -> {formatted}")
                    return formatted

            elif format_type == 'integer' and value:
                try:
                    formatted = int(float(value))
                    self.logger.debug(f"  Integer formatting: {value} -> {formatted}")
                    return formatted
                except:
                    return value

            elif format_type == 'number' and value:
                try:
                    decimal_places = column_config.get('decimal_places', 2)
                    formatted = round(float(value), decimal_places)
                    self.logger.debug(f"  Number formatting: {value} -> {formatted}")
                    return formatted
                except:
                    return value

            elif format_type == 'percentage' and value:
                try:
                    formatted = float(value) / 100  # Excel stores percentage as decimal
                    self.logger.debug(f"  Percentage formatting: {value} -> {formatted}")
                    return formatted
                except:
                    return value

            # Handle mapping
            mapping = column_config.get('mapping', {})
            if mapping and value in mapping:
                formatted = mapping[value]
                self.logger.debug(f"  Mapping: {value} -> {formatted}")
                return formatted

            return value

        except Exception as e:
            self.logger.error(f"Error applying column formatting: {e}")
            return value

    def _apply_column_style(self, cell, column_config: Dict):
        """Apply column styling ke cell"""
        try:
            format_type = column_config.get('format', 'text')

            # Set number format
            if format_type == 'integer':
                cell.number_format = '#,##0'
            elif format_type == 'number':
                decimal_places = column_config.get('decimal_places', 2)
                cell.number_format = f'#,##0.{"".join(["0"] * decimal_places)}'
            elif format_type == 'percentage':
                decimal_places = column_config.get('decimal_places', 2)
                cell.number_format = f'0.{"".join(["0"] * decimal_places)}%'
            elif format_type == 'date':
                cell.number_format = column_config.get('format_string', 'dd/mm/yyyy')

            # Set alignment
            if format_type in ['integer', 'number', 'percentage']:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        except Exception as e:
            self.logger.error(f"Error applying column style: {e}")

    def _copy_template_row(self, sheet, source_row: int, target_row: int, row_count: int = 1):
        """Copy template row dengan formatting dan debug"""
        try:
            self.logger.debug(f"Copying template row {source_row} to {target_row} ({row_count} rows)")

            for row_offset in range(row_count):
                source_row_num = source_row + row_offset
                target_row_num = target_row + row_offset

                for col in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(row=source_row_num, column=col)
                    target_cell = sheet.cell(row=target_row_num, column=col)

                    # Copy value
                    target_cell.value = source_cell.value

                    # Copy style
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()

        except Exception as e:
            self.logger.error(f"Error copying template row: {e}")

    def save_processed_template(self, output_path: str) -> bool:
        """Save processed template ke file baru dengan debug"""
        try:
            self.logger.info(f"Saving processed template to: {output_path}")
            self.workbook.save(output_path)
            self.logger.info(f"Template saved successfully: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Error saving processed template: {e}", exc_info=True)
            return False

    def create_copy(self) -> 'TemplateProcessorEnhanced':
        """Create copy dari template processor"""
        try:
            self.logger.debug("Creating copy of template processor")

            # Ensure template_path is a string
            template_path_str = str(self.template_path) if hasattr(self.template_path, '__str__') else self.template_path

            # Create new instance
            new_processor = TemplateProcessorEnhanced(template_path_str, self.formula_path)

            # Copy workbook from current instance
            new_processor.workbook = openpyxl.load_workbook(template_path_str)

            self.logger.debug("Template processor copy created successfully")
            return new_processor

        except Exception as e:
            self.logger.error(f"Error creating template processor copy: {e}")
            raise

    def get_template_info(self) -> Dict[str, Any]:
        """Get template information dengan debug"""
        try:
            info = {
                'template_path': self.template_path,
                'formula_path': self.formula_path,
                'sheets': self.workbook.sheetnames,
                'total_placeholders': sum(len(p) for p in self.placeholders.values()),
                'repeating_sections': list(self.repeating_sections.keys()),
                'formulas_loaded': len(self.formulas) > 0
            }

            self.logger.debug(f"Template info: {info}")
            return info

        except Exception as e:
            self.logger.error(f"Error getting template info: {e}")
            return {'error': str(e)}