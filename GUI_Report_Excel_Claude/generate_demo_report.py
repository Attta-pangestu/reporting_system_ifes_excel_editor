#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Demo Report dengan Sample Data
Menunjukkan bagaimana Adaptive Report Generator bekerja dengan real template
"""

import os
import sys
from datetime import datetime, timedelta
import random

# Add current directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    from adaptive_excel_processor import AdaptiveExcelProcessor
except ImportError as e:
    print(f"Import Error: {e}")
    sys.exit(1)

def generate_demo_report():
    """Generate demo report dengan sample data realistis"""
    print("=" * 80)
    print("DEMO REPORT GENERATION - ADAPTIVE EXCEL PROCESSOR")
    print("=" * 80)

    # Configuration
    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "templates",
        "Template_Laporan_FFB_Analysis.xlsx"
    )

    # Output path
    output_dir = os.path.expanduser("~/Desktop")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f"Demo_Adaptive_Report_{timestamp}.xlsx")

    print(f"Template: {template_path}")
    print(f"Output: {output_path}")
    print("-" * 80)

    try:
        # Step 1: Test Template Analysis
        print("Step 1: Analyzing Template...")
        if not os.path.exists(template_path):
            raise Exception(f"Template not found: {template_path}")

        processor = AdaptiveExcelProcessor(template_path, debug_mode=True)
        template_summary = processor.get_template_summary()

        print(f"[OK] Template analyzed successfully")
        print(f"   Total placeholders: {template_summary['capabilities']['total_placeholders']}")
        print(f"   Sheets with placeholders: {template_summary['capabilities']['sheets_with_placeholders']}")
        print(f"   Supports dynamic tables: {template_summary['capabilities']['supports_dynamic_tables']}")

        # Step 2: Generate Realistic Sample Data
        print("\nStep 2: Generating Sample Data...")

        # Generate date range for September 2025
        start_date = datetime(2025, 9, 1)
        end_date = datetime(2025, 9, 30)

        # Sample employee names (realistic Indonesian names)
        employee_names = [
            "Ahmad Sutrisno", "Budi Santoso", "Chandra Wijaya", "Dedi Kurniawan",
            "Eko Prasetyo", "Fajar Setiawan", "Gunawan Wijaya", "Hadi Susilo",
            "Indra Permadi", "Joko Widodo", "Kartika Sari", "Lukman Hakim",
            "Muhammad Rizki", "Nurul Hidayah", "Oman Sutanto", "Prihadi Utama",
            "Qoriah Aminah", "Rudi Hartono", "Siti Nurjanah", "Toni Sugianto",
            "Usman Hidayat", "Vera Lestari", "Wahyu Hidayat", "Yuliana Susanti"
        ]

        # Sample field names
        field_names = [
            "BLOCK A01", "BLOCK A02", "BLOCK A03", "BLOCK B01", "BLOCK B02",
            "BLOCK B03", "BLOCK C01", "BLOCK C02", "BLOCK C03", "BLOCK D01",
            "BLOCK D02", "BLOCK E01", "BLOCK E02", "BLOCK F01", "BLOCK F02"
        ]

        # Generate daily performance data
        daily_performance = []
        current_date = start_date
        while current_date <= end_date:
            # Simulate daily transaction data
            base_transactions = random.randint(40, 80)

            daily_performance.append({
                'TRANSDATE': current_date.strftime('%Y-%m-%d'),
                'JUMLAH_TRANSAKSI': base_transactions,
                'RIPE_BUNCHES': int(base_transactions * random.uniform(0.7, 0.9)),
                'UNRIPE_BUNCHES': int(base_transactions * random.uniform(0.1, 0.3)),
                'BLACK_BUNCHES': int(base_transactions * random.uniform(0.02, 0.08)),
                'ROTTEN_BUNCHES': int(base_transactions * random.uniform(0.01, 0.05)),
                'LONGSTALK_BUNCHES': int(base_transactions * random.uniform(0.01, 0.04)),
                'RAT_DAMAGE_BUNCHES': int(base_transactions * random.uniform(0.01, 0.03)),
                'LOOSE_FRUIT_KG': round(random.uniform(15, 45), 1)
            })

            current_date += timedelta(days=1)

        # Generate employee performance data
        employee_performance = []

        # Kerani (PM) - 8 employees
        for i in range(8):
            emp_name = employee_names[i]
            base_transactions = random.randint(150, 250)

            employee_performance.append({
                'EMPLOYEE_NAME': emp_name,
                'RECORDTAG': 'KERANI',
                'JUMLAH_TRANSAKSI': base_transactions,
                'TOTAL_RIPE': int(base_transactions * random.uniform(0.7, 0.9)),
                'TOTAL_UNRIPE': int(base_transactions * random.uniform(0.1, 0.3)),
                'TOTAL_BLACK': int(base_transactions * random.uniform(0.02, 0.08)),
                'TOTAL_ROTTEN': int(base_transactions * random.uniform(0.01, 0.05)),
                'TOTAL_LONGSTALK': int(base_transactions * random.uniform(0.01, 0.04)),
                'TOTAL_RATDAMAGE': int(base_transactions * random.uniform(0.01, 0.03)),
                'TOTAL_LOOSEFRUIT': round(random.uniform(300, 800), 1)
            })

        # Mandor (P1) - 6 employees
        for i in range(6):
            emp_name = employee_names[8 + i]
            base_transactions = random.randint(120, 200)

            employee_performance.append({
                'EMPLOYEE_NAME': emp_name,
                'RECORDTAG': 'MANDOR',
                'JUMLAH_TRANSAKSI': base_transactions,
                'TOTAL_RIPE': int(base_transactions * random.uniform(0.7, 0.9)),
                'TOTAL_UNRIPE': int(base_transactions * random.uniform(0.1, 0.3)),
                'TOTAL_BLACK': int(base_transactions * random.uniform(0.02, 0.08)),
                'TOTAL_ROTTEN': int(base_transactions * random.uniform(0.01, 0.05)),
                'TOTAL_LONGSTALK': int(base_transactions * random.uniform(0.01, 0.04)),
                'TOTAL_RATDAMAGE': int(base_transactions * random.uniform(0.01, 0.03)),
                'TOTAL_LOOSEFRUIT': round(random.uniform(200, 600), 1)
            })

        # Asisten (P5) - 4 employees
        for i in range(4):
            emp_name = employee_names[14 + i]
            base_transactions = random.randint(80, 150)

            employee_performance.append({
                'EMPLOYEE_NAME': emp_name,
                'RECORDTAG': 'ASISTEN',
                'JUMLAH_TRANSAKSI': base_transactions,
                'TOTAL_RIPE': int(base_transactions * random.uniform(0.7, 0.9)),
                'TOTAL_UNRIPE': int(base_transactions * random.uniform(0.1, 0.3)),
                'TOTAL_BLACK': int(base_transactions * random.uniform(0.02, 0.08)),
                'TOTAL_ROTTEN': int(base_transactions * random.uniform(0.01, 0.05)),
                'TOTAL_LONGSTALK': int(base_transactions * random.uniform(0.01, 0.04)),
                'TOTAL_RATDAMAGE': int(base_transactions * random.uniform(0.01, 0.03)),
                'TOTAL_LOOSEFRUIT': round(random.uniform(150, 400), 1)
            })

        # Generate field performance data
        field_performance = []
        for field_name in field_names:
            base_transactions = random.randint(200, 400)

            field_performance.append({
                'FIELDNAME': field_name,
                'JUMLAH_TRANSAKSI': base_transactions,
                'TOTAL_RIPE': int(base_transactions * random.uniform(0.7, 0.9)),
                'TOTAL_UNRIPE': int(base_transactions * random.uniform(0.1, 0.3)),
                'TOTAL_BLACK': int(base_transactions * random.uniform(0.02, 0.08)),
                'TOTAL_ROTTEN': int(base_transactions * random.uniform(0.01, 0.05)),
                'TOTAL_LONGSTALK': int(base_transactions * random.uniform(0.01, 0.04)),
                'TOTAL_RATDAMAGE': int(base_transactions * random.uniform(0.01, 0.03)),
                'TOTAL_LOOSEFRUIT': round(random.uniform(400, 900), 1)
            })

        # Generate quality analysis data
        quality_analysis = []
        current_date = start_date
        while current_date <= end_date:
            # Sample some days for quality analysis
            if random.random() > 0.3:  # 70% of days have quality data
                total_ripe = random.randint(100, 200)
                total_defect = int(total_ripe * random.uniform(0.05, 0.15))

                quality_analysis.append({
                    'TRANSDATE': current_date.strftime('%Y-%m-%d'),
                    'TOTAL_RIPE': total_ripe,
                    'TOTAL_BLACK': int(total_defect * random.uniform(0.3, 0.5)),
                    'TOTAL_ROTTEN': int(total_defect * random.uniform(0.2, 0.4)),
                    'TOTAL_RATDAMAGE': int(total_defect * random.uniform(0.1, 0.3)),
                    'PERCENTAGE_DEFECT': round((total_defect / total_ripe * 100), 2) if total_ripe > 0 else 0
                })

            current_date += timedelta(days=1)

        print(f"   Generated sample data:")
        print(f"   - Daily performance: {len(daily_performance)} days")
        print(f"   - Employee performance: {len(employee_performance)} employees")
        print(f"   - Field performance: {len(field_performance)} fields")
        print(f"   - Quality analysis: {len(quality_analysis)} records")

        # Step 3: Calculate Summary Variables
        print("\nStep 3: Calculating Summary Variables...")

        total_transactions = sum(day['JUMLAH_TRANSAKSI'] for day in daily_performance)
        total_ripe = sum(day['RIPE_BUNCHES'] for day in daily_performance)
        total_unripe = sum(day['UNRIPE_BUNCHES'] for day in daily_performance)

        variables = {
            # Report metadata
            'report_title': 'LAPORAN ANALISIS FRESH FRUIT BUNCH (FFB)',
            'estate_name': 'PGE 2B DEMO',
            'report_period': f"1 September 2025 - 30 September 2025",
            'generated_date': datetime.now().strftime('%d %B %Y'),
            'generated_time': datetime.now().strftime('%H:%M:%S'),

            # Dashboard summaries
            'total_transactions': total_transactions,
            'total_ripe_bunches': total_ripe,
            'total_unripe_bunches': total_unripe,
            'avg_ripe_per_transaction': round(total_ripe / total_transactions, 2) if total_transactions > 0 else 0,
            'quality_percentage': round(88.5, 2),  # Simulated quality percentage
            'verification_rate': round(92.3, 2),  # Simulated verification rate
            'daily_average_transactions': round(total_transactions / len(daily_performance), 2),
            'daily_average_ripe': round(total_ripe / len(daily_performance), 2),

            # Peak and low performance days
            'peak_performance_day': max(daily_performance, key=lambda x: x['JUMLAH_TRANSAKSI'])['TRANSDATE'],
            'low_performance_day': min(daily_performance, key=lambda x: x['JUMLAH_TRANSAKSI'])['TRANSDATE']
        }

        print("   Summary variables calculated:")
        for key, value in variables.items():
            if 'total' in key or 'rate' in key or 'average' in key:
                print(f"     {key}: {value}")

        # Step 4: Prepare Adaptive Data Structure
        print("\nStep 4: Preparing Adaptive Data Structure...")
        adaptive_data = {
            # Static variables
            **variables,

            # Dynamic data for repeating sections
            'Dashboard': [variables],  # Summary data for dashboard
            'Harian': daily_performance,
            'Karyawan': employee_performance,
            'Field': field_performance,
            'Kualitas': quality_analysis
        }

        print(f"[OK] Adaptive data structure prepared with {len(adaptive_data)} keys")

        # Step 5: Generate Adaptive Report
        print("\nStep 5: Generating Adaptive Report...")
        success = processor.generate_report(adaptive_data, output_path)

        if success:
            print("[OK] Report generated successfully!")
            print(f"   Output file: {output_path}")

            # Check file size
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"   File size: {file_size:,} bytes")

                # List sheets in the generated file
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(output_path)
                    print(f"   Sheets: {wb.sheetnames}")

                    # Check data in each sheet
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        max_row = ws.max_row
                        max_col = ws.max_column

                        # Count non-empty cells
                        non_empty_cells = 0
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value is not None:
                                    non_empty_cells += 1

                        print(f"   {sheet_name}: {max_row} rows x {max_col} columns ({non_empty_cells} non-empty cells)")

                        # Show sample data for key sheets
                        if sheet_name == 'Karyawan' and max_row > 1:
                            print(f"      Sample employees from {sheet_name}:")
                            for row in range(2, min(5, max_row)):  # Skip header row
                                emp_cell = ws.cell(row=row, column=1)  # Assuming employee name in column A
                                if emp_cell.value:
                                    tag_cell = ws.cell(row=row, column=2)  # Assuming record tag in column B
                                    trans_cell = ws.cell(row=row, column=3)  # Assuming transactions in column C
                                    print(f"        {emp_cell.value} - {tag_cell.value} - {trans_cell.value} transactions")

                        elif sheet_name == 'Harian' and max_row > 1:
                            print(f"      Sample daily data from {sheet_name}:")
                            for row in range(2, min(5, max_row)):
                                date_cell = ws.cell(row=row, column=1)
                                trans_cell = ws.cell(row=row, column=2)
                                if date_cell.value:
                                    print(f"        {date_cell.value}: {trans_cell.value} transactions")

                except Exception as e:
                    print(f"   Could not read sheet info: {e}")

            # Ask user if they want to open the file
            try:
                user_input = input("\nDo you want to open the generated demo report? (y/n): ").strip().lower()
                if user_input in ['y', 'yes']:
                    os.startfile(output_path)
                    print("[OK] Demo report opened in Excel")
            except Exception as e:
                print(f"Could not open file automatically: {e}")
                print(f"Please open manually: {output_path}")

        else:
            raise Exception("Report generation failed")

        print("\n" + "=" * 80)
        print("DEMO REPORT GENERATION COMPLETED SUCCESSFULLY!")
        print("The demo shows how the adaptive system works with real template!")
        print("=" * 80)

        return True

    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function"""
    print("Starting Demo Report Generation...")

    # Check if required files exist
    required_files = [
        "templates/Template_Laporan_FFB_Analysis.xlsx",
        "adaptive_excel_processor.py"
    ]

    missing_files = []
    for file_path in required_files:
        if not os.path.exists(file_path):
            missing_files.append(file_path)

    if missing_files:
        print("[ERROR] Missing required files:")
        for file_path in missing_files:
            print(f"   - {file_path}")
        return

    # Run demo
    success = generate_demo_report()

    if success:
        print("\n[SUCCESS] Demo completed successfully!")
        print("Your demo report is ready on Desktop!")
        print("\nThis demonstrates how the Adaptive Report Generator works:")
        print("1. Analyzes Excel template structure automatically")
        print("2. Detects repeating sections and tables")
        print("3. Fills placeholders with dynamic data")
        print("4. Preserves all Excel formatting")
        print("5. Handles employee names and other repeating data")
    else:
        print("\n[FAILED] Demo failed. Please check the error messages above.")

if __name__ == "__main__":
    main()