"""
Template Processor Engine
Menangani pembacaan template Excel dan pemrosesan placeholder variables
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
import json
import os
from typing import Dict, List, Any, Tuple, Optional
import pandas as pd
from datetime import datetime, date
import logging

class TemplateProcessor:
    """
    Template processor untuk mengelola Excel template dengan placeholder variables
    """

    def __init__(self, template_path: str, formula_path: str):
        """
        Inisialisasi template processor

        Args:
            template_path: Path ke file template Excel
            formula_path: Path ke file formula definition (JSON)
        """
        self.template_path = template_path
        self.formula_path = formula_path
        self.workbook = None
        self.formulas = {}
        self.placeholders = {}
        self.repeating_sections = {}

        # Setup logging
        self.logger = logging.getLogger(__name__)

        # Load template dan formulas
        self._load_template()
        self._load_formulas()
        self._scan_placeholders()
        self._parse_repeating_sections()

    def _load_template(self):
        """Load template Excel"""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file tidak ditemukan: {self.template_path}")

        self.workbook = openpyxl.load_workbook(self.template_path)
        self.logger.info(f"Template dimuat: {self.template_path}")
        self.logger.info(f"Sheets tersedia: {self.workbook.sheetnames}")

    def _load_formulas(self):
        """Load formula definitions dari file JSON"""
        if not os.path.exists(self.formula_path):
            raise FileNotFoundError(f"Formula file tidak ditemukan: {self.formula_path}")

        with open(self.formula_path, 'r', encoding='utf-8') as f:
            self.formulas = json.load(f)

        self.logger.info(f"Formula definitions loaded dari: {self.formula_path}")

    def _scan_placeholders(self):
        """Scan template untuk menemukan semua placeholder variables"""
        placeholder_pattern = r'\{\{([^}]+)\}\}'

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_placeholders = []

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        matches = re.findall(placeholder_pattern, cell.value)
                        for match in matches:
                            placeholder_info = {
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'placeholder': match,
                                'full_placeholder': f'{{{{{match}}}}}',
                                'original_value': cell.value
                            }
                            sheet_placeholders.append(placeholder_info)

            if sheet_placeholders:
                self.placeholders[sheet_name] = sheet_placeholders

        self.logger.info(f"Found {sum(len(p) for p in self.placeholders.values())} placeholders")

    def _parse_repeating_sections(self):
        """Parse repeating sections dari formula definitions"""
        repeating_config = self.formulas.get('repeating_sections', {})

        for section_name, section_config in repeating_config.items():
            sheet_name = section_config.get('sheet_name')
            if not sheet_name:
                continue

            self.repeating_sections[sheet_name] = {
                'name': section_name,
                'data_source': section_config.get('data_source'),
                'start_row': section_config.get('start_row', 1),
                'template_rows': section_config.get('template_rows', 1),
                'columns': section_config.get('columns', {}),
                'group_by': section_config.get('group_by')
            }

        self.logger.info(f"Parsed {len(self.repeating_sections)} repeating sections")

    def get_sheet_placeholders(self, sheet_name: str) -> List[Dict]:
        """Get placeholders untuk sheet tertentu"""
        return self.placeholders.get(sheet_name, [])

    def get_placeholders(self) -> Dict[str, List[Dict]]:
        """Get all placeholders from all sheets"""
        return self.placeholders

    def get_repeating_section(self, sheet_name: str) -> Optional[Dict]:
        """Get repeating section configuration untuk sheet tertentu"""
        return self.repeating_sections.get(sheet_name)

    def process_sheet_placeholders(self, sheet_name: str, data: Dict[str, Any]) -> bool:
        """
        Process placeholders pada sheet tertentu dengan data yang disediakan

        Args:
            sheet_name: Nama sheet yang akan diproses
            data: Data untuk substitusi placeholder

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.logger.error(f"Sheet {sheet_name} tidak ditemukan")
                return False

            sheet = self.workbook[sheet_name]
            placeholders = self.get_sheet_placeholders(sheet_name)

            for placeholder_info in placeholders:
                placeholder = placeholder_info['placeholder']
                cell_coord = placeholder_info['cell']
                cell = sheet[cell_coord]

                # Get value untuk placeholder
                value = self._get_placeholder_value(placeholder, data)

                # Apply formatting jika diperlukan
                value = self._apply_formatting(placeholder, value, data)

                # Set cell value
                cell.value = value

                # Apply cell styling
                self._apply_cell_style(cell, placeholder, value)

            self.logger.info(f"Processed {len(placeholders)} placeholders di sheet {sheet_name}")
            return True

        except Exception as e:
            self.logger.error(f"Error processing sheet {sheet_name}: {e}")
            return False

    def process_repeating_section(self, sheet_name: str, data: List[Dict[str, Any]]) -> bool:
        """
        Process repeating section pada sheet dengan data array

        Args:
            sheet_name: Nama sheet
            data: Array data untuk repeating section

        Returns:
            True jika berhasil, False jika gagal
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.logger.error(f"Sheet {sheet_name} tidak ditemukan")
                return False

            repeating_section = self.get_repeating_section(sheet_name)
            if not repeating_section:
                self.logger.warning(f"No repeating section found for sheet {sheet_name}")
                return False

            sheet = self.workbook[sheet_name]
            start_row = repeating_section['start_row']
            template_rows = repeating_section['template_rows']
            columns = repeating_section['columns']

            # Delete existing template rows (keep one for template)
            sheet.delete_rows(start_row + template_rows, sheet.max_row - start_row - template_rows + 1)

            # Process each data row
            for i, row_data in enumerate(data):
                target_row = start_row + i

                # Copy template row if needed
                if i > 0:
                    self._copy_template_row(sheet, start_row, target_row, template_rows)

                # Fill data in columns
                for col_letter, column_config in columns.items():
                    field_name = column_config.get('field')
                    if field_name and field_name in row_data:
                        cell = sheet[f"{col_letter}{target_row}"]
                        value = row_data[field_name]

                        # Apply formatting
                        value = self._apply_column_formatting(value, column_config)

                        # Set value
                        cell.value = value

                        # Apply styling
                        self._apply_column_style(cell, column_config)

            self.logger.info(f"Processed {len(data)} rows in repeating section for sheet {sheet_name}")
            return True

        except Exception as e:
            self.logger.error(f"Error processing repeating section in sheet {sheet_name}: {e}")
            return False

    def _get_placeholder_value(self, placeholder: str, data: Dict[str, Any]) -> Any:
        """Get value untuk placeholder dari data"""
        # Handle complex placeholders (formatting, calculations)
        if '|' in placeholder:
            placeholder = placeholder.split('|')[0]

        # Direct lookup
        if placeholder in data:
            return data[placeholder]

        # Nested lookup (dot notation)
        if '.' in placeholder:
            keys = placeholder.split('.')
            value = data
            for key in keys:
                if isinstance(value, dict) and key in value:
                    value = value[key]
                else:
                    return None
            return value

        # Default values
        defaults = {
            'current_date': datetime.now().strftime('%d %B %Y'),
            'current_time': datetime.now().strftime('%H:%M:%S'),
            'report_title': 'LAPORAN ANALISIS FFB',
            'report_period': 'Periode Laporan'
        }

        return defaults.get(placeholder, f'{{{placeholder}}}')

    def _apply_formatting(self, placeholder: str, value: Any, data: Dict[str, Any]) -> Any:
        """Apply formatting ke value"""
        if '|' in placeholder:
            format_spec = placeholder.split('|', 1)[1]
            if format_spec.startswith('format:'):
                format_string = format_spec[7:]  # Remove 'format:'
                if isinstance(value, (datetime, date)):
                    return value.strftime(format_string)
                elif isinstance(value, (int, float)):
                    return format(value, format_string)

        return value

    def _apply_cell_style(self, cell, placeholder: str, value: Any):
        """Apply styling ke cell berdasarkan value"""
        # Remove placeholder styling
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == 'FFFF5F5':
            cell.fill = PatternFill(fill_type=None)

        if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF6B6B':
            cell.font = cell.font.copy(color='000000')

        if cell.font and cell.font.italic:
            cell.font = cell.font.copy(italic=False)

        # Apply numeric formatting
        if isinstance(value, (int, float)):
            if '.' in str(value):
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '#,##0'

    def _apply_column_formatting(self, value: Any, column_config: Dict) -> Any:
        """Apply column formatting ke value"""
        format_type = column_config.get('format', 'text')

        if format_type == 'date' and value:
            if isinstance(value, str):
                try:
                    value = datetime.strptime(value, '%Y-%m-%d')
                except:
                    pass

            if isinstance(value, (datetime, date)):
                format_string = column_config.get('format_string', 'dd/mm/yyyy')
                return value.strftime(format_string)

        elif format_type == 'integer' and value:
            try:
                return int(float(value))
            except:
                return value

        elif format_type == 'number' and value:
            try:
                decimal_places = column_config.get('decimal_places', 2)
                return round(float(value), decimal_places)
            except:
                return value

        elif format_type == 'percentage' and value:
            try:
                return float(value) / 100  # Excel stores percentage as decimal
            except:
                return value

        # Handle mapping
        mapping = column_config.get('mapping', {})
        if mapping and value in mapping:
            return mapping[value]

        return value

    def _apply_column_style(self, cell, column_config: Dict):
        """Apply column styling ke cell"""
        format_type = column_config.get('format', 'text')

        # Set number format
        if format_type == 'integer':
            cell.number_format = '#,##0'
        elif format_type == 'number':
            decimal_places = column_config.get('decimal_places', 2)
            cell.number_format = f'#,##0.{"".join(["0"] * decimal_places)}'
        elif format_type == 'percentage':
            decimal_places = column_config.get('decimal_places', 2)
            cell.number_format = f'0.{"".join(["0"] * decimal_places)}%'
        elif format_type == 'date':
            cell.number_format = column_config.get('format_string', 'dd/mm/yyyy')

        # Set alignment
        if format_type in ['integer', 'number', 'percentage']:
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    def _copy_template_row(self, sheet, source_row: int, target_row: int, row_count: int = 1):
        """Copy template row dengan formatting"""
        for row_offset in range(row_count):
            source_row_num = source_row + row_offset
            target_row_num = target_row + row_offset

            for col in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=source_row_num, column=col)
                target_cell = sheet.cell(row=target_row_num, column=col)

                # Copy value
                target_cell.value = source_cell.value

                # Copy style
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()

    def save_processed_template(self, output_path: str) -> bool:
        """Save processed template ke file baru"""
        try:
            self.workbook.save(output_path)
            self.logger.info(f"Processed template saved: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Error saving processed template: {e}")
            return False

    def create_copy(self) -> 'TemplateProcessor':
        """Create copy dari template processor"""
        # Ensure template_path is a string, not a TemplateProcessor object
        template_path_str = str(self.template_path) if hasattr(self.template_path, '__str__') else self.template_path

        # Create new instance
        new_processor = TemplateProcessor(template_path_str, self.formula_path)

        # Copy workbook from current instance
        new_processor.workbook = openpyxl.load_workbook(template_path_str)

        return new_processor

    def get_template_info(self) -> Dict[str, Any]:
        """Get template information"""
        return {
            'template_path': self.template_path,
            'formula_path': self.formula_path,
            'sheets': self.workbook.sheetnames,
            'total_placeholders': sum(len(p) for p in self.placeholders.values()),
            'repeating_sections': list(self.repeating_sections.keys()),
            'formulas_loaded': len(self.formulas) > 0
        }