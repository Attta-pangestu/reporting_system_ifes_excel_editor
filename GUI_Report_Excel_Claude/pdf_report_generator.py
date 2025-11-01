"""
PDF Report Generator
Mengkonversi Excel report ke format PDF dengan professional formatting
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
import tempfile
import shutil

# Try to import pdfkit and wkhtmltopdf
try:
    import pdfkit
    PDFKIT_AVAILABLE = True
except ImportError:
    PDFKIT_AVAILABLE = False
    logging.warning("pdfkit not available, PDF export will be limited")

# Try to import reportlab for direct PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("reportlab not available, using alternative PDF generation")

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

class PDFReportGenerator:
    """
    Generator untuk membuat PDF report dari Excel template dan data
    """

    def __init__(self):
        """Initialize PDF report generator"""
        self.logger = logging.getLogger(__name__)

    def generate_pdf_report(self, excel_file_path: str, output_path: str = None) -> str:
        """
        Generate PDF report from Excel file

        Args:
            excel_file_path: Path ke Excel file yang sudah di-generate
            output_path: Path untuk output PDF (optional)

        Returns:
            Path ke generated PDF file
        """
        try:
            if not os.path.exists(excel_file_path):
                raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

            # Generate output filename
            if output_path is None:
                base_name = os.path.splitext(os.path.basename(excel_file_path))[0]
                output_path = os.path.join(
                    os.path.dirname(excel_file_path),
                    f"{base_name}.pdf"
                )

            self.logger.info(f"Generating PDF from: {excel_file_path}")
            self.logger.info(f"Output PDF: {output_path}")

            # Try different PDF generation methods
            if PDFKIT_AVAILABLE:
                return self._generate_pdf_with_pdfkit(excel_file_path, output_path)
            elif REPORTLAB_AVAILABLE:
                return self._generate_pdf_with_reportlab(excel_file_path, output_path)
            else:
                return self._generate_pdf_simple(excel_file_path, output_path)

        except Exception as e:
            self.logger.error(f"Error generating PDF: {e}")
            raise

    def _generate_pdf_with_pdfkit(self, excel_file_path: str, output_path: str) -> str:
        """Generate PDF using pdfkit (requires wkhtmltopdf)"""
        try:
            # Convert Excel to HTML first (simplified approach)
            html_content = self._excel_to_html(excel_file_path)

            # Convert HTML to PDF
            options = {
                'page-size': 'A4',
                'margin-top': '0.75in',
                'margin-right': '0.75in',
                'margin-bottom': '0.75in',
                'margin-left': '0.75in',
                'encoding': "UTF-8",
                'no-outline': None
            }

            pdfkit.from_string(html_content, output_path, options=options)
            self.logger.info(f"PDF generated successfully with pdfkit: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Error generating PDF with pdfkit: {e}")
            # Fallback to reportlab
            if REPORTLAB_AVAILABLE:
                return self._generate_pdf_with_reportlab(excel_file_path, output_path)
            else:
                return self._generate_pdf_simple(excel_file_path, output_path)

    def _generate_pdf_with_reportlab(self, excel_file_path: str, output_path: str) -> str:
        """Generate PDF using reportlab"""
        try:
            # Read Excel data
            workbook = openpyxl.load_workbook(excel_file_path)
            story = []

            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1,  # Center
                textColor=colors.darkblue
            )

            # Add title
            title = "LAPORAN ANALISIS FFB (FRESH FRUIT BUNCH)"
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))

            # Process each sheet
            for sheet_name in workbook.sheetnames:
                if sheet_name == 'Dashboard':
                    self._add_dashboard_section(workbook[sheet_name], story, styles)
                elif sheet_name in ['Harian', 'Karyawan', 'Field', 'Kualitas']:
                    self._add_table_section(workbook[sheet_name], story, styles)

                story.append(Spacer(1, 20))

            # Build PDF
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            doc.build(story)

            self.logger.info(f"PDF generated successfully with reportlab: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Error generating PDF with reportlab: {e}")
            return self._generate_pdf_simple(excel_file_path, output_path)

    def _add_dashboard_section(self, sheet, story, styles):
        """Add dashboard section to PDF"""
        # Create dashboard summary
        data = []

        # Read data from dashboard sheet
        for row in sheet.iter_rows(min_row=1, max_row=25, min_col=1, max_col=10):
            row_data = []
            has_data = False
            for cell in row:
                if cell.value:
                    row_data.append(str(cell.value))
                    has_data = True
                else:
                    row_data.append("")

            if has_data:
                data.append(row_data)

        if data:
            # Create table for dashboard
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)

    def _add_table_section(self, sheet, story, styles):
        """Add table section to PDF"""
        # Add section title
        section_title = sheet.title.replace('_', ' ').upper()
        title_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        story.append(Paragraph(section_title, title_style))

        # Read data from sheet
        data = []
        headers = []

        # Get headers (first row)
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                headers.append("")

        if headers:
            data.append(headers)

        # Get data rows
        for row in sheet.iter_rows(min_row=2, max_row=50, min_col=1, max_col=len(headers)+1):
            row_data = []
            has_data = False
            for cell in row:
                if cell.value:
                    row_data.append(str(cell.value))
                    has_data = True
                else:
                    row_data.append("")

            if has_data:
                data.append(row_data)

        if len(data) > 1:  # Has headers and data
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8)
            ]))
            story.append(table)

    def _excel_to_html(self, excel_file_path: str) -> str:
        """Convert Excel to HTML (simplified)"""
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            html_parts = []

            html_parts.append("""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Laporan Analisis FFB</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; }
                    h1 { color: #2c3e50; text-align: center; }
                    table { border-collapse: collapse; width: 100%; margin: 20px 0; }
                    th { background-color: #3498db; color: white; padding: 8px; text-align: left; }
                    td { border: 1px solid #ddd; padding: 8px; }
                    tr:nth-child(even) { background-color: #f2f2f2; }
                    .section { margin: 30px 0; }
                </style>
            </head>
            <body>
                <h1>LAPORAN ANALISIS FFB (FRESH FRUIT BUNCH)</h1>
            """)

            for sheet_name in workbook.sheetnames:
                html_parts.append(f'<div class="section">')
                html_parts.append(f'<h2>{sheet_name.replace("_", " ").upper()}</h2>')

                sheet = workbook[sheet_name]

                # Create table
                html_parts.append('<table>')

                # Headers
                html_parts.append('<tr>')
                for cell in sheet[1]:
                    if cell.value:
                        html_parts.append(f'<th>{cell.value}</th>')
                html_parts.append('</tr>')

                # Data rows
                for row in sheet.iter_rows(min_row=2, max_row=20, min_col=1, max_col=20):
                    has_data = False
                    row_html = ['<tr>']
                    for cell in row:
                        if cell.value:
                            row_html.append(f'<td>{cell.value}</td>')
                            has_data = True
                        else:
                            row_html.append('<td></td>')

                    if has_data:
                        row_html.append('</tr>')
                        html_parts.append(''.join(row_html))

                html_parts.append('</table>')
                html_parts.append('</div>')

            html_parts.append('</body></html>')

            return ''.join(html_parts)

        except Exception as e:
            self.logger.error(f"Error converting Excel to HTML: {e}")
            return f"<html><body><h1>Error generating report: {e}</h1></body></html>"

    def _generate_pdf_simple(self, excel_file_path: str, output_path: str) -> str:
        """Generate PDF using simple approach"""
        try:
            # Read Excel data and create simple text-based PDF
            workbook = openpyxl.load_workbook(excel_file_path)

            # Create a simple text report first
            text_report = self._create_text_report(workbook)

            # Convert text to PDF (this is a simplified approach)
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as temp_file:
                temp_file.write(text_report)

                # Use system print to PDF if available (this is a fallback)
                try:
                    import win32print
                    import win32api

                    # Print to PDF using Windows API
                    printer_info = win32print.GetPrinter(win32print.GetDefaultPrinter())
                    printer_name = printer_info[2]  # Get printer name

                    win32api.ShellExecute(
                        0, "printto", temp_file.name,
                        f'"{printer_name}"', "/D:PDFcreator", 0
                    )

                except ImportError:
                    # Fallback: Save as text file
                    shutil.copy2(temp_file.name, output_path.replace('.pdf', '.txt'))
                    output_path = output_path.replace('.pdf', '.txt')

            return output_path

        except Exception as e:
            self.logger.error(f"Error generating simple PDF: {e}")
            raise

    def _create_text_report(self, workbook) -> str:
        """Create text report from Excel workbook"""
        lines = []
        lines.append("=" * 80)
        lines.append("LAPORAN ANALISIS FFB (FRESH FRUIT BUNCH)")
        lines.append("=" * 80)
        lines.append(f"Generated on: {datetime.now().strftime('%d %B %Y %H:%M:%S')}")
        lines.append("")

        for sheet_name in workbook.sheetnames:
            lines.append(f"SHEET: {sheet_name.upper()}")
            lines.append("-" * 40)

            sheet = workbook[sheet_name]

            for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), 1):
                row_data = []
                has_data = False
                for cell in row:
                    if cell.value:
                        row_data.append(str(cell.value))
                        has_data = True
                    else:
                        row_data.append("")

                if has_data:
                    line = " | ".join(f"{val:20}" for val in row_data)
                    lines.append(line)

            lines.append("")

        return "\n".join(lines)

    def generate_analysis_report(self, query_results: Dict[str, Any], parameters: Dict[str, Any],
                                output_path: str) -> str:
        """
        Generate analysis PDF report directly from database query results

        Args:
            query_results: Query results from database
            parameters: Report parameters
            output_path: Path untuk output PDF

        Returns:
            Path ke generated PDF file
        """
        try:
            self.logger.info("Generating analysis PDF report from database results...")

            if not REPORTLAB_AVAILABLE:
                raise ImportError("reportlab is required for PDF generation")

            story = []
            styles = getSampleStyleSheet()

            # Create custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1,
                textColor=colors.darkblue
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkgreen
            )

            # Title
            title = "LAPORAN ANALISIS FFB - DATABASE REPORT"
            story.append(Paragraph(title, title_style))

            # Report info
            info_data = [
                ["Parameter", "Value"],
                ["Periode", f"{parameters.get('start_date', 'N/A')} s/d {parameters.get('end_date', 'N/A')}"],
                ["Estate", parameters.get('estate_name', 'N/A')],
                ["Generated", datetime.now().strftime('%d %B %Y %H:%M:%S')]
            ]

            info_table = Table(info_data)
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(info_table)
            story.append(Spacer(1, 20))

            # Analysis sections
            if 'transaction_summary' in query_results and query_results['transaction_summary']:
                self._add_transaction_analysis(query_results['transaction_summary'], story, styles)

            if 'daily_performance' in query_results and query_results['daily_performance']:
                self._add_daily_performance_analysis(query_results['daily_performance'], story, styles)

            if 'field_performance' in query_results and query_results['field_performance']:
                self._add_field_performance_analysis(query_results['field_performance'], story, styles)

            # Build PDF
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            doc.build(story)

            self.logger.info(f"Analysis PDF generated successfully: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Error generating analysis PDF: {e}")
            raise

    def _add_transaction_analysis(self, transaction_data, story, styles):
        """Add transaction analysis section"""
        story.append(Paragraph("TRANSACTION SUMMARY", styles['Heading2']))

        # Extract data from transaction summary
        if isinstance(transaction_data, list) and len(transaction_data) > 0 and transaction_data[0].get('rows'):
            rows = transaction_data[0]['rows']

            data = [["Metric", "Value"]]
            for row in rows:
                for key, value in row.items():
                    if key != 'TOTAL_TRANSAKSI':  # Add all metrics
                        data.append([key.replace('_', ' ').title(), str(value)])

            if len(data) > 1:
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)

        story.append(Spacer(1, 12))

    def _add_daily_performance_analysis(self, daily_data, story, styles):
        """Add daily performance analysis section"""
        story.append(Paragraph("DAILY PERFORMANCE ANALYSIS", styles['Heading2']))

        # Extract daily performance data
        if isinstance(daily_data, list) and len(daily_data) > 0 and daily_data[0].get('rows'):
            rows = daily_data[0]['rows']

            data = [["Date", "Transactions", "Ripe Bunches", "Unripe Bunches"]]

            for row in rows:
                data.append([
                    row.get('TRANSDATE', 'N/A'),
                    row.get('JUMLAH_TRANSAKSI', 'N/A'),
                    row.get('RIPE_BUNCHES', 'N/A'),
                    row.get('UNRIPE_BUNCHES', 'N/A')
                ])

            if len(data) > 1:
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)

        story.append(Spacer(1, 12))

    def _add_field_performance_analysis(self, field_data, story, styles):
        """Add field performance analysis section"""
        story.append(Paragraph("FIELD PERFORMANCE ANALYSIS", styles['Heading2']))

        # Extract field performance data
        if isinstance(field_data, list) and len(field_data) > 0 and field_data[0].get('rows'):
            rows = field_data[0]['rows']

            data = [["Field Name", "Transactions", "Total Ripe", "Total Unripe"]]

            for row in rows:
                data.append([
                    row.get('FIELDNAME', 'N/A'),
                    row.get('JUMLAH_TRANSAKSI', 'N/A'),
                    row.get('TOTAL_RIPE', 'N/A'),
                    row.get('TOTAL_UNRIPE', 'N/A')
                ])

            if len(data) > 1:
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)

        story.append(Spacer(1, 12))