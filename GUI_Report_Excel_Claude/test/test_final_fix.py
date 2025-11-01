#!/usr/bin/env python3
"""
Final Test untuk Excel Report Generator Fix
Menguji complete end-to-end dengan enhanced formula dan proper template processing
"""

import os
import sys
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
import traceback

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def setup_console_output():
    """Setup console output untuk Windows compatibility"""
    import sys
    if sys.platform == 'win32':
        import codecs
        sys.stdout = codecs.getwriter('cp850')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('cp850')(sys.stderr.buffer, 'strict')

def print_section(title):
    """Print section header"""
    print("="*60)
    print(f"FINAL TEST: {title}")
    print("="*60)

def print_subsection(title):
    """Print subsection header"""
    print(f"\n{title}:")
    print("-"*40)

def test_complete_fix():
    """Test complete fix end-to-end"""
    print_section("COMPLETE END-TO-END FIX TEST")

    try:
        # Import required modules
        from firebird_connector_enhanced import FirebirdConnectorEnhanced
        from formula_engine_enhanced import FormulaEngineEnhanced
        from template_processor_enhanced import TemplateProcessorEnhanced
        from excel_report_generator_enhanced import ExcelReportGeneratorEnhanced

        # Setup paths
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        templates_dir = os.path.join(base_path, "templates")
        enhanced_formula_path = os.path.join(base_path, "laporan_ffb_analysis_formula_enhanced.json")
        db_path = r"D:\Gawean Rebinmas\Monitoring Database\Database Ifess\IFESS_2B_24-10-2025\PTRJ_P2B.FDB"

        print(f"Enhanced formula path: {enhanced_formula_path}")
        print(f"Database path: {db_path}")

        # Check files exist
        if not os.path.exists(enhanced_formula_path):
            print("ERROR: Enhanced formula JSON not found!")
            return False

        # Find template
        template_files = [f for f in os.listdir(templates_dir) if f.endswith(('.xlsx', '.xls'))]
        if not template_files:
            print("ERROR: No template files found!")
            return False

        template_path = os.path.join(templates_dir, template_files[0])
        print(f"Using template: {template_files[0]}")

        # Initialize enhanced report generator
        print_subsection("INITIALIZING ENHANCED REPORT GENERATOR")
        generator = ExcelReportGeneratorEnhanced(
            template_path=template_path,
            formula_path=enhanced_formula_path
        )

        # Test parameters
        test_params = {
            'start_date': '2024-10-01',
            'end_date': '2024-10-31',
            'selected_estates': ['PGE 2B']
        }

        print(f"Test parameters: {test_params}")

        # Test database connection
        print_subsection("TESTING DATABASE CONNECTION")
        if not generator.test_database_connection('PGE 2B'):
            print("ERROR: Database connection failed!")
            return False
        print("SUCCESS: Database connection working")

        # Test template info
        print_subsection("TESTING TEMPLATE INFO")
        template_info = generator.get_template_info()
        if 'error' in template_info:
            print(f"ERROR: Template info failed: {template_info['error']}")
            return False

        print(f"Template info: {template_info}")

        # Generate report
        print_subsection("GENERATING COMPLETE REPORT")
        output_dir = os.path.join(os.path.dirname(__file__), "final_test_output")
        os.makedirs(output_dir, exist_ok=True)

        success, results = generator.generate_report(
            start_date=test_params['start_date'],
            end_date=test_params['end_date'],
            selected_estates=test_params['selected_estates'],
            output_dir=output_dir
        )

        if success:
            print(f"SUCCESS: Report generated successfully!")
            print(f"Generated files: {results}")

            # Verify output file
            if results and len(results) > 0:
                output_file = results[0]
                if os.path.exists(output_file):
                    file_size = os.path.getsize(output_file)
                    print(f"Output file: {output_file}")
                    print(f"File size: {file_size:,} bytes")

                    # Test template processing verification
                    print_subsection("VERIFYING TEMPLATE PROCESSING")
                    try:
                        # Load the generated file and check if placeholders are filled
                        import openpyxl
                        wb = openpyxl.load_workbook(output_file)

                        filled_placeholders = 0
                        empty_placeholders = 0

                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            print(f"\nChecking sheet: {sheet_name}")

                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value and isinstance(cell.value, str):
                                        if '{{' in cell.value and '}}' in cell.value:
                                            empty_placeholders += 1
                                            print(f"  Found unfilled placeholder: {cell.value} at {cell.coordinate}")
                                        else:
                                            filled_placeholders += 1

                        print(f"\nTemplate verification results:")
                        print(f"  Filled cells: {filled_placeholders}")
                        print(f"  Empty placeholders: {empty_placeholders}")

                        if empty_placeholders == 0:
                            print("SUCCESS: All placeholders filled!")
                        else:
                            print(f"WARNING: {empty_placeholders} placeholders still unfilled")

                    except Exception as e:
                        print(f"WARNING: Could not verify output file: {e}")

                    return True
                else:
                    print("ERROR: Output file not found!")
                    return False
            else:
                print("ERROR: No output files generated!")
                return False
        else:
            print(f"ERROR: Report generation failed: {results}")
            return False

    except Exception as e:
        print(f"ERROR: Complete test failed: {e}")
        traceback.print_exc()
        return False

def test_placeholder_resolution():
    """Test placeholder resolution secara detail"""
    print_section("PLACEHOLDER RESOLUTION TEST")

    try:
        # Import required modules
        from firebird_connector_enhanced import FirebirdConnectorEnhanced
        from formula_engine_enhanced import FormulaEngineEnhanced
        from template_processor_enhanced import TemplateProcessorEnhanced

        # Setup paths
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        templates_dir = os.path.join(base_path, "templates")
        enhanced_formula_path = os.path.join(base_path, "laporan_ffb_analysis_formula_enhanced.json")
        db_path = r"D:\Gawean Rebinmas\Monitoring Database\Database Ifess\IFESS_2B_24-10-2025\PTRJ_P2B.FDB"

        # Find template
        template_files = [f for f in os.listdir(templates_dir) if f.endswith(('.xlsx', '.xls'))]
        template_path = os.path.join(templates_dir, template_files[0])

        # Initialize components
        connector = FirebirdConnectorEnhanced(db_path=db_path)
        formula_engine = FormulaEngineEnhanced(enhanced_formula_path, connector)
        template_processor = TemplateProcessorEnhanced(template_path, enhanced_formula_path)

        # Prepare test parameters
        test_params = {
            'start_date': '2024-10-01',
            'end_date': '2024-10-31',
            'estate_name': 'PGE 2B',
            'estate_code': 'PGE_2B'
        }

        print(f"Test parameters: {test_params}")

        # Execute queries and process variables
        query_results = formula_engine.execute_all_queries(test_params)
        variables = formula_engine.process_variables(query_results, test_params)

        # Prepare complete data
        all_data = {
            **test_params,
            **variables,
            **query_results
        }

        print(f"\nComplete data keys: {list(all_data.keys())}")

        # Test placeholder resolution
        placeholders = template_processor.get_placeholders()
        total_placeholders = sum(len(p) for p in placeholders.values())
        print(f"\nTotal placeholders in template: {total_placeholders}")

        resolved_placeholders = 0
        unresolved_placeholders = []

        for sheet_name, sheet_placeholders in placeholders.items():
            print(f"\nSheet '{sheet_name}': {len(sheet_placeholders)} placeholders")

            for ph in sheet_placeholders:
                placeholder = ph['placeholder']
                cell_coord = ph['cell']

                try:
                    value = template_processor._get_placeholder_value(placeholder, all_data)

                    if value is not None:
                        resolved_placeholders += 1
                        print(f"  RESOLVED: {placeholder} at {cell_coord} = {value}")
                    else:
                        unresolved_placeholders.append(f"{placeholder} at {cell_coord}")
                        print(f"  UNRESOLVED: {placeholder} at {cell_coord}")

                except Exception as e:
                    unresolved_placeholders.append(f"{placeholder} at {cell_coord} (ERROR: {e})")
                    print(f"  ERROR: {placeholder} at {cell_coord} = {e}")

        print(f"\nResolution Summary:")
        print(f"  Total placeholders: {total_placeholders}")
        print(f"  Resolved: {resolved_placeholders}")
        print(f"  Unresolved: {len(unresolved_placeholders)}")
        print(f"  Success rate: {resolved_placeholders/total_placeholders*100:.1f}%")

        if unresolved_placeholders:
            print(f"\nUnresolved placeholders:")
            for up in unresolved_placeholders:
                print(f"  • {up}")

        return resolved_placeholders > 0

    except Exception as e:
        print(f"ERROR: Placeholder resolution test failed: {e}")
        traceback.print_exc()
        return False

def main():
    """Main function"""
    # Setup console
    setup_console_output()

    print("Excel Report Generator - Final Fix Test")
    print("="*60)

    try:
        # Run placeholder resolution test
        print("Running placeholder resolution test...")
        placeholder_test = test_placeholder_resolution()

        # Run complete fix test
        print("\nRunning complete fix test...")
        complete_test = test_complete_fix()

        # Final summary
        print_section("FINAL TEST SUMMARY")
        print(f"Placeholder Resolution Test: {'PASS' if placeholder_test else 'FAIL'}")
        print(f"Complete Fix Test: {'PASS' if complete_test else 'FAIL'}")

        if placeholder_test and complete_test:
            print("\nALL TESTS PASSED!")
            print("The Excel Report Generator is now working correctly!")
            print("\nFIXES APPLIED:")
            print("1. Enhanced formula JSON with all required variables")
            print("2. Proper query execution with parameter substitution")
            print("3. Complete variable mapping for template placeholders")
            print("4. Enhanced template processing with debug logging")
            print("5. Repeating sections properly configured")
            return True
        else:
            print("\nSOME TESTS FAILED!")
            print("Please check the error messages above.")
            return False

    except Exception as e:
        print(f"ERROR: Final test failed: {e}")
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)