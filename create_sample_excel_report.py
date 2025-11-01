#!/usr/bin/env python3
"""
Create Sample Excel Report dengan data yang realistis sesuai format asli
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os

class SampleExcelReportCreator:
    def __init__(self):
        self.filename = "Laporan_Kinerja_Kerani_Mandor_Asisten_Sample.xlsx"

    def create_sample_report(self):
        """Create sample Excel report with realistic data"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Kinerja"

        # Set column widths
        column_widths = {
            'A': 15, 'B': 20, 'C': 25, 'D': 12, 'E': 18, 'F': 25, 'G': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Create styles
        self.create_styles()

        # Add headers
        headers = [
            "ESTATE", "DIVISI", "KARYAWAN", "ROLE",
            "JUMLAH TRANSAKSI", "PERSENTASE TERVERIFIKASI", "KETERANGAN PERBEDAAN"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        # Sample data based on realistic scenarios
        sample_data = [
            # PGE 1A - DIVISI 1
            ["PGE 1A", "DIVISI 1", "== DIVISI 1 TOTAL ==", "SUMMARY", 1850, "86.05% (1592)", ""],
            ["PGE 1A", "DIVISI 1", "AHMAD RIZKI", "KERANI", 520, "92.31% (480)", "12 perbedaan (2.5%)"],
            ["PGE 1A", "DIVISI 1", "BUDI SANTOSO", "KERANI", 480, "88.54% (425)", "18 perbedaan (4.2%)"],
            ["PGE 1A", "DIVISI 1", "CHANDRA KUSUMA", "KERANI", 450, "84.44% (380)", "25 perbedaan (6.6%)"],
            ["PGE 1A", "DIVISI 1", "DANI PRASTYO", "KERANI", 400, "79.00% (316)", "35 perbedaan (11.1%)"],
            ["PGE 1A", "DIVISI 1", "EKO WIBOWO", "MANDOR", 320, "17.30%", ""],
            ["PGE 1A", "DIVISI 1", "FERY HANDOKO", "ASISTEN", 280, "15.14%", ""],

            # Separator
            ["", "", "", "", "", "", ""],

            # PGE 1A - DIVISI 2
            ["PGE 1A", "DIVISI 2", "== DIVISI 2 TOTAL ==", "SUMMARY", 1420, "83.52% (1186)", ""],
            ["PGE 1A", "DIVISI 2", "GUNAWAN SUSILO", "KERANI", 380, "89.47% (340)", "8 perbedaan (2.4%)"],
            ["PGE 1A", "DIVISI 2", "HENDRA SETIAWAN", "KERANI", 350, "85.71% (300)", "15 perbedaan (5.0%)"],
            ["PGE 1A", "DIVISI 2", "IRFAN BACHDIM", "KERANI", 320, "81.25% (260)", "22 perbedaan (8.5%)"],
            ["PGE 1A", "DIVISI 2", "JOKO SANTOSO", "KERANI", 370, "78.38% (290)", "28 perbedaan (9.7%)"],
            ["PGE 1A", "DIVISI 2", "KARNA SUDARMO", "MANDOR", 280, "19.72%", ""],
            ["PGE 1A", "DIVISI 2", "LUKMAN HAKIM", "ASISTEN", 240, "16.90%", ""],

            # Separator
            ["", "", "", "", "", "", ""],

            # PGE 1B - DIVISI 3
            ["PGE 1B", "DIVISI 3", "== DIVISI 3 TOTAL ==", "SUMMARY", 1680, "84.23% (1415)", ""],
            ["PGE 1B", "DIVISI 3", "MUHAMMAD YUSUF", "KERANI", 450, "91.11% (410)", "10 perbedaan (2.4%)"],
            ["PGE 1B", "DIVISI 3", "NUR HIDAYAT", "KERANI", 420, "87.62% (368)", "16 perbedaan (4.3%)"],
            ["PGE 1B", "DIVISI 3", "OKTO MANURUNG", "KERANI", 400, "83.75% (335)", "20 perbedaan (6.0%)"],
            ["PGE 1B", "DIVISI 3", "PRAYOGO UTOMO", "KERANI", 410, "76.10% (312)", "38 perbedaan (12.2%)"],
            ["PGE 1B", "DIVISI 3", "QUSENTINO ARI", "MANDOR", 300, "17.86%", ""],
            ["PGE 1B", "DIVISI 3", "RONAL TAMPUBOLON", "ASISTEN", 260, "15.48%", ""],

            # Separator
            ["", "", "", "", "", "", ""],

            # IJL - DIVISI 4
            ["IJL", "DIVISI 4", "== DIVISI 4 TOTAL ==", "SUMMARY", 1250, "87.20% (1090)", ""],
            ["IJL", "DIVISI 4", "SAFULLAH BAHRI", "KERANI", 380, "93.42% (355)", "6 perbedaan (1.7%)"],
            ["IJL", "DIVISI 4", "TRI AJI SETYO", "KERANI", 350, "88.57% (310)", "12 perbedaan (3.9%)"],
            ["IJL", "DIVISI 4", "UGO SANTOSO", "KERANI", 320, "84.38% (270)", "18 perbedaan (6.7%)"],
            ["IJL", "DIVISI 4", "VICTOR PANGGABEAN", "KERANI", 200, "78.00% (156)", "25 perbedaan (16.0%)"],
            ["IJL", "DIVISI 4", "WAWAN SETIAWAN", "MANDOR", 240, "19.20%", ""],
            ["IJL", "DIVISI 4", "XENA KARTIKA", "ASISTEN", 220, "17.60%", ""],

            # Separator
            ["", "", "", "", "", "", ""],

            # DME - DIVISI 5
            ["DME", "DIVISI 5", "== DIVISI 5 TOTAL ==", "SUMMARY", 980, "89.29% (875)", ""],
            ["DME", "DIVISI 5", "YOGI HENDRAWAN", "KERANI", 320, "94.69% (303)", "5 perbedaan (1.7%)"],
            ["DME", "DIVISI 5", "ZAINAL ARIFIN", "KERANI", 280, "90.00% (252)", "8 perbedaan (3.2%)"],
            ["DME", "DIVISI 5", "ADI SANTOSA", "KERANI", 250, "85.60% (214)", "15 perbedaan (7.0%)"],
            ["DME", "DIVISI 5", "BAMBANG SUTRISNO", "KERANI", 130, "82.31% (107)", "12 perbedaan (11.2%)"],
            ["DME", "DIVISI 5", "CECEP SOPIAN", "MANDOR", 180, "18.37%", ""],
            ["DME", "DIVISI 5", "DEDI KURNIAWAN", "ASISTEN", 160, "16.33%", ""]
        ]

        # Apply data with styling
        row_num = 2
        grand_kerani = 0
        grand_verified = 0

        for row_data in sample_data:
            if not any(row_data):  # Empty separator row
                for col_idx in range(1, 8):
                    cell = ws.cell(row=row_num, column=col_idx, value="")
                    cell.border = self.thin_border
                row_num += 1
                continue

            # Check if this is a summary or total row
            if "TOTAL ==" in str(row_data[2]) or "GRAND TOTAL" in str(row_data[0]):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.fill = self.summary_fill
                    cell.font = self.summary_font
                    cell.alignment = self.center_alignment
                    cell.border = self.thin_border

                # Add to grand total
                if row_data[4] and isinstance(row_data[4], (int, float)):
                    grand_kerani += row_data[4]
                if "(" in str(row_data[5]) and ")" in str(row_data[5]):
                    try:
                        verified_part = str(row_data[5]).split("(")[1].split(")")[0]
                        grand_verified += int(verified_part)
                    except:
                        pass

            else:
                # Employee row - determine role for styling
                role = str(row_data[3]).upper() if len(row_data) > 3 else ""

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.border = self.thin_border

                    if role == "KERANI":
                        cell.fill = self.kerani_fill
                        cell.font = self.default_font
                        cell.alignment = self.center_alignment

                        if col_idx == 6:  # Percentage column
                            cell.font = self.kerani_percentage_font
                        elif col_idx == 7:  # Keterangan column
                            cell.font = self.keterangan_font

                    elif role == "MANDOR":
                        cell.fill = self.mandor_fill
                        cell.font = self.default_font
                        cell.alignment = self.center_alignment

                        if col_idx == 6:  # Percentage column
                            cell.font = self.mandor_percentage_font

                    elif role == "ASISTEN":
                        cell.fill = self.asisten_fill
                        cell.font = self.default_font
                        cell.alignment = self.center_alignment

                        if col_idx == 6:  # Percentage column
                            cell.font = self.asisten_percentage_font

                    else:
                        cell.font = self.default_font
                        cell.alignment = self.center_alignment

            row_num += 1

        # Add Grand Total row
        grand_verification_rate = (grand_verified / grand_kerani * 100) if grand_kerani > 0 else 0
        grand_total_data = [
            "=== GRAND TOTAL ===", "", "", "",
            grand_kerani,
            f"{grand_verification_rate:.2f}% ({grand_verified})",
            ""
        ]

        for col_idx, value in enumerate(grand_total_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = self.grand_total_fill
            cell.font = self.grand_total_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        # Add summary section below the table
        row_num += 2
        summary_titles = [
            "PENJELASAN LAPORAN KINERJA",
            "",
            "• KERANI: % transaksi yang sudah diverifikasi (ada duplikat TRANSNO dengan P1/P5) dari total yang ia buat.",
            "• MANDOR/ASISTEN: % transaksi yang ia buat per total Kerani di divisi tersebut.",
            "• SUMMARY: % verifikasi keseluruhan divisi (Total Transaksi Kerani Terverifikasi / Total Transaksi Kerani).",
            "• GRAND TOTAL: % verifikasi keseluruhan untuk semua estate yang dipilih.",
            "",
            "KETERANGAN PERBEDAAN INPUT (INDIKATOR KINERJA):",
            "",
            "• Metodologi: Untuk setiap transaksi KERANI yang terverifikasi, sistem menghitung jumlah field yang berbeda.",
            "• Field yang dibandingkan: RIPEBCH, UNRIPEBCH, BLACKBCH, ROTTENBCH, LONGSTALKBCH, RATDMGBCH, LOOSEFRUIT.",
            "• Format: 'X perbedaan (Y%)' dimana Y% = (X perbedaan / Jumlah transaksi terverifikasi) × 100.",
            "• Interpretasi: Semakin banyak perbedaan, semakin besar kemungkinan ada ketidakakuratan dalam input data."
        ]

        for title in summary_titles:
            if title:
                cell = ws.cell(row=row_num, column=1, value=title)
                if title == "PENJELASAN LAPORAN KINERJA" or title == "KETERANGAN PERBEDAAN INPUT (INDIKATOR KINERJA):":
                    cell.font = Font(name="Arial", size=12, bold=True, color="2D3748")
                else:
                    cell.font = Font(name="Arial", size=10, color="4A5568")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row_num += 1

        # Add footer
        row_num += 1
        footer_text = f"Laporan dibuat pada: {datetime.now().strftime('%d %B %Y, %H:%M:%S')} | Sistem Analisis Real-time | PT. Rebinmas Jaya"
        cell = ws.cell(row=row_num, column=1, value=footer_text)
        cell.font = Font(name="Arial", size=8, color="718096", italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Merge footer cells
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)

        # Save file
        filepath = os.path.join(os.getcwd(), self.filename)
        wb.save(filepath)

        print(f"Sample Excel report created: {filepath}")
        print(f"Contains realistic data with {len([r for r in sample_data if r[0]])} data rows")
        print(f"Same format and styling as the original PDF report")

        return filepath

    def create_styles(self):
        """Create all styles for the Excel report"""
        # Header styles
        self.header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        self.header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Summary styles
        self.summary_fill = PatternFill(start_color="4299E1", end_color="4299E1", fill_type="solid")
        self.summary_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

        # Kerani styles
        self.kerani_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
        self.kerani_percentage_font = Font(name="Arial", size=8, color="E53E3E")
        self.keterangan_font = Font(name="Arial", size=8, color="C53030", bold=True)

        # Mandor styles
        self.mandor_fill = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
        self.mandor_percentage_font = Font(name="Arial", size=8, color="38A169")

        # Asisten styles
        self.asisten_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
        self.asisten_percentage_font = Font(name="Arial", size=8, color="3182CE")

        # Grand total styles
        self.grand_total_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        self.grand_total_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        # Default styles
        self.default_font = Font(name="Arial", size=8)
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def main():
    creator = SampleExcelReportCreator()
    creator.create_sample_report()

if __name__ == "__main__":
    main()