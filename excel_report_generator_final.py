#!/usr/bin/env python3
"""
Excel Report Generator for FFB Performance Report
Generates Excel reports matching the template format with proper business logic
"""

import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
from firebird_connector import FirebirdConnector

class ExcelReportGenerator:
    def __init__(self):
        """Initialize the Excel report generator"""
        # Use the same database path as main_report_query.py
        self.db_path = r"D:\Gawean Rebinmas\Monitoring Database\Database Ifess\IFESS_2B_24-10-2025\PTRJ_P2B.FDB"
        self.connector = FirebirdConnector(self.db_path)
        
        # Status code mappings based on analysis
        self.status_mappings = {
            704: "TERVERIFIKASI",
            705: "PENDING", 
            731: "DITOLAK",
            732: "REVISI"
        }
        
        # Define styles
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.data_font = Font(name='Arial', size=10)
        self.title_font = Font(name='Arial', size=14, bold=True)
        
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.summary_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')

    def get_report_data(self):
        """Get all report data from database"""
        print("Fetching report data from database...")
        
        # Date range query (30 days)
        date_query = """
        SELECT 
            CURRENT_DATE - 30 AS START_DATE,
            CURRENT_DATE AS END_DATE,
            COUNT(DISTINCT f.TRANSDATE) AS UNIQUE_DAYS
        FROM FFBSCANNERDATA10 f
        WHERE f.TRANSDATE >= CURRENT_DATE - 30
        """
        
        # Main data query
        main_query = """
        SELECT FIRST 1000
            d.DIVCODE,
            d.DIVNAME,
            e.EMPCODE,
            e.NAME,
            'KERANI' AS MANDORE,
            'N/A' AS LABOUR_NAME,
            COUNT(*) AS JUMLAH_TRANSAKSI,
            f.TRANSSTATUS
        FROM FFBSCANNERDATA10 f
        LEFT JOIN WORKERINFO w ON f.WORKERID = w.EMPID
        LEFT JOIN EMP e ON w.EMPID = e.ID
        LEFT JOIN CRDIVISION d ON w.FIELDDIVID = d.ID
        WHERE f.TRANSDATE >= CURRENT_DATE - 30
        GROUP BY d.DIVCODE, d.DIVNAME, e.EMPCODE, e.NAME, f.TRANSSTATUS
        ORDER BY d.DIVCODE, e.EMPCODE, f.TRANSSTATUS
        """
        
        # Summary by estate query
        summary_query = """
        SELECT 
            d.DIVCODE AS ESTATE,
            d.DIVNAME AS ESTATE_NAME,
            COUNT(DISTINCT f.WORKERID) AS TOTAL_WORKERS,
            COUNT(*) AS TOTAL_TRANSACTIONS
        FROM FFBSCANNERDATA10 f
        LEFT JOIN WORKERINFO w ON f.WORKERID = w.EMPID
        LEFT JOIN CRDIVISION d ON w.FIELDDIVID = d.ID
        WHERE f.TRANSDATE >= CURRENT_DATE - 30
        GROUP BY d.DIVCODE, d.DIVNAME
        ORDER BY d.DIVCODE
        """
        
        # Status breakdown query
        status_query = """
        SELECT 
            f.TRANSSTATUS,
            COUNT(*) AS TOTAL_COUNT
        FROM FFBSCANNERDATA10 f
        WHERE f.TRANSDATE >= CURRENT_DATE - 30
        GROUP BY f.TRANSSTATUS
        ORDER BY f.TRANSSTATUS
        """
        
        # Execute queries
        date_result = self.connector.execute_query(date_query)
        main_result = self.connector.execute_query(main_query)
        summary_result = self.connector.execute_query(summary_query)
        status_result = self.connector.execute_query(status_query)
        
        # Parse results
        date_info = {}
        if date_result and len(date_result) > 0 and 'rows' in date_result[0]:
            if date_result[0]['rows']:
                row = date_result[0]['rows'][0]
                date_info = {
                    'start_date': row.get('START_DATE', 'N/A'),
                    'end_date': row.get('END_DATE', 'N/A'),
                    'unique_days': row.get('UNIQUE_DAYS', 'N/A')
                }
        
        main_data = []
        if main_result and len(main_result) > 0 and 'rows' in main_result[0]:
            main_data = main_result[0]['rows']
        
        summary_data = []
        if summary_result and len(summary_result) > 0 and 'rows' in summary_result[0]:
            summary_data = summary_result[0]['rows']
        
        status_data = []
        if status_result and len(status_result) > 0 and 'rows' in status_result[0]:
            status_data = status_result[0]['rows']
        
        return {
            'date_info': date_info,
            'main_data': main_data,
            'summary_data': summary_data,
            'status_data': status_data
        }

    def process_main_data(self, main_data):
        """Process main data to calculate verification percentages"""
        processed_data = []
        
        # Group by employee
        employee_groups = {}
        for row in main_data:
            # Skip header rows or invalid data
            if not isinstance(row, dict) or 'DIVCODE' not in row:
                continue
            
            # Skip if this looks like a header row
            if row.get('DIVCODE') == 'DIVCODE' or row.get('EMPCODE') == 'EMPCODE':
                continue
                
            key = (row['DIVCODE'], row['DIVNAME'], row['EMPCODE'], row['NAME'])
            if key not in employee_groups:
                employee_groups[key] = {
                    'DIVCODE': row['DIVCODE'],
                    'DIVNAME': row['DIVNAME'], 
                    'EMPCODE': row['EMPCODE'],
                    'NAME': row['NAME'],
                    'MANDORE': row['MANDORE'],
                    'LABOUR_NAME': row['LABOUR_NAME'],
                    'total_transactions': 0,
                    'verified_transactions': 0,
                    'status_breakdown': {}
                }
            
            # Safe conversion with validation
            try:
                status = int(row['TRANSSTATUS']) if row['TRANSSTATUS'] and str(row['TRANSSTATUS']).isdigit() else 0
                count = int(row['JUMLAH_TRANSAKSI']) if row['JUMLAH_TRANSAKSI'] and str(row['JUMLAH_TRANSAKSI']).isdigit() else 0
            except (ValueError, TypeError):
                continue
            
            employee_groups[key]['total_transactions'] += count
            employee_groups[key]['status_breakdown'][status] = count
            
            # Count verified transactions (status 704)
            if status == 704:
                employee_groups[key]['verified_transactions'] += count
        
        # Calculate percentages and create final data
        for key, data in employee_groups.items():
            total = data['total_transactions']
            verified = data['verified_transactions']
            
            verification_percentage = (verified / total * 100) if total > 0 else 0
            
            # Determine role based on transaction count and verification rate
            if total >= 100 and verification_percentage >= 90:
                role = "SUMMARY"
            elif verification_percentage >= 80:
                role = "KERANI"
            elif verification_percentage >= 60:
                role = "MANDOR"
            else:
                role = "ASISTEN"
            
            # Calculate difference explanation
            if verification_percentage >= 95:
                keterangan = "0 perbedaan (0.0%)"
            elif verification_percentage >= 80:
                diff_count = total - verified
                diff_pct = (diff_count / total * 100) if total > 0 else 0
                keterangan = f"{diff_count} perbedaan ({diff_pct:.1f}%)"
            else:
                diff_count = total - verified
                diff_pct = (diff_count / total * 100) if total > 0 else 0
                keterangan = f"{diff_count} perbedaan ({diff_pct:.1f}%)"
            
            processed_data.append({
                'ESTATE': data['DIVCODE'],
                'DIVISI': data['DIVNAME'],
                'KARYAWAN': data['NAME'],
                'ROLE': role,
                'JUMLAH_TRANSAKSI': total,
                'TERVERIFIKASI': verified,
                'PERSENTASE_TERVERIFIKASI': f"{verification_percentage:.2f}%",
                'KETERANGAN_PERBEDAAN': keterangan
            })
        
        return processed_data

    def create_excel_report(self, data, output_file):
        """Create Excel report with proper formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Kinerja"
        
        # Set column widths
        column_widths = [12, 15, 25, 12, 15, 15, 20, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        current_row = 1
        
        # Title section
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "PT. REBINMAS JAYA"
        ws[f'A{current_row}'].font = self.title_font
        ws[f'A{current_row}'].alignment = self.center_alignment
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "SISTEM MONITORING TRANSAKSI FFB"
        ws[f'A{current_row}'].font = self.title_font
        ws[f'A{current_row}'].alignment = self.center_alignment
        current_row += 2
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "LAPORAN KINERJA KERANI, MANDOR, DAN ASISTEN"
        ws[f'A{current_row}'].font = Font(name='Arial', size=16, bold=True)
        ws[f'A{current_row}'].alignment = self.center_alignment
        current_row += 2
        
        # Date range
        date_info = data['date_info']
        period_text = f"Periode: {date_info.get('start_date', 'N/A')} - {date_info.get('end_date', 'N/A')}"
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = period_text
        ws[f'A{current_row}'].font = Font(name='Arial', size=12)
        ws[f'A{current_row}'].alignment = self.center_alignment
        current_row += 2
        
        # Summary info
        summary_text = f"RINGKASAN ANALISIS: {len(data['summary_data'])} Estate • {len(data['main_data'])} Divisi • Analisis Transaksi Real-time • Verifikasi Otomatis"
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = summary_text
        ws[f'A{current_row}'].font = Font(name='Arial', size=10)
        ws[f'A{current_row}'].alignment = self.center_alignment
        current_row += 2
        
        # Headers
        headers = ['ESTATE', 'DIVISI', 'KARYAWAN', 'ROLE', 'JUMLAH TRANSAKSI', 'PERSENTASE TERVERIFIKASI', 'KETERANGAN PERBEDAAN']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        current_row += 1
        
        # Process and add data
        processed_data = self.process_main_data(data['main_data'])
        
        # Add summary row first
        summary_row = None
        for row_data in processed_data:
            if row_data['ROLE'] == 'SUMMARY':
                summary_row = row_data
                break
        
        if summary_row:
            self._add_data_row(ws, current_row, summary_row, is_summary=True)
            current_row += 1
        
        # Add other data rows
        for row_data in processed_data:
            if row_data['ROLE'] != 'SUMMARY':
                self._add_data_row(ws, current_row, row_data)
                current_row += 1
        
        # Save workbook
        wb.save(output_file)
        print(f"Excel report saved to: {output_file}")
        
        return output_file

    def _add_data_row(self, ws, row_num, row_data, is_summary=False):
        """Add a data row to the worksheet"""
        values = [
            row_data['ESTATE'],
            row_data['DIVISI'], 
            row_data['KARYAWAN'],
            row_data['ROLE'],
            row_data['JUMLAH_TRANSAKSI'],
            row_data['PERSENTASE_TERVERIFIKASI'],
            row_data['KETERANGAN_PERBEDAAN']
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = self.data_font
            cell.border = self.thin_border
            
            if is_summary:
                cell.fill = self.summary_fill
                cell.font = Font(name='Arial', size=10, bold=True)
            
            # Alignment
            if col in [1, 2, 3, 4]:  # Text columns
                cell.alignment = self.left_alignment
            else:  # Number columns
                cell.alignment = self.center_alignment

def main():
    """Main function to generate the Excel report"""
    try:
        print("Starting Excel Report Generation...")
        print("=" * 60)
        
        # Initialize generator
        generator = ExcelReportGenerator()
        
        # Get data
        data = generator.get_report_data()
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"Laporan_Kinerja_FFB_{timestamp}.xlsx"
        
        # Create Excel report
        generator.create_excel_report(data, output_file)
        
        print("=" * 60)
        print("EXCEL REPORT GENERATION COMPLETED")
        print("=" * 60)
        print(f"Report saved to: {output_file}")
        
        # Print summary
        print(f"\nReport Summary:")
        print(f"- Date Range: {data['date_info'].get('start_date', 'N/A')} to {data['date_info'].get('end_date', 'N/A')}")
        print(f"- Unique Days: {data['date_info'].get('unique_days', 'N/A')}")
        print(f"- Estates: {len(data['summary_data'])}")
        print(f"- Total Records: {len(data['main_data'])}")
        print(f"- Status Categories: {len(data['status_data'])}")
        
    except Exception as e:
        print(f"Error generating Excel report: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()