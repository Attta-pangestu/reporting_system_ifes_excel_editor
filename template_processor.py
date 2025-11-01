"""
Template Processor untuk Excel Report Generator
Menangani pemrosesan template Excel dengan placeholder dan formula
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
import json
import os
from typing import Dict, List, Any, Tuple, Optional
import pandas as pd
import logging
import time
from datetime import datetime

class TemplateProcessor:
    """
    Kelas untuk memproses template Excel dengan placeholder dan formula
    """
    
    def __init__(self, template_path: str, formula_path: str):
        """
        Inisialisasi TemplateProcessor
        
        Args:
            template_path: Path ke file template Excel
            formula_path: Path ke file definisi formula JSON
        """
        # Setup logging
        self.logger = logging.getLogger(__name__)
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.DEBUG)
        
        self.logger.info(f"Initializing TemplateProcessor with template: {template_path}")
        self.logger.info(f"Formula definitions path: {formula_path}")
        
        self.template_path = template_path
        self.formula_path = formula_path
        self.template_wb = None
        self.formulas = {}
        self.placeholders = {}
        
        # Load template dan formula
        start_time = time.time()
        self._load_template()
        self._load_formulas()
        self._scan_placeholders()
        
        load_time = time.time() - start_time
        self.logger.info(f"Template processor initialization completed in {load_time:.3f} seconds")
        self.logger.debug(f"Found {len(self.placeholders)} sheets with placeholders")
        
        # Log summary of loaded components
        total_placeholders = sum(len(sheet_placeholders) for sheet_placeholders in self.placeholders.values())
        self.logger.info(f"Total placeholders found: {total_placeholders}")
        self.logger.info(f"Formula definitions loaded: {len(self.formulas.get('variables', {}))}")
    
    def _load_template(self):
        """Load template Excel file"""
        try:
            self.logger.debug(f"Loading template file: {self.template_path}")
            
            if not os.path.exists(self.template_path):
                self.logger.error(f"Template file not found: {self.template_path}")
                raise FileNotFoundError(f"Template file tidak ditemukan: {self.template_path}")
            
            file_size = os.path.getsize(self.template_path)
            self.logger.debug(f"Template file size: {file_size} bytes")
            
            start_time = time.time()
            self.template_wb = openpyxl.load_workbook(self.template_path)
            load_time = time.time() - start_time
            
            self.logger.info(f"Template loaded successfully in {load_time:.3f} seconds")
            self.logger.debug(f"Template contains {len(self.template_wb.sheetnames)} sheets: {self.template_wb.sheetnames}")
            
        except Exception as e:
            self.logger.error(f"Error loading template: {str(e)}")
            raise
    
    def _load_formulas(self):
        """Load formula definitions dari JSON file"""
        try:
            self.logger.debug(f"Loading formula definitions: {self.formula_path}")
            
            if not os.path.exists(self.formula_path):
                self.logger.warning(f"Formula file not found: {self.formula_path}, using empty formulas")
                self.formulas = {}
                return
            
            with open(self.formula_path, 'r', encoding='utf-8') as f:
                self.formulas = json.load(f)
            
            self.logger.info(f"Formula definitions loaded successfully")
            self.logger.debug(f"Formula structure: {list(self.formulas.keys())}")
            
            # Log detailed formula information
            if 'variables' in self.formulas:
                var_count = len(self.formulas['variables'])
                self.logger.debug(f"Variables defined: {var_count}")
                for var_name, var_def in self.formulas['variables'].items():
                    self.logger.debug(f"Variable '{var_name}': type={var_def.get('type', 'unknown')}")
            
            if 'repeating_sections' in self.formulas:
                section_count = len(self.formulas['repeating_sections'])
                self.logger.debug(f"Repeating sections defined: {section_count}")
                
        except json.JSONDecodeError as e:
            self.logger.error(f"Invalid JSON in formula file: {str(e)}")
            self.formulas = {}
        except Exception as e:
            self.logger.error(f"Error loading formulas: {str(e)}")
            self.formulas = {}
    
    def _scan_placeholders(self):
        """Scan semua sheet untuk menemukan placeholder"""
        self.logger.info("Starting placeholder scanning across all sheets")
        start_time = time.time()
        
        self.placeholders = {}
        placeholder_pattern = re.compile(r'\{\{([^}]+)\}\}')
        total_cells_scanned = 0
        total_placeholders_found = 0
        
        for sheet_name in self.template_wb.sheetnames:
            self.logger.debug(f"Scanning sheet: {sheet_name}")
            sheet = self.template_wb[sheet_name]
            sheet_placeholders = {}
            sheet_cells_scanned = 0
            sheet_placeholders_found = 0
            
            # Scan setiap cell dalam sheet
            for row in sheet.iter_rows():
                for cell in row:
                    sheet_cells_scanned += 1
                    if cell.value and isinstance(cell.value, str):
                        matches = placeholder_pattern.findall(cell.value)
                        if matches:
                            for match in matches:
                                placeholder_key = match.strip()
                                if placeholder_key not in sheet_placeholders:
                                    sheet_placeholders[placeholder_key] = []
                                
                                sheet_placeholders[placeholder_key].append({
                                    'cell': cell.coordinate,
                                    'original_value': cell.value,
                                    'row': cell.row,
                                    'column': cell.column
                                })
                                sheet_placeholders_found += 1
                                
                                self.logger.debug(f"Found placeholder '{placeholder_key}' in {sheet_name}!{cell.coordinate}: {cell.value}")
            
            if sheet_placeholders:
                self.placeholders[sheet_name] = sheet_placeholders
                self.logger.info(f"Sheet '{sheet_name}': {len(sheet_placeholders)} unique placeholders, {sheet_placeholders_found} total occurrences")
            else:
                self.logger.debug(f"Sheet '{sheet_name}': No placeholders found")
            
            total_cells_scanned += sheet_cells_scanned
            total_placeholders_found += sheet_placeholders_found
            self.logger.debug(f"Sheet '{sheet_name}': Scanned {sheet_cells_scanned} cells")
        
        scan_time = time.time() - start_time
        self.logger.info(f"Placeholder scanning completed in {scan_time:.3f} seconds")
        self.logger.info(f"Total cells scanned: {total_cells_scanned}")
        self.logger.info(f"Total placeholder occurrences found: {total_placeholders_found}")
        
        # Log summary by sheet
        for sheet_name, sheet_placeholders in self.placeholders.items():
            unique_placeholders = list(sheet_placeholders.keys())
            self.logger.debug(f"Sheet '{sheet_name}' placeholders: {unique_placeholders}")
    
    def get_placeholder_info(self) -> Dict:
        """Mendapatkan informasi placeholder yang ditemukan"""
        return self.placeholders
    
    def get_formula_definitions(self) -> Dict:
        """Mendapatkan definisi formula yang dimuat"""
        return self.formulas
    
    def process_template(self, data_context: Dict[str, Any]) -> openpyxl.Workbook:
        """
        Proses template dengan data context yang diberikan
        
        Args:
            data_context: Dictionary berisi data untuk mengisi placeholder
            
        Returns:
            openpyxl.Workbook: Workbook yang sudah diproses
        """
        self.logger.info("Starting template processing")
        self.logger.debug(f"Data context keys: {list(data_context.keys())}")
        
        # Log data context details
        for key, value in data_context.items():
            if isinstance(value, (list, dict)):
                self.logger.debug(f"Data context '{key}': {type(value).__name__} with {len(value)} items")
            else:
                self.logger.debug(f"Data context '{key}': {type(value).__name__} = {str(value)[:100]}")
        
        start_time = time.time()
        
        # Buat workbook baru untuk hasil
        processed_wb = openpyxl.Workbook()
        processed_wb.remove(processed_wb.active)  # Remove default sheet
        
        # Process setiap sheet
        for sheet_name in self.template_wb.sheetnames:
            self.logger.info(f"Processing sheet: {sheet_name}")
            sheet_start_time = time.time()
            
            source_sheet = self.template_wb[sheet_name]
            target_sheet = processed_wb.create_sheet(title=sheet_name)
            
            # Copy struktur sheet
            self.logger.debug(f"Copying sheet structure for: {sheet_name}")
            self._copy_sheet_structure(source_sheet, target_sheet)
            
            # Process placeholders
            self.logger.debug(f"Processing placeholders for: {sheet_name}")
            self._process_sheet_placeholders(target_sheet, sheet_name, data_context)
            
            # Process repeating sections
            self.logger.debug(f"Processing repeating sections for: {sheet_name}")
            self._process_repeating_sections(target_sheet, sheet_name, data_context)
            
            sheet_time = time.time() - sheet_start_time
            self.logger.info(f"Sheet '{sheet_name}' processed in {sheet_time:.3f} seconds")
        
        total_time = time.time() - start_time
        self.logger.info(f"Template processing completed in {total_time:.3f} seconds")
        
        return processed_wb
    
    def _copy_sheet_structure(self, source_sheet, target_sheet):
        """Copy struktur sheet termasuk formatting"""
        self.logger.debug(f"Copying sheet structure from {source_sheet.title}")
        cells_copied = 0
        
        # Copy semua cell values dan formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                target_cell.value = cell.value
                cells_copied += 1
                
                # Copy formatting
                if cell.font:
                    target_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                
                if cell.fill:
                    target_cell.fill = PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color
                    )
                
                if cell.border:
                    target_cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
                
                if cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text
                    )
        
        # Copy merged cells
        merged_count = 0
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
            merged_count += 1
        
        self.logger.debug(f"Copied {cells_copied} cells and {merged_count} merged ranges")
    
    def _process_sheet_placeholders(self, sheet, sheet_name: str, data_context: Dict[str, Any]):
        """Proses placeholder dalam sheet"""
        placeholders = self.placeholders.get(sheet_name, {})
        
        if not placeholders:
            self.logger.debug(f"No placeholders to process in sheet: {sheet_name}")
            return
        
        self.logger.info(f"Processing {len(placeholders)} unique placeholders in sheet: {sheet_name}")
        
        processed_count = 0
        failed_count = 0
        
        for placeholder_key, locations in placeholders.items():
            self.logger.debug(f"Processing placeholder: {placeholder_key}")
            
            # Dapatkan nilai untuk placeholder ini
            start_time = time.time()
            value = self._get_placeholder_value(placeholder_key, data_context)
            value_time = time.time() - start_time
            
            self.logger.debug(f"Placeholder '{placeholder_key}' resolved to: {str(value)[:200]} (took {value_time:.3f}s)")
            
            # Update semua lokasi placeholder ini
            for location in locations:
                try:
                    cell = sheet[location['cell']]
                    original_value = location['original_value']
                    
                    # Replace placeholder dengan nilai
                    if isinstance(original_value, str):
                        new_value = original_value.replace(f"{{{{{placeholder_key}}}}}", str(value))
                        cell.value = new_value
                        processed_count += 1
                        
                        self.logger.debug(f"Updated {location['cell']}: '{original_value}' -> '{new_value}'")
                    else:
                        self.logger.warning(f"Non-string original value in {location['cell']}: {type(original_value)}")
                        
                except Exception as e:
                    failed_count += 1
                    self.logger.error(f"Failed to update placeholder in {location['cell']}: {str(e)}")
        
        self.logger.info(f"Placeholder processing completed for {sheet_name}: {processed_count} successful, {failed_count} failed")
    
    def _get_placeholder_value(self, placeholder_key: str, data_context: Dict[str, Any]) -> Any:
        """Mendapatkan nilai untuk placeholder berdasarkan formula definition"""
        self.logger.debug(f"Resolving placeholder value for: {placeholder_key}")
        
        # Cek apakah ada definisi formula untuk placeholder ini
        variables = self.formulas.get('variables', {})
        
        if placeholder_key in variables:
            self.logger.debug(f"Found formula definition for: {placeholder_key}")
            formula_def = variables[placeholder_key]
            
            try:
                start_time = time.time()
                result = self._execute_formula(formula_def, data_context)
                exec_time = time.time() - start_time
                
                self.logger.debug(f"Formula execution for '{placeholder_key}' completed in {exec_time:.3f}s: {str(result)[:100]}")
                return result
                
            except Exception as e:
                self.logger.error(f"Error executing formula for '{placeholder_key}': {str(e)}")
                return f"[ERROR: {placeholder_key}]"
        
        # Jika tidak ada formula, cari langsung di data_context
        if placeholder_key in data_context:
            value = data_context[placeholder_key]
            self.logger.debug(f"Found direct value for '{placeholder_key}': {type(value).__name__}")
            return value
        
        # Cek nested keys (dot notation)
        if '.' in placeholder_key:
            try:
                keys = placeholder_key.split('.')
                current_value = data_context
                
                for key in keys:
                    if isinstance(current_value, dict) and key in current_value:
                        current_value = current_value[key]
                    else:
                        raise KeyError(f"Key '{key}' not found")
                
                self.logger.debug(f"Found nested value for '{placeholder_key}': {type(current_value).__name__}")
                return current_value
                
            except Exception as e:
                self.logger.warning(f"Failed to resolve nested key '{placeholder_key}': {str(e)}")
        
        # Default value jika tidak ditemukan
        self.logger.warning(f"Placeholder '{placeholder_key}' not found in data context or formulas")
        return f"[{placeholder_key}]"
    
    def _execute_formula(self, formula_def: Dict, data_context: Dict[str, Any]) -> Any:
        """Eksekusi formula definition untuk mendapatkan nilai"""
        formula_type = formula_def.get('type', 'direct')
        self.logger.debug(f"Executing formula type: {formula_type}")
        
        try:
            if formula_type == 'direct':
                # Ambil langsung dari data_context
                source_key = formula_def.get('source', '')
                default_value = formula_def.get('default', '')
                
                self.logger.debug(f"Direct formula - source: {source_key}, default: {default_value}")
                
                result = data_context.get(source_key, default_value)
                self.logger.debug(f"Direct formula result: {type(result).__name__}")
                return result
            
            elif formula_type == 'calculation':
                # Lakukan kalkulasi berdasarkan expression
                expression = formula_def.get('expression', '')
                self.logger.debug(f"Calculation formula - expression: {expression}")
                
                result = self._evaluate_expression(expression, data_context)
                self.logger.debug(f"Calculation result: {result}")
                return result
            
            elif formula_type == 'aggregation':
                # Lakukan agregasi data
                self.logger.debug(f"Aggregation formula - config: {formula_def}")
                result = self._perform_aggregation(formula_def, data_context)
                self.logger.debug(f"Aggregation result: {result}")
                return result
            
            elif formula_type == 'formatting':
                # Format data sesuai template
                self.logger.debug(f"Formatting formula - config: {formula_def}")
                result = self._format_data(formula_def, data_context)
                self.logger.debug(f"Formatting result: {result}")
                return result
            
            else:
                self.logger.warning(f"Unknown formula type: {formula_type}")
                return formula_def.get('default', '')
                
        except Exception as e:
            self.logger.error(f"Error executing {formula_type} formula: {str(e)}")
            return formula_def.get('default', f'[ERROR: {formula_type}]')
    
    def _evaluate_expression(self, expression: str, data_context: Dict[str, Any]) -> Any:
        """Evaluasi mathematical expression dengan data context"""
        self.logger.debug(f"Evaluating expression: {expression}")
        
        try:
            original_expression = expression
            substitution_count = 0
            
            # Replace variables dalam expression dengan nilai dari data_context
            for key, value in data_context.items():
                if isinstance(value, (int, float)):
                    placeholder = f"{{{key}}}"
                    if placeholder in expression:
                        expression = expression.replace(placeholder, str(value))
                        substitution_count += 1
                        self.logger.debug(f"Substituted {placeholder} with {value}")
            
            self.logger.debug(f"Expression after substitution ({substitution_count} replacements): {expression}")
            
            # Evaluasi expression (hati-hati dengan security)
            # Hanya izinkan operasi matematika dasar
            allowed_chars = set('0123456789+-*/.() ')
            if all(c in allowed_chars for c in expression):
                result = eval(expression)
                self.logger.debug(f"Expression evaluation successful: {original_expression} = {result}")
                return result
            else:
                self.logger.warning(f"Expression contains invalid characters: {expression}")
                return 0
                
        except Exception as e:
            self.logger.error(f"Error evaluating expression '{expression}': {str(e)}")
            return 0
    
    def _perform_aggregation(self, formula_def: Dict, data_context: Dict[str, Any]) -> Any:
        """Perform data aggregation"""
        agg_type = formula_def.get('aggregation_type', 'sum')
        source_key = formula_def.get('source', '')
        
        self.logger.debug(f"Performing {agg_type} aggregation on source: {source_key}")
        
        source_data = data_context.get(source_key, [])
        
        if not isinstance(source_data, list):
            self.logger.warning(f"Source data for aggregation is not a list: {type(source_data)}")
            return 0
        
        self.logger.debug(f"Aggregating {len(source_data)} items")
        
        try:
            if agg_type == 'sum':
                result = sum(source_data)
            elif agg_type == 'count':
                result = len(source_data)
            elif agg_type == 'average':
                result = sum(source_data) / len(source_data) if source_data else 0
            elif agg_type == 'max':
                result = max(source_data) if source_data else 0
            elif agg_type == 'min':
                result = min(source_data) if source_data else 0
            else:
                self.logger.warning(f"Unknown aggregation type: {agg_type}")
                result = 0
            
            self.logger.debug(f"Aggregation {agg_type} result: {result}")
            return result
            
        except Exception as e:
            self.logger.error(f"Error performing {agg_type} aggregation: {str(e)}")
            return 0
    
    def _format_data(self, formula_def: Dict, data_context: Dict[str, Any]) -> str:
        """Format data sesuai template"""
        source_key = formula_def.get('source', '')
        format_template = formula_def.get('format', '{value}')
        
        self.logger.debug(f"Formatting data from source: {source_key} with template: {format_template}")
        
        value = data_context.get(source_key, '')
        
        try:
            result = format_template.format(value=value)
            self.logger.debug(f"Formatting result: {value} -> {result}")
            return result
            
        except Exception as e:
            self.logger.error(f"Error formatting data: {str(e)}")
            return str(value)
    
    def _process_repeating_sections(self, sheet, sheet_name: str, data_context: Dict[str, Any]):
        """Proses repeating sections untuk data dinamis"""
        repeating_config = self.formulas.get('repeating_sections', {}).get(sheet_name, {})
        
        if not repeating_config:
            self.logger.debug(f"No repeating sections configured for sheet: {sheet_name}")
            return
        
        self.logger.info(f"Processing {len(repeating_config)} repeating sections in sheet: {sheet_name}")
        
        for section_name, config in repeating_config.items():
            self.logger.debug(f"Processing repeating section: {section_name}")
            start_time = time.time()
            
            try:
                self._process_single_repeating_section(sheet, config, data_context)
                section_time = time.time() - start_time
                self.logger.debug(f"Repeating section '{section_name}' processed in {section_time:.3f}s")
                
            except Exception as e:
                self.logger.error(f"Error processing repeating section '{section_name}': {str(e)}")
    
    def _process_single_repeating_section(self, sheet, config: Dict, data_context: Dict[str, Any]):
        """Proses satu repeating section"""
        start_row = config.get('start_row', 1)
        template_rows = config.get('template_rows', 1)
        data_source = config.get('data_source', '')
        
        self.logger.debug(f"Repeating section config - start_row: {start_row}, template_rows: {template_rows}, data_source: {data_source}")
        
        # Ambil data untuk repeating section
        repeat_data = data_context.get(data_source, [])
        
        if not repeat_data:
            self.logger.warning(f"No data found for repeating section source: {data_source}")
            return
        
        if not isinstance(repeat_data, list):
            self.logger.warning(f"Repeating section data is not a list: {type(repeat_data)}")
            return
        
        self.logger.info(f"Processing repeating section with {len(repeat_data)} data items")
        
        # Insert rows untuk data tambahan
        rows_needed = len(repeat_data) * template_rows
        current_rows = template_rows
        
        if rows_needed > current_rows:
            rows_to_insert = rows_needed - current_rows
            self.logger.debug(f"Inserting {rows_to_insert} additional rows")
            
            # Insert additional rows
            for i in range(rows_to_insert):
                sheet.insert_rows(start_row + current_rows + i)
        
        # Fill data untuk setiap item
        processed_items = 0
        for idx, item_data in enumerate(repeat_data):
            self.logger.debug(f"Processing repeat item {idx + 1}/{len(repeat_data)}: {list(item_data.keys()) if isinstance(item_data, dict) else type(item_data)}")
            
            row_offset = idx * template_rows
            
            # Process template rows untuk item ini
            for template_row_idx in range(template_rows):
                current_row = start_row + row_offset + template_row_idx
                
                # Process placeholders dalam row ini
                cells_processed = 0
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=current_row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        cell_value = cell.value
                        
                        # Replace placeholders dengan data item
                        if isinstance(item_data, dict):
                            for key, value in item_data.items():
                                placeholder = f"{{{{{key}}}}}"
                                if placeholder in cell_value:
                                    cell_value = cell_value.replace(placeholder, str(value))
                                    cells_processed += 1
                        
                        if cell_value != original_value:
                            cell.value = cell_value
                            self.logger.debug(f"Updated repeat cell {cell.coordinate}: '{original_value}' -> '{cell_value}'")
                
                self.logger.debug(f"Processed {cells_processed} placeholders in row {current_row}")
            
            processed_items += 1
        
        self.logger.info(f"Repeating section completed: {processed_items} items processed")
    
    def save_processed_template(self, processed_wb: openpyxl.Workbook, output_path: str):
        """Simpan template yang sudah diproses"""
        self.logger.info(f"Saving processed template to: {output_path}")
        
        try:
            start_time = time.time()
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
                self.logger.debug(f"Created output directory: {output_dir}")
            
            processed_wb.save(output_path)
            save_time = time.time() - start_time
            
            # Verify file was saved
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                self.logger.info(f"Template saved successfully in {save_time:.3f}s - Size: {file_size} bytes")
                self.logger.info(f"Output file: {output_path}")
            else:
                self.logger.error(f"Failed to save template - file not found after save operation")
                
        except Exception as e:
            self.logger.error(f"Error saving processed template: {str(e)}")
            raise