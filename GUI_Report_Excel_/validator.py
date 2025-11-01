#!/usr/bin/env python3
"""
Template Validator
Sistem validasi untuk memastikan konsistensi template dan kompatibilitas database

Fungsi utama:
- Validasi konsistensi antara file Excel dan JSON
- Validasi kompatibilitas struktur data dengan database
- Validasi placeholder dan mapping
- Validasi query SQL dan parameter

Author: AI Assistant
Date: 2025-10-31
"""

import openpyxl
import json
import re
import os
from typing import Dict, List, Any, Optional, Tuple
import logging
from pathlib import Path
import time
from datetime import datetime

class TemplateValidator:
    def __init__(self):
        """Initialize Template Validator"""
        # Setup comprehensive logging
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('validation_debug.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
        # Create dedicated validation logger
        self.validation_logger = logging.getLogger('validation_debug')
        self.validation_logger.setLevel(logging.DEBUG)
        
        # Validation results
        self.validation_results = {
            "errors": [],
            "warnings": [],
            "info": []
        }
        
        # Validation metrics
        self.validation_metrics = {
            'start_time': None,
            'end_time': None,
            'total_duration': 0,
            'stages_completed': 0,
            'stages_failed': 0,
            'data_processed': {}
        }
        
    def validate_template(self, template_info: Dict[str, Any], 
                         db_connector=None) -> Dict[str, List[str]]:
        """
        Comprehensive template validation
        
        Args:
            template_info: Template information dictionary
            db_connector: Optional database connector for database validation
            
        Returns:
            Dictionary with validation results (errors, warnings, info)
        """
        # Initialize validation session
        self.validation_metrics['start_time'] = time.time()
        validation_session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        self.validation_logger.info(f"=== VALIDATION SESSION START: {validation_session_id} ===")
        self.validation_logger.debug(f"Template info keys: {list(template_info.keys())}")
        self.validation_logger.debug(f"Template info size: {len(str(template_info))} characters")
        
        self.validation_results = {
            "errors": [],
            "warnings": [],
            "info": []
        }
        
        try:
            template_name = template_info.get('name', template_info.get('template_info', {}).get('name', 'Unknown'))
            self.logger.info(f"Starting validation for template: {template_name}")
            self.validation_logger.info(f"Template name: {template_name}")
            
            # Log template structure overview
            self._log_template_overview(template_info)
            
            # Stage 1: Basic template structure validation
            self._validate_template_structure_enhanced(template_info)
            
            # Stage 2: Excel file validation
            self._validate_excel_file_enhanced(template_info)
            
            # Stage 3: JSON configuration validation
            self._validate_json_config_enhanced(template_info)
            
            # Stage 4: Placeholder consistency validation
            self._validate_placeholder_consistency_enhanced(template_info)
            
            # Stage 5: Query validation
            self._validate_queries_enhanced(template_info)
            
            # Stage 6: Database compatibility validation (if connector provided)
            if db_connector:
                self._validate_database_compatibility_enhanced(template_info, db_connector)
            else:
                self.validation_logger.warning("Database connector not provided - skipping database compatibility validation")
                
            # Finalize validation
            self._finalize_validation()
            
        except Exception as e:
            self.validation_metrics['stages_failed'] += 1
            error_msg = f"Validation process error: {str(e)}"
            self.validation_results["errors"].append(error_msg)
            self.logger.error(error_msg)
            self.validation_logger.error(f"CRITICAL VALIDATION ERROR: {error_msg}")
            self.validation_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            
        # Calculate final metrics
        self.validation_metrics['end_time'] = time.time()
        self.validation_metrics['total_duration'] = self.validation_metrics['end_time'] - self.validation_metrics['start_time']
        
        self.validation_logger.info(f"=== VALIDATION SESSION END: {validation_session_id} ===")
         self.validation_logger.info(f"Total duration: {self.validation_metrics['total_duration']:.2f} seconds")
         self.validation_logger.info(f"Stages completed: {self.validation_metrics['stages_completed']}")
         self.validation_logger.info(f"Stages failed: {self.validation_metrics['stages_failed']}")
         self.validation_logger.info(f"Final results - Errors: {len(self.validation_results['errors'])}, Warnings: {len(self.validation_results['warnings'])}, Info: {len(self.validation_results['info'])}")
             
         return self.validation_results
        
    def _finalize_validation(self):
        """Finalize validation process and log summary"""
        try:
            self.validation_logger.info("=== VALIDATION FINALIZATION ===")
            
            # Calculate success rate
            total_stages = self.validation_metrics['stages_completed'] + self.validation_metrics['stages_failed']
            if total_stages > 0:
                success_rate = (self.validation_metrics['stages_completed'] / total_stages) * 100
                self.validation_logger.info(f"Validation success rate: {success_rate:.1f}%")
            
            # Log data processing metrics
            data_processed = self.validation_metrics.get('data_processed', {})
            if data_processed:
                self.validation_logger.info("Data processing metrics:")
                for metric, value in data_processed.items():
                    self.validation_logger.info(f"  {metric}: {value}")
            
            # Determine overall validation status
            error_count = len(self.validation_results['errors'])
            warning_count = len(self.validation_results['warnings'])
            
            if error_count == 0:
                if warning_count == 0:
                    status = "PASSED (No issues)"
                else:
                    status = f"PASSED (With {warning_count} warnings)"
                self.validation_logger.info(f"✓ Overall validation status: {status}")
            else:
                status = f"FAILED ({error_count} errors, {warning_count} warnings)"
                self.validation_logger.error(f"✗ Overall validation status: {status}")
            
            # Add final summary to results
            summary_msg = f"Validation completed: {status}"
            self.validation_results["info"].append(summary_msg)
            
        except Exception as e:
            self.validation_logger.error(f"Error during validation finalization: {str(e)}")
    
    def _log_template_overview(self, template_info: Dict[str, Any]):
        """Log comprehensive template overview for debugging"""
        try:
            self.validation_logger.debug("=== TEMPLATE OVERVIEW ===")
            
            # Log top-level structure
            for key, value in template_info.items():
                if isinstance(value, dict):
                    self.validation_logger.debug(f"Key '{key}': Dictionary with {len(value)} items")
                elif isinstance(value, list):
                    self.validation_logger.debug(f"Key '{key}': List with {len(value)} items")
                else:
                    self.validation_logger.debug(f"Key '{key}': {type(value).__name__} - {str(value)[:100]}...")
            
            # Log queries overview
            queries = template_info.get("queries", [])
            self.validation_logger.debug(f"Queries found: {len(queries)}")
            for i, query in enumerate(queries):
                if isinstance(query, dict):
                    query_name = query.get('name', f'Query_{i}')
                    self.validation_logger.debug(f"  Query {i}: {query_name}")
                    
            # Log mappings overview
            mappings = template_info.get("mappings", {})
            self.validation_logger.debug(f"Mappings found: {len(mappings)} items")
            
        except Exception as e:
            self.validation_logger.error(f"Error logging template overview: {str(e)}")
        
    def _validate_template_structure_enhanced(self, template_info: Dict[str, Any]):
        """Enhanced template structure validation with detailed logging"""
        stage_start = time.time()
        self.validation_logger.info("=== STAGE 1: Template Structure Validation ===")
        
        try:
            # Required top-level keys
            required_keys = ["template_info", "queries", "mappings"]
            self.validation_logger.debug(f"Checking for required keys: {required_keys}")
            
            missing_keys = []
            for key in required_keys:
                if key not in template_info:
                    missing_keys.append(key)
                    error_msg = f"Missing required key: {key}"
                    self.validation_results["errors"].append(error_msg)
                    self.validation_logger.error(f"STRUCTURE ERROR: {error_msg}")
                else:
                    self.validation_logger.debug(f"✓ Required key found: {key}")
                    
            if missing_keys:
                self.validation_logger.error(f"Missing keys summary: {missing_keys}")
            else:
                self.validation_logger.info("✓ All required top-level keys present")
                    
            # Template info validation
            if "template_info" in template_info:
                template_meta = template_info["template_info"]
                self.validation_logger.debug(f"Template metadata: {template_meta}")
                
                required_meta_keys = ["name", "excel_file"]
                self.validation_logger.debug(f"Checking template metadata keys: {required_meta_keys}")
                
                for key in required_meta_keys:
                    if key not in template_meta:
                        error_msg = f"Missing template_info key: {key}"
                        self.validation_results["errors"].append(error_msg)
                        self.validation_logger.error(f"METADATA ERROR: {error_msg}")
                    else:
                        value = template_meta[key]
                        self.validation_logger.debug(f"✓ Template metadata '{key}': {value}")
                        
            # File paths validation
            excel_path = template_info.get("excel_path")
            if excel_path:
                self.validation_logger.debug(f"Checking Excel file path: {excel_path}")
                if not os.path.exists(excel_path):
                    error_msg = f"Excel file not found: {excel_path}"
                    self.validation_results["errors"].append(error_msg)
                    self.validation_logger.error(f"FILE ERROR: {error_msg}")
                else:
                    file_size = os.path.getsize(excel_path)
                    self.validation_logger.info(f"✓ Excel file found: {excel_path} (Size: {file_size} bytes)")
            else:
                self.validation_logger.warning("Excel path not specified in template_info")
                
            json_path = template_info.get("json_path")
            if json_path:
                self.validation_logger.debug(f"Checking JSON file path: {json_path}")
                if not os.path.exists(json_path):
                    error_msg = f"JSON file not found: {json_path}"
                    self.validation_results["errors"].append(error_msg)
                    self.validation_logger.error(f"FILE ERROR: {error_msg}")
                else:
                    file_size = os.path.getsize(json_path)
                    self.validation_logger.info(f"✓ JSON file found: {json_path} (Size: {file_size} bytes)")
            else:
                self.validation_logger.warning("JSON path not specified in template_info")
                
            self.validation_metrics['stages_completed'] += 1
            stage_duration = time.time() - stage_start
            self.validation_logger.info(f"✓ Stage 1 completed in {stage_duration:.2f} seconds")
                
        except Exception as e:
            self.validation_metrics['stages_failed'] += 1
            error_msg = f"Template structure validation error: {str(e)}"
            self.validation_results["errors"].append(error_msg)
            self.validation_logger.error(f"STAGE 1 CRITICAL ERROR: {error_msg}")

    def _validate_excel_file_enhanced(self, template_info: Dict[str, Any]):
        """Enhanced Excel file validation with detailed logging"""
        stage_start = time.time()
        self.validation_logger.info("=== STAGE 2: Excel File Validation ===")
        
        try:
            excel_path = template_info.get("excel_path")
            if not excel_path:
                self.validation_logger.warning("No Excel path provided - skipping Excel validation")
                return
                
            if not os.path.exists(excel_path):
                error_msg = f"Excel file not found: {excel_path}"
                self.validation_results["errors"].append(error_msg)
                self.validation_logger.error(f"EXCEL ERROR: {error_msg}")
                return
                
            self.validation_logger.debug(f"Loading Excel file: {excel_path}")
            file_size = os.path.getsize(excel_path)
            self.validation_logger.debug(f"Excel file size: {file_size} bytes")
                
            # Load Excel file
            workbook = openpyxl.load_workbook(excel_path)
            self.validation_logger.debug(f"Excel workbook loaded successfully")
            
            # Check worksheets
            worksheet_count = len(workbook.worksheets)
            self.validation_logger.debug(f"Found {worksheet_count} worksheets")
            
            if worksheet_count == 0:
                error_msg = "Excel file has no worksheets"
                self.validation_results["errors"].append(error_msg)
                self.validation_logger.error(f"EXCEL ERROR: {error_msg}")
                return
                
            # Log worksheet details
            for i, sheet in enumerate(workbook.worksheets):
                sheet_name = sheet.title
                max_row = sheet.max_row
                max_col = sheet.max_column
                self.validation_logger.debug(f"  Worksheet {i}: '{sheet_name}' ({max_row} rows, {max_col} columns)")
                
            # Find placeholders in Excel
            excel_placeholders = set()
            placeholder_locations = []
            
            self.validation_logger.debug("Scanning for placeholders in Excel...")
            
            for sheet_idx, sheet in enumerate(workbook.worksheets):
                sheet_name = sheet.title
                sheet_placeholders = set()
                
                for row_idx, row in enumerate(sheet.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value and isinstance(cell.value, str):
                            placeholders = re.findall(r'\{\{([^}]+)\}\}', cell.value)
                            if placeholders:
                                for placeholder in placeholders:
                                    excel_placeholders.add(placeholder)
                                    sheet_placeholders.add(placeholder)
                                    location = f"{sheet_name}!{cell.coordinate}"
                                    placeholder_locations.append({
                                        'placeholder': placeholder,
                                        'location': location,
                                        'cell_value': cell.value
                                    })
                                    self.validation_logger.debug(f"    Found placeholder '{placeholder}' at {location}: {cell.value}")
                
                self.validation_logger.debug(f"  Sheet '{sheet_name}': {len(sheet_placeholders)} unique placeholders")
                
            # Store found placeholders for later comparison
            template_info["_excel_placeholders"] = excel_placeholders
            template_info["_placeholder_locations"] = placeholder_locations
            
            # Log placeholder summary
            self.validation_logger.info(f"✓ Excel placeholder scan complete: {len(excel_placeholders)} unique placeholders found")
            for placeholder in sorted(excel_placeholders):
                self.validation_logger.debug(f"    Placeholder: {placeholder}")
            
            info_msg = f"Found {len(excel_placeholders)} placeholders in Excel file"
            self.validation_results["info"].append(info_msg)
            
            # Store metrics
            self.validation_metrics['data_processed']['excel_worksheets'] = worksheet_count
            self.validation_metrics['data_processed']['excel_placeholders'] = len(excel_placeholders)
            
            workbook.close()
            self.validation_logger.debug("Excel workbook closed")
            
            self.validation_metrics['stages_completed'] += 1
            stage_duration = time.time() - stage_start
            self.validation_logger.info(f"✓ Stage 2 completed in {stage_duration:.2f} seconds")
            
        except Exception as e:
            self.validation_metrics['stages_failed'] += 1
            error_msg = f"Excel file validation error: {str(e)}"
            self.validation_results["errors"].append(error_msg)
            self.validation_logger.error(f"STAGE 2 CRITICAL ERROR: {error_msg}")
            self.validation_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            
    def _validate_json_config_enhanced(self, template_info: Dict[str, Any]):
        """Enhanced JSON configuration validation with detailed logging"""
        stage_start = time.time()
        self.validation_logger.info("=== STAGE 3: JSON Configuration Validation ===")
        
        try:
            # Validate queries structure
            queries = template_info.get("queries", [])
            self.validation_logger.debug(f"Found {len(queries)} queries to validate")
            
            if not isinstance(queries, list):
                error_msg = f"Queries must be a list, got {type(queries)}"
                self.validation_results["errors"].append(error_msg)
                self.validation_logger.error(f"QUERIES STRUCTURE ERROR: {error_msg}")
                return
                
            for i, query in enumerate(queries):
                query_name = query.get('name', f'Query_{i}') if isinstance(query, dict) else f'Query_{i}'
                self.validation_logger.debug(f"Validating query {i}: {query_name}")
                
                if not isinstance(query, dict):
                    error_msg = f"Query {i} must be a dictionary"
                    self.validation_results["errors"].append(error_msg)
                    self.validation_logger.error(f"QUERY STRUCTURE ERROR: {error_msg}")
                    continue
                    
                # Required query fields
                required_query_fields = ["name", "sql"]
                for field in required_query_fields:
                    if field not in query:
                        error_msg = f"Query {i} missing required field: {field}"
                        self.validation_results["errors"].append(error_msg)
                        self.validation_logger.error(f"QUERY FIELD ERROR: {error_msg}")
                    else:
                        self.validation_logger.debug(f"  ✓ Query {i} has required field: {field}")
                        
                # Validate SQL
                sql = query.get("sql", "")
                if not sql.strip():
                    error_msg = f"Query {i} has empty SQL"
                    self.validation_results["errors"].append(error_msg)
                    self.validation_logger.error(f"QUERY SQL ERROR: {error_msg}")
                else:
                    sql_length = len(sql.strip())
                    self.validation_logger.debug(f"  ✓ Query {i} SQL length: {sql_length} characters")
                    
                # Validate output format
                output_format = query.get("output_format", "table")
                valid_formats = ["table", "single_value", "list"]
                if output_format not in valid_formats:
                    warning_msg = f"Query {i} has invalid output_format: {output_format}"
                    self.validation_results["warnings"].append(warning_msg)
                    self.validation_logger.warning(f"QUERY FORMAT WARNING: {warning_msg}")
                else:
                    self.validation_logger.debug(f"  ✓ Query {i} output format: {output_format}")
                    
            # Validate mappings
            mappings = template_info.get("mappings", {})
            self.validation_logger.debug(f"Found {len(mappings)} mappings to validate")
            
            if not isinstance(mappings, dict):
                error_msg = f"Mappings must be a dictionary, got {type(mappings)}"
                self.validation_results["errors"].append(error_msg)
                self.validation_logger.error(f"MAPPINGS STRUCTURE ERROR: {error_msg}")
            else:
                for mapping_name, mapping_value in mappings.items():
                    self.validation_logger.debug(f"  Mapping '{mapping_name}': {type(mapping_value)}")
                
            # Validate transformations
            transformations = template_info.get("transformations", [])
            if not isinstance(transformations, list):
                warning_msg = f"Transformations should be a list, got {type(transformations)}"
                self.validation_results["warnings"].append(warning_msg)
                self.validation_logger.warning(f"TRANSFORMATIONS WARNING: {warning_msg}")
            else:
                self.validation_logger.debug(f"Found {len(transformations)} transformations")
                
            # Store metrics
            self.validation_metrics['data_processed']['json_queries'] = len(queries)
            self.validation_metrics['data_processed']['json_mappings'] = len(mappings)
            self.validation_metrics['data_processed']['json_transformations'] = len(transformations)
            
            self.validation_metrics['stages_completed'] += 1
            stage_duration = time.time() - stage_start
            self.validation_logger.info(f"✓ Stage 3 completed in {stage_duration:.2f} seconds")
                
        except Exception as e:
            self.validation_metrics['stages_failed'] += 1
            error_msg = f"JSON config validation error: {str(e)}"
            self.validation_results["errors"].append(error_msg)
            self.validation_logger.error(f"STAGE 3 CRITICAL ERROR: {error_msg}")
            self.validation_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            
    def _validate_placeholder_consistency_enhanced(self, template_info: Dict[str, Any]):
        """Enhanced placeholder consistency validation with detailed logging"""
        stage_start = time.time()
        self.validation_logger.info("=== STAGE 4: Placeholder Consistency Validation ===")
        
        try:
            excel_placeholders = template_info.get("_excel_placeholders", set())
            mappings = template_info.get("mappings", {})
            queries = template_info.get("queries", [])
            
            self.validation_logger.debug(f"Excel placeholders count: {len(excel_placeholders)}")
            self.validation_logger.debug(f"Excel placeholders: {sorted(excel_placeholders)}")
            
            # Get all available mappings
            available_mappings = set(mappings.keys())
            self.validation_logger.debug(f"Direct mappings: {sorted(available_mappings)}")
            
            # Add query names as available mappings
            query_mappings = set()
            for query in queries:
                query_name = query.get("name")
                if query_name:
                    available_mappings.add(query_name)
                    query_mappings.add(query_name)
                    self.validation_logger.debug(f"  Query mapping: {query_name}")
                    
            self.validation_logger.debug(f"Query mappings: {sorted(query_mappings)}")
                    
            # Add built-in mappings
            built_in_mappings = {
                "current_date", "current_time", "current_timestamp",
                "REPORT_DATE", "REPORT_TIME", "GENERATED_TIMESTAMP"
            }
            available_mappings.update(built_in_mappings)
            self.validation_logger.debug(f"Built-in mappings: {sorted(built_in_mappings)}")
            
            self.validation_logger.debug(f"Total available mappings: {len(available_mappings)}")
            
            # Check for unmapped placeholders
            unmapped_placeholders = excel_placeholders - available_mappings
            if unmapped_placeholders:
                for placeholder in sorted(unmapped_placeholders):
                    warning_msg = f"Placeholder '{placeholder}' in Excel has no mapping"
                    self.validation_results["warnings"].append(warning_msg)
                    self.validation_logger.warning(f"UNMAPPED PLACEHOLDER: {warning_msg}")
                    
                    # Log placeholder locations for debugging
                    placeholder_locations = template_info.get("_placeholder_locations", [])
                    for location_info in placeholder_locations:
                        if location_info['placeholder'] == placeholder:
                            self.validation_logger.warning(f"  Location: {location_info['location']} - {location_info['cell_value']}")
            else:
                self.validation_logger.info("✓ All Excel placeholders have mappings")
                
            # Check for unused mappings
            unused_mappings = available_mappings - excel_placeholders
            unused_non_builtin = unused_mappings - built_in_mappings
            
            if unused_non_builtin:
                for mapping in sorted(unused_non_builtin):
                    info_msg = f"Mapping '{mapping}' is defined but not used in Excel"
                    self.validation_results["info"].append(info_msg)
                    self.validation_logger.info(f"UNUSED MAPPING: {info_msg}")
            else:
                self.validation_logger.info("✓ All non-built-in mappings are used")
                    
            # Summary
            summary_msg = f"Placeholder consistency check: {len(excel_placeholders)} Excel placeholders, {len(available_mappings)} available mappings"
            self.validation_results["info"].append(summary_msg)
            self.validation_logger.info(f"CONSISTENCY SUMMARY: {summary_msg}")
            
            # Store metrics
            self.validation_metrics['data_processed']['placeholder_mappings'] = len(available_mappings)
            self.validation_metrics['data_processed']['unmapped_placeholders'] = len(unmapped_placeholders)
            self.validation_metrics['data_processed']['unused_mappings'] = len(unused_non_builtin)
            
            self.validation_metrics['stages_completed'] += 1
            stage_duration = time.time() - stage_start
            self.validation_logger.info(f"✓ Stage 4 completed in {stage_duration:.2f} seconds")
            
        except Exception as e:
            self.validation_metrics['stages_failed'] += 1
            error_msg = f"Placeholder consistency validation error: {str(e)}"
            self.validation_results["errors"].append(error_msg)
            self.validation_logger.error(f"STAGE 4 CRITICAL ERROR: {error_msg}")
            self.validation_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            
    def _validate_queries_enhanced(self, template_info: Dict[str, Any]):
        """Enhanced SQL queries validation with detailed logging"""
        stage_start = time.time()
        self.validation_logger.info("=== STAGE 5: SQL Queries Validation ===")
        
        try:
            queries = template_info.get("queries", [])
            self.validation_logger.debug(f"Validating {len(queries)} SQL queries")
            
            valid_queries = 0
            queries_with_issues = 0
            
            for i, query in enumerate(queries):
                query_name = query.get("name", f"Query_{i}")
                sql = query.get("sql", "")
                
                self.validation_logger.debug(f"Validating query '{query_name}'")
                
                # Basic SQL validation
                if not sql.strip():
                    warning_msg = f"Query '{query_name}' has empty SQL"
                    self.validation_results["warnings"].append(warning_msg)
                    self.validation_logger.warning(f"EMPTY SQL: {warning_msg}")
                    queries_with_issues += 1
                    continue
                    
                sql_length = len(sql.strip())
                self.validation_logger.debug(f"  SQL length: {sql_length} characters")
                    
                # Check for basic SQL structure
                sql_upper = sql.upper().strip()
                
                # Must start with SELECT, INSERT, UPDATE, DELETE, or WITH
                valid_starts = ["SELECT", "INSERT", "UPDATE", "DELETE", "WITH"]
                sql_start_valid = any(sql_upper.startswith(start) for start in valid_starts)
                
                if not sql_start_valid:
                    warning_msg = f"Query '{query_name}' may have invalid SQL structure"
                    self.validation_results["warnings"].append(warning_msg)
                    self.validation_logger.warning(f"SQL STRUCTURE: {warning_msg}")
                    queries_with_issues += 1
                else:
                    detected_type = next(start for start in valid_starts if sql_upper.startswith(start))
                    self.validation_logger.debug(f"  ✓ SQL type: {detected_type}")
                    
                # Check for potential SQL injection risks
                dangerous_keywords = ["DROP", "DELETE", "TRUNCATE", "ALTER", "CREATE"]
                found_dangerous = []
                
                for keyword in dangerous_keywords:
                    if keyword in sql_upper and not sql_upper.startswith("SELECT"):
                        found_dangerous.append(keyword)
                        
                if found_dangerous:
                    warning_msg = f"Query '{query_name}' contains potentially dangerous keywords: {found_dangerous}"
                    self.validation_results["warnings"].append(warning_msg)
                    self.validation_logger.warning(f"DANGEROUS SQL: {warning_msg}")
                    queries_with_issues += 1
                else:
                    self.validation_logger.debug(f"  ✓ No dangerous keywords detected")
                        
                # Validate parameters
                parameters = query.get("parameters", [])
                if parameters and not isinstance(parameters, list):
                    error_msg = f"Query '{query_name}' parameters must be a list"
                    self.validation_results["errors"].append(error_msg)
                    self.validation_logger.error(f"PARAMETER ERROR: {error_msg}")
                    queries_with_issues += 1
                else:
                    self.validation_logger.debug(f"  Parameters: {len(parameters)} defined")
                    
                # Check for parameter placeholders in SQL
                param_placeholders = re.findall(r'\?|\$\d+|:\w+', sql)
                placeholder_count = len(param_placeholders)
                parameter_count = len(parameters)
                
                self.validation_logger.debug(f"  SQL placeholders: {placeholder_count}, Parameters: {parameter_count}")
                
                if placeholder_count != parameter_count:
                    warning_msg = f"Query '{query_name}' parameter count mismatch: {placeholder_count} placeholders, {parameter_count} parameters"
                    self.validation_results["warnings"].append(warning_msg)
                    self.validation_logger.warning(f"PARAMETER MISMATCH: {warning_msg}")
                    queries_with_issues += 1
                else:
                    self.validation_logger.debug(f"  ✓ Parameter count matches")
                    
                # If no issues found, count as valid
                if sql.strip() and sql_start_valid and not found_dangerous and isinstance(parameters, list) and placeholder_count == parameter_count:
                    valid_queries += 1
                    self.validation_logger.debug(f"  ✓ Query '{query_name}' validation passed")
                    
            # Summary
            self.validation_logger.info(f"Query validation summary: {valid_queries} valid, {queries_with_issues} with issues")
            
            # Store metrics
            self.validation_metrics['data_processed']['valid_queries'] = valid_queries
            self.validation_metrics['data_processed']['queries_with_issues'] = queries_with_issues
            
            self.validation_metrics['stages_completed'] += 1
            stage_duration = time.time() - stage_start
            self.validation_logger.info(f"✓ Stage 5 completed in {stage_duration:.2f} seconds")
                    
        except Exception as e:
            self.validation_metrics['stages_failed'] += 1
            error_msg = f"Query validation error: {str(e)}"
            self.validation_results["errors"].append(error_msg)
            self.validation_logger.error(f"STAGE 5 CRITICAL ERROR: {error_msg}")
            self.validation_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            
    def _validate_database_compatibility_enhanced(self, template_info: Dict[str, Any], db_connector):
        """Enhanced database compatibility validation with detailed logging"""
        stage_start = time.time()
        self.validation_logger.info("=== STAGE 6: Database Compatibility Validation ===")
        
        try:
            # Check database connection
            self.validation_logger.debug("Testing database connection...")
            if not db_connector.test_connection():
                error_msg = "Cannot connect to database for compatibility check"
                self.validation_results["errors"].append(error_msg)
                self.validation_logger.error(f"DB CONNECTION ERROR: {error_msg}")
                return
            else:
                self.validation_logger.info("✓ Database connection successful")
                
            # Get database schema information
            try:
                self.validation_logger.debug("Retrieving database schema information...")
                tables_info = db_connector.get_tables_info()
                available_tables = set(table.upper() for table in tables_info.keys())
                self.validation_logger.debug(f"Found {len(available_tables)} tables in database")
                for table in sorted(available_tables):
                    self.validation_logger.debug(f"  Available table: {table}")
            except Exception as e:
                warning_msg = f"Could not retrieve database schema: {str(e)}"
                self.validation_results["warnings"].append(warning_msg)
                self.validation_logger.warning(f"SCHEMA WARNING: {warning_msg}")
                available_tables = set()
                
            # Check required tables
            required_tables = template_info.get("database_requirements", {}).get("required_tables", [])
            self.validation_logger.debug(f"Checking {len(required_tables)} required tables")
            
            missing_tables = []
            for table in required_tables:
                if table.upper() not in available_tables:
                    missing_tables.append(table)
                    error_msg = f"Required table '{table}' not found in database"
                    self.validation_results["errors"].append(error_msg)
                    self.validation_logger.error(f"MISSING TABLE: {error_msg}")
                else:
                    self.validation_logger.debug(f"  ✓ Required table found: {table}")
                    
            if not missing_tables:
                self.validation_logger.info("✓ All required tables found in database")
                    
            # Validate queries against database
            queries = template_info.get("queries", [])
            self.validation_logger.debug(f"Validating {len(queries)} queries against database")
            
            valid_queries = 0
            invalid_queries = 0
            
            for query in queries:
                query_name = query.get("name", "Unknown")
                sql = query.get("sql", "")
                
                if sql.strip():
                    self.validation_logger.debug(f"Validating query '{query_name}' against database")
                    try:
                        # Try to validate SQL (without executing)
                        is_valid = db_connector.validate_sql(sql)
                        if not is_valid:
                            warning_msg = f"Query '{query_name}' may have SQL syntax issues"
                            self.validation_results["warnings"].append(warning_msg)
                            self.validation_logger.warning(f"SQL VALIDATION: {warning_msg}")
                            invalid_queries += 1
                        else:
                            self.validation_logger.debug(f"  ✓ Query '{query_name}' SQL validation passed")
                            valid_queries += 1
                    except Exception as e:
                        warning_msg = f"Could not validate query '{query_name}': {str(e)}"
                        self.validation_results["warnings"].append(warning_msg)
                        self.validation_logger.warning(f"QUERY VALIDATION ERROR: {warning_msg}")
                        invalid_queries += 1
                        
            # Summary
            compatibility_summary = f"Database compatibility check completed - {valid_queries} valid queries, {invalid_queries} with issues"
            self.validation_results["info"].append(compatibility_summary)
            self.validation_logger.info(f"DB COMPATIBILITY SUMMARY: {compatibility_summary}")
            
            # Store metrics
            self.validation_metrics['data_processed']['db_tables_available'] = len(available_tables)
            self.validation_metrics['data_processed']['db_tables_missing'] = len(missing_tables)
            self.validation_metrics['data_processed']['db_queries_valid'] = valid_queries
            self.validation_metrics['data_processed']['db_queries_invalid'] = invalid_queries
            
            self.validation_metrics['stages_completed'] += 1
            stage_duration = time.time() - stage_start
            self.validation_logger.info(f"✓ Stage 6 completed in {stage_duration:.2f} seconds")
            
        except Exception as e:
            self.validation_metrics['stages_failed'] += 1
            error_msg = f"Database compatibility validation error: {str(e)}"
            self.validation_results["errors"].append(error_msg)
            self.validation_logger.error(f"STAGE 6 CRITICAL ERROR: {error_msg}")
            self.validation_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            
    def validate_template_files(self, excel_path: str, json_path: str) -> Dict[str, List[str]]:
        """Validate template files before loading"""
        results = {
            "errors": [],
            "warnings": [],
            "info": []
        }
        
        try:
            # Check file existence
            if not os.path.exists(excel_path):
                results["errors"].append(f"Excel file not found: {excel_path}")
                
            if not os.path.exists(json_path):
                results["errors"].append(f"JSON file not found: {json_path}")
                
            if results["errors"]:
                return results
                
            # Validate Excel file
            try:
                workbook = openpyxl.load_workbook(excel_path)
                if len(workbook.worksheets) == 0:
                    results["errors"].append("Excel file has no worksheets")
                workbook.close()
                results["info"].append("Excel file structure is valid")
            except Exception as e:
                results["errors"].append(f"Excel file validation error: {str(e)}")
                
            # Validate JSON file
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                    
                # Check required JSON structure
                if not isinstance(json_data, dict):
                    results["errors"].append("JSON file must contain a dictionary")
                else:
                    required_keys = ["template_info", "queries", "mappings"]
                    for key in required_keys:
                        if key not in json_data:
                            results["errors"].append(f"JSON file missing required key: {key}")
                            
                results["info"].append("JSON file structure is valid")
                
            except json.JSONDecodeError as e:
                results["errors"].append(f"JSON file syntax error: {str(e)}")
            except Exception as e:
                results["errors"].append(f"JSON file validation error: {str(e)}")
                
        except Exception as e:
            results["errors"].append(f"Template files validation error: {str(e)}")
            
        return results
        
    def generate_validation_report(self, validation_results: Dict[str, List[str]]) -> str:
        """Generate a formatted validation report"""
        try:
            report_lines = []
            report_lines.append("=" * 60)
            report_lines.append("TEMPLATE VALIDATION REPORT")
            report_lines.append("=" * 60)
            
            # Summary
            error_count = len(validation_results.get("errors", []))
            warning_count = len(validation_results.get("warnings", []))
            info_count = len(validation_results.get("info", []))
            
            report_lines.append(f"\nSUMMARY:")
            report_lines.append(f"  Errors: {error_count}")
            report_lines.append(f"  Warnings: {warning_count}")
            report_lines.append(f"  Info: {info_count}")
            
            # Errors
            if error_count > 0:
                report_lines.append(f"\nERRORS ({error_count}):")
                for i, error in enumerate(validation_results["errors"], 1):
                    report_lines.append(f"  {i}. {error}")
                    
            # Warnings
            if warning_count > 0:
                report_lines.append(f"\nWARNINGS ({warning_count}):")
                for i, warning in enumerate(validation_results["warnings"], 1):
                    report_lines.append(f"  {i}. {warning}")
                    
            # Info
            if info_count > 0:
                report_lines.append(f"\nINFORMATION ({info_count}):")
                for i, info in enumerate(validation_results["info"], 1):
                    report_lines.append(f"  {i}. {info}")
                    
            # Conclusion
            report_lines.append("\n" + "=" * 60)
            if error_count == 0:
                report_lines.append("VALIDATION PASSED - Template is ready for use")
            else:
                report_lines.append("VALIDATION FAILED - Please fix errors before using template")
            report_lines.append("=" * 60)
            
            return "\n".join(report_lines)
            
        except Exception as e:
            return f"Error generating validation report: {str(e)}"