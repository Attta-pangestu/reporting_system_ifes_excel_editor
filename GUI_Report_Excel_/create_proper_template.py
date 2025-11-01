#!/usr/bin/env python3
"""
Create Proper Excel Template with Placeholders
Generates an Excel template that matches the JSON mappings with proper {{placeholder}} format
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_excel_template():
    """Create Excel template with proper placeholders"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Kinerja FFB"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5282', end_color='2C5282', fill_type='solid')
    
    title_font = Font(name='Arial', size=14, bold=True)
    subtitle_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Report Header Section
    ws['A1'] = 'LAPORAN KINERJA FRESH FRUIT BUNCH (FFB) SCANNER'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:H1')
    
    ws['A2'] = 'Periode: {{PERIOD_TEXT}}'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center_alignment
    ws.merge_cells('A2:H2')
    
    ws['A3'] = 'Estate: {{ESTATE_NAME}}'
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = left_alignment
    
    ws['E3'] = 'Tanggal Generate: {{REPORT_DATE}} {{REPORT_TIME}}'
    ws['E3'].font = normal_font
    ws['E3'].alignment = left_alignment
    
    # Summary Section
    ws['A5'] = 'RINGKASAN KINERJA'
    ws['A5'].font = subtitle_font
    ws.merge_cells('A5:B5')
    
    # Summary data placeholders
    summary_data = [
        ('Total Transaksi:', '{{TRANSACTION_COUNT}}'),
        ('Divisi:', '{{DIVISION_NAME}}'),
        ('Persentase Verifikasi:', '{{VERIFICATION_PERCENTAGE}}%'),
        ('Generated:', '{{GENERATED_TIMESTAMP}}')
    ]
    
    for i, (label, placeholder) in enumerate(summary_data, start=6):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = normal_font
        ws[f'B{i}'] = placeholder
        ws[f'B{i}'].font = normal_font
    
    # Main Data Table Headers
    headers = [
        'ESTATE', 'DIVISI', 'KARYAWAN', 'ROLE', 
        'JUMLAH TRANSAKSI', 'PERSENTASE TERVERIFIKASI', 
        'KETERANGAN PERBEDAAN', 'PERIODE LAPORAN'
    ]
    
    header_row = 12
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Data placeholders - these will be replaced with actual data
    data_start_row = 13
    
    # Employee data placeholder
    ws[f'A{data_start_row}'] = '{{EMPLOYEE_DATA}}'
    ws[f'A{data_start_row}'].font = normal_font
    
    # Transaction data placeholder  
    ws[f'A{data_start_row + 2}'] = '{{TRANSACTION_DATA}}'
    ws[f'A{data_start_row + 2}'].font = normal_font
    
    # Kerani data placeholder
    ws[f'A{data_start_row + 4}'] = '{{KERANI_DATA}}'
    ws[f'A{data_start_row + 4}'].font = normal_font
    
    # Mandor data placeholder
    ws[f'A{data_start_row + 6}'] = '{{MANDOR_DATA}}'
    ws[f'A{data_start_row + 6}'].font = normal_font
    
    # Asisten data placeholder
    ws[f'A{data_start_row + 8}'] = '{{ASISTEN_DATA}}'
    ws[f'A{data_start_row + 8}'].font = normal_font
    
    # Division data placeholder
    ws[f'A{data_start_row + 10}'] = '{{DIVISION_DATA}}'
    ws[f'A{data_start_row + 10}'].font = normal_font
    
    # Individual field placeholders for data mapping
    placeholder_row = data_start_row + 15
    individual_placeholders = [
        '{{EMPLOYEE_NAME}}', '{{ROLE_TYPE}}', '{{DIFFERENCE_DESCRIPTION}}'
    ]
    
    for i, placeholder in enumerate(individual_placeholders):
        ws[f'A{placeholder_row + i}'] = placeholder
        ws[f'A{placeholder_row + i}'].font = normal_font
    
    # Set column widths
    column_widths = [15, 20, 25, 12, 18, 22, 30, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add some sample data rows to show structure
    sample_data_row = data_start_row + 20
    ws[f'A{sample_data_row}'] = 'Sample Data (will be replaced):'
    ws[f'A{sample_data_row}'].font = Font(italic=True, color='666666')
    
    sample_row = sample_data_row + 1
    sample_values = [
        '{{ESTATE_NAME}}', '{{DIVISION_NAME}}', '{{EMPLOYEE_NAME}}', 
        '{{ROLE_TYPE}}', '{{TRANSACTION_COUNT}}', '{{VERIFICATION_PERCENTAGE}}', 
        '{{DIFFERENCE_DESCRIPTION}}', '{{PERIOD_TEXT}}'
    ]
    
    for col, value in enumerate(sample_values, start=1):
        cell = ws.cell(row=sample_row, column=col, value=value)
        cell.font = Font(size=9, italic=True, color='666666')
        cell.border = thin_border
    
    # Save the template
    template_path = 'templates/laporan_kinerja_template.xlsx'
    
    # Create templates directory if it doesn't exist
    os.makedirs('templates', exist_ok=True)
    
    wb.save(template_path)
    print(f"Excel template created successfully: {template_path}")
    
    # Print summary of placeholders added
    placeholders_added = [
        '{{PERIOD_TEXT}}', '{{ESTATE_NAME}}', '{{REPORT_DATE}}', '{{REPORT_TIME}}',
        '{{TRANSACTION_COUNT}}', '{{DIVISION_NAME}}', '{{VERIFICATION_PERCENTAGE}}',
        '{{GENERATED_TIMESTAMP}}', '{{EMPLOYEE_DATA}}', '{{TRANSACTION_DATA}}',
        '{{KERANI_DATA}}', '{{MANDOR_DATA}}', '{{ASISTEN_DATA}}', '{{DIVISION_DATA}}',
        '{{EMPLOYEE_NAME}}', '{{ROLE_TYPE}}', '{{DIFFERENCE_DESCRIPTION}}'
    ]
    
    print(f"\nPlaceholders added to template: {len(placeholders_added)}")
    for placeholder in placeholders_added:
        print(f"  - {placeholder}")
    
    return template_path

if __name__ == "__main__":
    create_excel_template()