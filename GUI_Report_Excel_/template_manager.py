#!/usr/bin/env python3
"""
Template Manager
Mengelola template Excel dan JSON untuk sistem report generator

Fungsi utama:
- Memuat dan memvalidasi template Excel + JSON
- Menyediakan informasi template yang tersedia
- Memastikan konsistensi antara file Excel dan JSON
- Validasi kompatibilitas dengan struktur database

Author: AI Assistant
Date: 2025-10-31
"""

import os
import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

class TemplateManager:
    def __init__(self, templates_dir: str = None):
        """
        Initialize Template Manager
        
        Args:
            templates_dir: Directory containing templates (default: ./templates)
        """
        if templates_dir is None:
            self.templates_dir = Path(__file__).parent / "templates"
        else:
            self.templates_dir = Path(templates_dir)
            
        # Ensure templates directory exists
        self.templates_dir.mkdir(exist_ok=True)
        
        # Setup logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
    def get_available_templates(self) -> List[Dict[str, Any]]:
        """
        Get list of available templates
        
        Returns:
            List of template information dictionaries
        """
        templates = []
        
        try:
            # Scan for Excel files in templates directory
            for excel_file in self.templates_dir.glob("*.xlsx"):
                # Check if corresponding JSON file exists
                json_file = excel_file.with_suffix('.json')
                
                if json_file.exists():
                    try:
                        template_info = self._load_template_info(excel_file, json_file)
                        if template_info:
                            templates.append(template_info)
                    except Exception as e:
                        self.logger.warning(f"Error loading template {excel_file.name}: {str(e)}")
                        
        except Exception as e:
            self.logger.error(f"Error scanning templates directory: {str(e)}")
            
        return templates
        
    def get_template_info(self, template_name: str) -> Optional[Dict[str, Any]]:
        """
        Get detailed information for a specific template
        
        Args:
            template_name: Name of the template
            
        Returns:
            Template information dictionary or None if not found
        """
        templates = self.get_available_templates()
        
        for template in templates:
            if template['name'] == template_name:
                return template
                
        return None
        
    def validate_template(self, template_name: str) -> Dict[str, Any]:
        """
        Validate template consistency and compatibility
        
        Args:
            template_name: Name of the template to validate
            
        Returns:
            Validation result dictionary
        """
        result = {
            'valid': False,
            'errors': [],
            'warnings': [],
            'info': {}
        }
        
        try:
            template_info = self.get_template_info(template_name)
            if not template_info:
                result['errors'].append(f"Template '{template_name}' not found")
                return result
                
            # Validate Excel file
            excel_validation = self._validate_excel_file(template_info['excel_path'])
            result['errors'].extend(excel_validation.get('errors', []))
            result['warnings'].extend(excel_validation.get('warnings', []))
            
            # Validate JSON file
            json_validation = self._validate_json_file(template_info['json_path'])
            result['errors'].extend(json_validation.get('errors', []))
            result['warnings'].extend(json_validation.get('warnings', []))
            
            # Validate consistency between Excel and JSON
            consistency_validation = self._validate_consistency(template_info)
            result['errors'].extend(consistency_validation.get('errors', []))
            result['warnings'].extend(consistency_validation.get('warnings', []))
            
            # Set valid flag
            result['valid'] = len(result['errors']) == 0
            result['info'] = template_info
            
        except Exception as e:
            result['errors'].append(f"Validation error: {str(e)}")
            
        return result
        
    def create_template_structure(self, template_name: str) -> Dict[str, str]:
        """
        Create basic template structure (Excel + JSON files)
        
        Args:
            template_name: Name for the new template
            
        Returns:
            Dictionary with paths to created files
        """
        try:
            # Create file paths
            excel_path = self.templates_dir / f"{template_name}.xlsx"
            json_path = self.templates_dir / f"{template_name}.json"
            
            # Create basic Excel template
            self._create_basic_excel_template(excel_path)
            
            # Create basic JSON configuration
            self._create_basic_json_template(json_path, template_name)
            
            return {
                'excel_file': str(excel_path),
                'json_file': str(json_path),
                'template_name': template_name
            }
            
        except Exception as e:
            raise Exception(f"Error creating template structure: {str(e)}")
            
    def _load_template_info(self, excel_file: Path, json_file: Path) -> Optional[Dict[str, Any]]:
        """Load template information from Excel and JSON files"""
        try:
            # Load JSON configuration
            with open(json_file, 'r', encoding='utf-8') as f:
                json_config = json.load(f)
                
            # Basic template info
            template_info = {
                'name': excel_file.stem,
                'excel_file': excel_file.name,
                'json_file': json_file.name,
                'excel_path': str(excel_file),
                'json_path': str(json_file),
                'description': json_config.get('description', ''),
                'version': json_config.get('version', '1.0'),
                'queries': json_config.get('queries', []),
                'transformations': json_config.get('transformations', []),
                'mappings': json_config.get('mappings', {}),
                'config': json_config
            }
            
            return template_info
            
        except Exception as e:
            self.logger.error(f"Error loading template info: {str(e)}")
            return None
            
    def _validate_excel_file(self, excel_path: str) -> Dict[str, List[str]]:
        """Validate Excel file structure"""
        result = {'errors': [], 'warnings': []}
        
        try:
            # Check if file exists and is readable
            if not os.path.exists(excel_path):
                result['errors'].append(f"Excel file not found: {excel_path}")
                return result
                
            # Try to open Excel file
            workbook = openpyxl.load_workbook(excel_path)
            
            # Check if workbook has at least one worksheet
            if len(workbook.worksheets) == 0:
                result['errors'].append("Excel file has no worksheets")
                
            # Check for data placeholders (cells with specific patterns)
            for sheet in workbook.worksheets:
                # Look for placeholder patterns like {{DATA}}, {{SUMMARY}}, etc.
                placeholder_found = False
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if '{{' in cell.value and '}}' in cell.value:
                                placeholder_found = True
                                break
                    if placeholder_found:
                        break
                        
                if not placeholder_found:
                    result['warnings'].append(f"No data placeholders found in sheet '{sheet.title}'")
                    
            workbook.close()
            
        except Exception as e:
            result['errors'].append(f"Error validating Excel file: {str(e)}")
            
        return result
        
    def _validate_json_file(self, json_path: str) -> Dict[str, List[str]]:
        """Validate JSON configuration file"""
        result = {'errors': [], 'warnings': []}
        
        try:
            # Check if file exists
            if not os.path.exists(json_path):
                result['errors'].append(f"JSON file not found: {json_path}")
                return result
                
            # Try to parse JSON
            with open(json_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                
            # Validate required fields
            required_fields = ['name', 'description', 'queries']
            for field in required_fields:
                if field not in config:
                    result['errors'].append(f"Missing required field in JSON: {field}")
                    
            # Validate queries structure
            if 'queries' in config:
                if not isinstance(config['queries'], list):
                    result['errors'].append("'queries' must be a list")
                else:
                    for i, query in enumerate(config['queries']):
                        if not isinstance(query, dict):
                            result['errors'].append(f"Query {i} must be a dictionary")
                        else:
                            if 'name' not in query:
                                result['errors'].append(f"Query {i} missing 'name' field")
                            if 'sql' not in query:
                                result['errors'].append(f"Query {i} missing 'sql' field")
                                
            # Validate transformations if present
            if 'transformations' in config:
                if not isinstance(config['transformations'], list):
                    result['warnings'].append("'transformations' should be a list")
                    
            # Validate mappings if present
            if 'mappings' in config:
                if not isinstance(config['mappings'], dict):
                    result['warnings'].append("'mappings' should be a dictionary")
                    
        except json.JSONDecodeError as e:
            result['errors'].append(f"Invalid JSON format: {str(e)}")
        except Exception as e:
            result['errors'].append(f"Error validating JSON file: {str(e)}")
            
        return result
        
    def _validate_consistency(self, template_info: Dict[str, Any]) -> Dict[str, List[str]]:
        """Validate consistency between Excel and JSON files"""
        result = {'errors': [], 'warnings': []}
        
        try:
            # Load Excel file
            workbook = openpyxl.load_workbook(template_info['excel_path'])
            
            # Get all placeholders from Excel
            excel_placeholders = set()
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # Extract placeholders like {{PLACEHOLDER}}
                            import re
                            placeholders = re.findall(r'\{\{([^}]+)\}\}', cell.value)
                            excel_placeholders.update(placeholders)
                            
            # Get query names from JSON
            json_queries = set()
            for query in template_info.get('queries', []):
                json_queries.add(query.get('name', ''))
                
            # Check if all Excel placeholders have corresponding queries
            for placeholder in excel_placeholders:
                if placeholder not in json_queries and placeholder not in ['DATE', 'TIME', 'TIMESTAMP']:
                    result['warnings'].append(f"Excel placeholder '{placeholder}' has no corresponding query in JSON")
                    
            # Check if all queries have corresponding placeholders
            for query_name in json_queries:
                if query_name and query_name not in excel_placeholders:
                    result['warnings'].append(f"JSON query '{query_name}' has no corresponding placeholder in Excel")
                    
            workbook.close()
            
        except Exception as e:
            result['errors'].append(f"Error validating consistency: {str(e)}")
            
        return result
        
    def _create_basic_excel_template(self, excel_path: Path):
        """Create a basic Excel template with placeholders"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Report"
        
        # Add basic structure
        sheet['A1'] = "LAPORAN DATA"
        sheet['A2'] = "Tanggal: {{DATE}}"
        sheet['A4'] = "Data Utama:"
        sheet['A5'] = "{{MAIN_DATA}}"
        
        sheet['A10'] = "Ringkasan:"
        sheet['A11'] = "{{SUMMARY_DATA}}"
        
        # Save workbook
        workbook.save(excel_path)
        workbook.close()
        
    def _create_basic_json_template(self, json_path: Path, template_name: str):
        """Create a basic JSON configuration template"""
        config = {
            "name": template_name,
            "description": f"Template {template_name}",
            "version": "1.0",
            "queries": [
                {
                    "name": "MAIN_DATA",
                    "description": "Query data utama",
                    "sql": "SELECT * FROM your_table WHERE condition = ?",
                    "parameters": [],
                    "output_format": "table"
                },
                {
                    "name": "SUMMARY_DATA", 
                    "description": "Query data ringkasan",
                    "sql": "SELECT COUNT(*) as total FROM your_table",
                    "parameters": [],
                    "output_format": "single_value"
                }
            ],
            "transformations": [
                {
                    "name": "format_numbers",
                    "description": "Format angka dengan pemisah ribuan",
                    "type": "number_format",
                    "format": "#,##0"
                }
            ],
            "mappings": {
                "DATE": "current_date",
                "TIME": "current_time",
                "TIMESTAMP": "current_timestamp"
            }
        }
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
            
    def get_template_queries(self, template_name: str) -> List[Dict[str, Any]]:
        """Get all queries for a specific template"""
        template_info = self.get_template_info(template_name)
        if template_info:
            return template_info.get('queries', [])
        return []
        
    def get_template_mappings(self, template_name: str) -> Dict[str, Any]:
        """Get all mappings for a specific template"""
        template_info = self.get_template_info(template_name)
        if template_info:
            return template_info.get('mappings', {})
        return {}
        
    def get_template_transformations(self, template_name: str) -> List[Dict[str, Any]]:
        """Get all transformations for a specific template"""
        template_info = self.get_template_info(template_name)
        if template_info:
            return template_info.get('transformations', [])
        return []