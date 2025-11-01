#!/usr/bin/env python3
"""
Create Sample Excel Template
Script untuk membuat template Excel contoh

Author: AI Assistant
Date: 2025-10-31
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_sample_template():
    """Create sample Excel template with placeholders"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan FFB"

    # Set up styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=12)
    normal_font = Font(size=10)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Report Header
    ws['A1'] = "LAPORAN KINERJA FFB SCANNER"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')

    # Report Info
    ws['A3'] = "Tanggal Laporan:"
    ws['B3'] = "{{REPORT_DATE}}"
    ws['A4'] = "Waktu Generate:"
    ws['B4'] = "{{REPORT_TIME}}"
    ws['A5'] = "Periode Data:"
    ws['B5'] = "{{DATE_RANGE}}"

    # Summary Section
    ws['A7'] = "RINGKASAN DATA"
    ws['A7'].font = subheader_font
    ws['A7'].fill = subheader_fill
    ws.merge_cells('A7:B7')

    ws['A8'] = "Total Transaksi:"
    ws['B8'] = "{{TOTAL_TRANSACTIONS}}"
    ws['A9'] = "Total Pekerja:"
    ws['B9'] = "{{TOTAL_WORKERS}}"
    ws['A10'] = "Total Estate:"
    ws['B10'] = "{{TOTAL_ESTATES}}"
    ws['A11'] = "Hari Unik:"
    ws['B11'] = "{{UNIQUE_DAYS}}"

    # Main Data Section
    ws['A13'] = "DATA UTAMA"
    ws['A13'].font = subheader_font
    ws['A13'].fill = subheader_fill
    ws.merge_cells('A13:F13')

    # Table Headers
    headers = ['Worker ID', 'Nama Pekerja', 'Status Transaksi', 'Jumlah Transaksi', 'Tanggal', 'Estate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=14, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data placeholder
    ws['A15'] = "{{MAIN_DATA_TABLE}}"

    # Estate Summary Section
    ws['A25'] = "RINGKASAN PER ESTATE"
    ws['A25'].font = subheader_font
    ws['A25'].fill = subheader_fill
    ws.merge_cells('A25:D25')

    # Estate Summary Headers
    estate_headers = ['Estate', 'Total Records', 'Unique Workers', 'Verified Count']
    for col, header in enumerate(estate_headers, 1):
        cell = ws.cell(row=26, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Estate data placeholder
    ws['A27'] = "{{ESTATE_SUMMARY}}"

    # Status Breakdown Section
    ws['A35'] = "BREAKDOWN STATUS TRANSAKSI"
    ws['A35'].font = subheader_font
    ws['A35'].fill = subheader_fill
    ws.merge_cells('A35:B35')

    # Status Headers
    status_headers = ['Status', 'Jumlah']
    for col, header in enumerate(status_headers, 1):
        cell = ws.cell(row=36, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Status data placeholder
    ws['A37'] = "{{STATUS_BREAKDOWN}}"

    # Footer
    ws['A45'] = "Generated: {{GENERATED_TIMESTAMP}}"
    ws['A45'].font = Font(size=8, italic=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the template
    template_path = os.path.join(os.path.dirname(__file__), "templates", "sample_template.xlsx")
    wb.save(template_path)
    wb.close()

    print(f"Sample Excel template created successfully at: {template_path}")
    return template_path

if __name__ == "__main__":
    create_sample_template()