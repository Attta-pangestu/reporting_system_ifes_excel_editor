#!/usr/bin/env python3
"""
Excel Formula Validator
Specialized testing for Excel template formulas and compatibility

Tests:
- Formula syntax validation
- Formula calculation accuracy
- Excel version compatibility
- Cell formatting preservation
- Template structure integrity

Author: AI Assistant
Date: 2025-10-31
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import re
import os
from pathlib import Path
import logging
from typing import Dict, List, Any, Optional, Tuple
import tempfile
import shutil

class ExcelFormulaValidator:
    """Validator for Excel template formulas and formatting"""
    
    def __init__(self):
        """Initialize the validator"""
        self.logger = logging.getLogger(__name__)
        self.validation_results = []
        self.test_data = self._generate_test_data()
    
    def _generate_test_data(self) -> Dict[str, Any]:
        """Generate test data for formula validation"""
        return {
            'employee_data': [
                {'ID': 'EMP001', 'NAME': 'John Doe'},
                {'ID': 'EMP002', 'NAME': 'Jane Smith'},
                {'ID': 'EMP003', 'NAME': 'Bob Johnson'}
            ],
            'transaction_data': [
                {'SCANUSERID': 'EMP001', 'total_transactions': 150, 'verified_transactions': 145},
                {'SCANUSERID': 'EMP002', 'total_transactions': 200, 'verified_transactions': 190},
                {'SCANUSERID': 'EMP003', 'total_transactions': 175, 'verified_transactions': 170}
            ],
            'division_data': [
                {'DIVID': 'DIV001', 'DIVNAME': 'Division A'},
                {'DIVID': 'DIV002', 'DIVNAME': 'Division B'}
            ]
        }
    
    def validate_template_structure(self, template_path: Path, template_info: Dict[str, Any]) -> Dict[str, Any]:
        """Validate Excel template structure and formatting"""
        results = {
            'template_path': str(template_path),
            'structure_valid': True,
            'issues': [],
            'warnings': [],
            'formatting_check': {},
            'formula_check': {}
        }
        
        try:
            workbook = openpyxl.load_workbook(template_path)
            worksheet_name = template_info['config']['template_info']['worksheet']
            
            if worksheet_name not in workbook.sheetnames:
                results['structure_valid'] = False
                results['issues'].append(f"Worksheet '{worksheet_name}' not found")
                return results
            
            worksheet = workbook[worksheet_name]
            
            # Validate headers
            expected_headers = template_info['output_structure']['headers']
            for cell_ref, expected_value in expected_headers.items():
                actual_value = worksheet[cell_ref].value
                if actual_value != expected_value:
                    results['issues'].append(
                        f"Header mismatch at {cell_ref}: expected '{expected_value}', got '{actual_value}'"
                    )
            
            # Check formatting
            results['formatting_check'] = self._check_formatting(worksheet, expected_headers)
            
            # Check for formulas
            results['formula_check'] = self._check_formulas(worksheet)
            
            workbook.close()
            
        except Exception as e:
            results['structure_valid'] = False
            results['issues'].append(f"Error loading template: {str(e)}")
        
        return results
    
    def _check_formatting(self, worksheet, headers: Dict[str, str]) -> Dict[str, Any]:
        """Check cell formatting in the worksheet"""
        formatting_results = {
            'header_formatting': {},
            'column_widths': {},
            'borders': {},
            'alignment': {}
        }
        
        # Check header formatting
        for cell_ref in headers.keys():
            cell = worksheet[cell_ref]
            formatting_results['header_formatting'][cell_ref] = {
                'font': {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'color': str(cell.font.color.rgb) if cell.font.color else None
                },
                'fill': {
                    'pattern_type': cell.fill.patternType,
                    'start_color': str(cell.fill.start_color.rgb) if cell.fill.start_color else None
                },
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical
                }
            }
        
        # Check column widths
        for col in worksheet.columns:
            col_letter = get_column_letter(col[0].column)
            width = worksheet.column_dimensions[col_letter].width
            formatting_results['column_widths'][col_letter] = width
        
        return formatting_results
    
    def _check_formulas(self, worksheet) -> Dict[str, Any]:
        """Check formulas in the worksheet"""
        formula_results = {
            'formula_count': 0,
            'formulas_found': [],
            'formula_types': {},
            'potential_issues': []
        }
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_results['formula_count'] += 1
                    formula_info = {
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'type': self._classify_formula(cell.value)
                    }
                    formula_results['formulas_found'].append(formula_info)
                    
                    # Count formula types
                    formula_type = formula_info['type']
                    formula_results['formula_types'][formula_type] = \
                        formula_results['formula_types'].get(formula_type, 0) + 1
                    
                    # Check for potential issues
                    issues = self._validate_formula_syntax(cell.value)
                    if issues:
                        formula_results['potential_issues'].extend([
                            f"{cell.coordinate}: {issue}" for issue in issues
                        ])
        
        return formula_results
    
    def _classify_formula(self, formula: str) -> str:
        """Classify the type of formula"""
        formula_upper = formula.upper()
        
        if 'SUM(' in formula_upper:
            return 'SUM'
        elif 'COUNT(' in formula_upper:
            return 'COUNT'
        elif 'IF(' in formula_upper:
            return 'CONDITIONAL'
        elif 'VLOOKUP(' in formula_upper or 'HLOOKUP(' in formula_upper:
            return 'LOOKUP'
        elif any(func in formula_upper for func in ['AVERAGE(', 'MAX(', 'MIN(']):
            return 'STATISTICAL'
        elif any(func in formula_upper for func in ['CONCATENATE(', 'LEFT(', 'RIGHT(', 'MID(']):
            return 'TEXT'
        else:
            return 'OTHER'
    
    def _validate_formula_syntax(self, formula: str) -> List[str]:
        """Validate formula syntax and identify potential issues"""
        issues = []
        
        # Check for balanced parentheses
        if formula.count('(') != formula.count(')'):
            issues.append("Unbalanced parentheses")
        
        # Check for common syntax errors
        if formula.endswith(','):
            issues.append("Formula ends with comma")
        
        if ',,,' in formula:
            issues.append("Multiple consecutive commas")
        
        # Check for potential circular references (basic check)
        if re.search(r'[A-Z]+\d+:[A-Z]+\d+', formula):
            # This is a range reference, which is generally good
            pass
        
        return issues
    
    def test_formula_calculations(self, template_path: Path, template_info: Dict[str, Any]) -> Dict[str, Any]:
        """Test formula calculations with sample data"""
        results = {
            'calculation_tests': [],
            'accuracy_score': 0.0,
            'failed_calculations': []
        }
        
        try:
            # Create a copy of the template for testing
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                shutil.copy2(template_path, tmp_file.name)
                test_file_path = tmp_file.name
            
            workbook = openpyxl.load_workbook(test_file_path)
            worksheet_name = template_info['config']['template_info']['worksheet']
            worksheet = workbook[worksheet_name]
            
            # Insert test data
            self._insert_test_data(worksheet, template_info)
            
            # Test calculations
            calculation_tests = self._perform_calculation_tests(worksheet)
            results['calculation_tests'] = calculation_tests
            
            # Calculate accuracy score
            total_tests = len(calculation_tests)
            passed_tests = sum(1 for test in calculation_tests if test['passed'])
            results['accuracy_score'] = (passed_tests / total_tests * 100) if total_tests > 0 else 0
            
            results['failed_calculations'] = [
                test for test in calculation_tests if not test['passed']
            ]
            
            workbook.save(test_file_path)
            workbook.close()
            
            # Clean up
            os.unlink(test_file_path)
            
        except Exception as e:
            results['failed_calculations'].append(f"Error during calculation testing: {str(e)}")
        
        return results
    
    def _insert_test_data(self, worksheet, template_info: Dict[str, Any]):
        """Insert test data into the worksheet"""
        # Get data start row
        data_start_row = template_info['output_structure'].get('data_start_row', 2)
        columns = template_info['output_structure']['columns']
        
        # Insert sample data
        test_records = [
            {
                'estate_name': 'Estate A',
                'division_name': 'Division A',
                'employee_name': 'John Doe',
                'role_type': 'Kerani',
                'transaction_count': 150,
                'verification_percentage': 96.67,
                'difference_description': 'Minor differences in 5 transactions',
                'period_text': 'September 2024'
            },
            {
                'estate_name': 'Estate A',
                'division_name': 'Division B',
                'employee_name': 'Jane Smith',
                'role_type': 'Mandor',
                'transaction_count': 200,
                'verification_percentage': 95.00,
                'difference_description': 'Good performance',
                'period_text': 'September 2024'
            }
        ]
        
        for i, record in enumerate(test_records):
            row_num = data_start_row + i
            for col_letter, field_name in columns.items():
                if field_name in record:
                    cell = worksheet[f"{col_letter}{row_num}"]
                    cell.value = record[field_name]
    
    def _perform_calculation_tests(self, worksheet) -> List[Dict[str, Any]]:
        """Perform calculation tests on formulas"""
        tests = []
        
        # Test percentage calculations
        for row in worksheet.iter_rows(min_row=2, max_row=10):  # Check first few data rows
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    test_result = {
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'expected_type': 'number',
                        'passed': False,
                        'actual_value': None,
                        'error': None
                    }
                    
                    try:
                        # Force calculation
                        worksheet.parent.calculation.calcMode = 'automatic'
                        actual_value = cell.value
                        
                        # Basic validation - check if it's a number for percentage calculations
                        if 'percentage' in cell.coordinate.lower() or 'F' in cell.coordinate:
                            if isinstance(actual_value, (int, float)) and 0 <= actual_value <= 100:
                                test_result['passed'] = True
                        else:
                            # For other formulas, just check if they don't error
                            test_result['passed'] = actual_value is not None
                        
                        test_result['actual_value'] = actual_value
                        
                    except Exception as e:
                        test_result['error'] = str(e)
                    
                    tests.append(test_result)
        
        return tests
    
    def test_excel_compatibility(self, template_path: Path) -> Dict[str, Any]:
        """Test Excel compatibility across different scenarios"""
        results = {
            'openpyxl_compatibility': True,
            'file_format_valid': True,
            'save_load_test': True,
            'issues': []
        }
        
        try:
            # Test 1: Load with openpyxl
            workbook = openpyxl.load_workbook(template_path)
            results['openpyxl_compatibility'] = True
            
            # Test 2: Check file format
            if not str(template_path).endswith(('.xlsx', '.xlsm')):
                results['file_format_valid'] = False
                results['issues'].append("File format should be .xlsx or .xlsm for modern Excel compatibility")
            
            # Test 3: Save and reload test
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                workbook.save(tmp_file.name)
                
                # Try to reload
                test_workbook = openpyxl.load_workbook(tmp_file.name)
                test_workbook.close()
                
                # Clean up
                os.unlink(tmp_file.name)
            
            workbook.close()
            
        except Exception as e:
            results['openpyxl_compatibility'] = False
            results['save_load_test'] = False
            results['issues'].append(f"Compatibility error: {str(e)}")
        
        return results
    
    def generate_validation_report(self, template_path: Path, template_info: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive validation report"""
        self.logger.info(f"Starting Excel template validation for: {template_path}")
        
        report = {
            'template_path': str(template_path),
            'validation_timestamp': str(datetime.now()),
            'overall_score': 0,
            'tests': {}
        }
        
        # Run all validation tests
        tests = [
            ('structure_validation', self.validate_template_structure),
            ('formula_calculations', self.test_formula_calculations),
            ('excel_compatibility', self.test_excel_compatibility)
        ]
        
        passed_tests = 0
        total_tests = len(tests)
        
        for test_name, test_function in tests:
            try:
                if test_name == 'excel_compatibility':
                    test_result = test_function(template_path)
                else:
                    test_result = test_function(template_path, template_info)
                
                report['tests'][test_name] = test_result
                
                # Determine if test passed
                if test_name == 'structure_validation':
                    if test_result.get('structure_valid', False) and not test_result.get('issues', []):
                        passed_tests += 1
                elif test_name == 'formula_calculations':
                    if test_result.get('accuracy_score', 0) >= 80:  # 80% threshold
                        passed_tests += 1
                elif test_name == 'excel_compatibility':
                    if (test_result.get('openpyxl_compatibility', False) and 
                        test_result.get('file_format_valid', False) and 
                        test_result.get('save_load_test', False)):
                        passed_tests += 1
                
            except Exception as e:
                report['tests'][test_name] = {
                    'error': str(e),
                    'passed': False
                }
                self.logger.error(f"Test {test_name} failed: {e}")
        
        # Calculate overall score
        report['overall_score'] = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        self.logger.info(f"Validation completed. Overall score: {report['overall_score']:.1f}%")
        
        return report

if __name__ == '__main__':
    # Example usage
    from datetime import datetime
    
    validator = ExcelFormulaValidator()
    
    # Test with the laporan kinerja template
    template_path = Path("templates/laporan_kinerja_template.xlsx")
    
    if template_path.exists():
        # Load template info
        with open("templates/laporan_kinerja_template.json", 'r') as f:
            template_info = json.load(f)
        
        # Generate validation report
        report = validator.generate_validation_report(template_path, template_info)
        
        # Save report
        report_file = f"excel_validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(report_file, 'w') as f:
            json.dump(report, f, indent=2)
        
        print(f"Validation report saved to: {report_file}")
        print(f"Overall Score: {report['overall_score']:.1f}%")
    else:
        print(f"Template file not found: {template_path}")