#!/usr/bin/env python3
"""
Report Processor
Memproses template Excel dengan data dari database

Fungsi utama:
- Memproses template Excel dengan data dari database
- Menerapkan transformasi data sesuai konfigurasi JSON
- Menghasilkan report final dengan format yang dipertahankan
- Menangani placeholder dan mapping data

Author: AI Assistant
Date: 2025-10-31
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Union
import logging
from pathlib import Path
import time
import uuid

class ReportProcessor:
    def __init__(self):
        """Initialize Report Processor with enhanced debug logging"""
        # Setup comprehensive logging
        self.setup_debug_logging()
        
        # Processing metrics
        self.processing_metrics = {
            'session_id': None,
            'start_time': None,
            'end_time': None,
            'total_duration': 0,
            'stages_completed': 0,
            'stages_failed': 0,
            'queries_executed': 0,
            'placeholders_processed': 0,
            'worksheets_processed': 0,
            'data_processed': {
                'total_records': 0,
                'query_results': {},
                'placeholder_replacements': 0,
                'formatting_operations': 0
            }
        }
        
    def setup_debug_logging(self):
        """Setup comprehensive debug logging system"""
        # Create logs directory if it doesn't exist
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        # Setup main logger
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.DEBUG)
        
        # Clear existing handlers
        self.logger.handlers.clear()
        
        # Create formatters
        detailed_formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - [%(funcName)s:%(lineno)d] - %(message)s'
        )
        
        # Console handler for INFO and above
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(detailed_formatter)
        self.logger.addHandler(console_handler)
        
        # File handler for all DEBUG messages
        debug_file_handler = logging.FileHandler(
            log_dir / f"report_processor_debug_{datetime.now().strftime('%Y%m%d')}.log",
            encoding='utf-8'
        )
        debug_file_handler.setLevel(logging.DEBUG)
        debug_file_handler.setFormatter(detailed_formatter)
        self.logger.addHandler(debug_file_handler)
        
        # Separate handler for rendering process
        self.rendering_logger = logging.getLogger(f"{__name__}.rendering")
        self.rendering_logger.setLevel(logging.DEBUG)
        
        rendering_handler = logging.FileHandler(
            log_dir / f"template_rendering_{datetime.now().strftime('%Y%m%d')}.log",
            encoding='utf-8'
        )
        rendering_handler.setLevel(logging.DEBUG)
        rendering_handler.setFormatter(detailed_formatter)
        self.rendering_logger.addHandler(rendering_handler)
        
        # Query execution logger
        self.query_logger = logging.getLogger(f"{__name__}.queries")
        self.query_logger.setLevel(logging.DEBUG)
        
        query_handler = logging.FileHandler(
            log_dir / f"query_execution_{datetime.now().strftime('%Y%m%d')}.log",
            encoding='utf-8'
        )
        query_handler.setLevel(logging.DEBUG)
        query_handler.setFormatter(detailed_formatter)
        self.query_logger.addHandler(query_handler)
        
    def generate_report(self, template_info: Dict[str, Any], 
                       db_connector, output_path: str) -> str:
        """
        Generate report based on template and database data with comprehensive debug logging
        
        Args:
            template_info: Template information dictionary
            db_connector: Database connector instance
            output_path: Output directory path
            
        Returns:
            Path to generated report file
        """
        # Initialize processing session
        session_id = str(uuid.uuid4())[:8]
        self.processing_metrics['session_id'] = session_id
        self.processing_metrics['start_time'] = time.time()
        
        self.logger.info(f"=== REPORT GENERATION SESSION START: {session_id} ===")
        self.logger.info(f"Template: {template_info.get('name', 'Unknown')}")
        self.logger.info(f"Output path: {output_path}")
        self.logger.debug(f"Template info keys: {list(template_info.keys())}")
        
        try:
            # Stage 1: Load template Excel file
            stage_start = time.time()
            self.logger.info("=== STAGE 1: Loading Template Excel File ===")
            
            excel_path = template_info.get('excel_path')
            if not excel_path or not os.path.exists(excel_path):
                raise FileNotFoundError(f"Template Excel file not found: {excel_path}")
                
            self.logger.debug(f"Loading Excel file: {excel_path}")
            workbook = openpyxl.load_workbook(excel_path)
            
            worksheet_names = workbook.sheetnames
            self.logger.info(f"✓ Template loaded successfully - {len(worksheet_names)} worksheets")
            self.logger.debug(f"Worksheets: {worksheet_names}")
            
            stage_duration = time.time() - stage_start
            self.logger.info(f"Stage 1 completed in {stage_duration:.2f} seconds")
            self.processing_metrics['stages_completed'] += 1
            
            # Stage 2: Execute queries and get data
            stage_start = time.time()
            self.logger.info("=== STAGE 2: Query Execution and Data Retrieval ===")
            
            query_results = self._execute_template_queries_enhanced(template_info, db_connector)
            
            stage_duration = time.time() - stage_start
            self.logger.info(f"Stage 2 completed in {stage_duration:.2f} seconds")
            self.processing_metrics['stages_completed'] += 1
            
            # Stage 3: Process each worksheet
            stage_start = time.time()
            self.logger.info("=== STAGE 3: Worksheet Processing and Template Rendering ===")
            
            for i, sheet in enumerate(workbook.worksheets):
                sheet_start = time.time()
                self.rendering_logger.info(f"Processing worksheet {i+1}/{len(workbook.worksheets)}: {sheet.title}")
                
                self._process_worksheet_enhanced(sheet, template_info, query_results)
                
                sheet_duration = time.time() - sheet_start
                self.rendering_logger.info(f"✓ Worksheet '{sheet.title}' processed in {sheet_duration:.2f} seconds")
                self.processing_metrics['worksheets_processed'] += 1
                
            stage_duration = time.time() - stage_start
            self.logger.info(f"Stage 3 completed in {stage_duration:.2f} seconds")
            self.processing_metrics['stages_completed'] += 1
            
            # Stage 4: Generate output and verify
            stage_start = time.time()
            self.logger.info("=== STAGE 4: Output Generation and Verification ===")
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            template_name = template_info.get('name', 'report')
            output_filename = f"{template_name}_{timestamp}.xlsx"
            output_file_path = os.path.join(output_path, output_filename)
            
            self.logger.debug(f"Output file path: {output_file_path}")
            
            # Save processed workbook
            workbook.save(output_file_path)
            workbook.close()
            
            # Verify output file
            if os.path.exists(output_file_path):
                file_size = os.path.getsize(output_file_path)
                self.logger.info(f"✓ Report file saved successfully: {output_file_path}")
                self.logger.debug(f"File size: {file_size:,} bytes")
                
                # Perform output verification
                self._verify_report_output(output_file_path, template_info)
            else:
                raise FileNotFoundError(f"Failed to create output file: {output_file_path}")
            
            stage_duration = time.time() - stage_start
            self.logger.info(f"Stage 4 completed in {stage_duration:.2f} seconds")
            self.processing_metrics['stages_completed'] += 1
            
            # Finalize session
            self._finalize_processing_session()
            
            self.logger.info(f"✓ Report generated successfully: {output_file_path}")
            return output_file_path
            
        except Exception as e:
            self.processing_metrics['stages_failed'] += 1
            error_msg = f"Error generating report: {str(e)}"
            self.logger.error(f"CRITICAL ERROR in session {session_id}: {error_msg}")
            self.logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            self.logger.debug(f"Processing metrics at failure: {self.processing_metrics}")
            raise
            
    def _execute_template_queries_enhanced(self, template_info: Dict[str, Any], 
                                          db_connector) -> Dict[str, Any]:
        """Execute template queries with comprehensive debug logging"""
        self.query_logger.info("=== QUERY EXECUTION PHASE START ===")
        
        queries = template_info.get("queries", [])
        mappings = template_info.get("mappings", {})
        
        self.query_logger.info(f"Found {len(queries)} queries to execute")
        self.query_logger.debug(f"Available mappings: {list(mappings.keys())}")
        
        query_results = {}
        total_records = 0
        
        try:
            for i, query_config in enumerate(queries):
                query_start = time.time()
                query_name = query_config.get("name", f"query_{i}")
                
                self.query_logger.info(f"--- Executing Query {i+1}/{len(queries)}: {query_name} ---")
                
                # Validate query configuration
                if not isinstance(query_config, dict):
                    error_msg = f"Query {i} is not a dictionary: {type(query_config)}"
                    self.query_logger.error(error_msg)
                    raise ValueError(error_msg)
                
                sql = query_config.get("sql", "").strip()
                if not sql:
                    error_msg = f"Query {query_name} has empty SQL"
                    self.query_logger.error(error_msg)
                    raise ValueError(error_msg)
                
                self.query_logger.debug(f"SQL Query ({len(sql)} chars): {sql[:200]}{'...' if len(sql) > 200 else ''}")
                
                # Execute query with timing
                try:
                    self.query_logger.debug(f"Executing query via database connector...")
                    result = db_connector.execute_query(sql)
                    
                    query_duration = time.time() - query_start
                    
                    # Process and validate results
                    if result is None:
                        self.query_logger.warning(f"Query {query_name} returned None")
                        processed_result = []
                        record_count = 0
                    elif isinstance(result, list):
                        processed_result = result
                        record_count = len(result)
                        self.query_logger.info(f"✓ Query {query_name} returned {record_count} records")
                    else:
                        # Handle single value or other types
                        processed_result = result
                        record_count = 1 if result is not None else 0
                        self.query_logger.info(f"✓ Query {query_name} returned single value: {type(result)}")
                    
                    # Log sample data for debugging
                    if isinstance(processed_result, list) and len(processed_result) > 0:
                        sample_record = processed_result[0]
                        if isinstance(sample_record, dict):
                            self.query_logger.debug(f"Sample record keys: {list(sample_record.keys())}")
                            self.query_logger.debug(f"Sample record: {dict(list(sample_record.items())[:3])}")
                        else:
                            self.query_logger.debug(f"Sample record type: {type(sample_record)}, value: {sample_record}")
                    
                    # Store results
                    query_results[query_name] = processed_result
                    total_records += record_count
                    
                    # Update metrics
                    self.processing_metrics['queries_executed'] += 1
                    self.processing_metrics['data_processed']['query_results'][query_name] = {
                        'record_count': record_count,
                        'execution_time': query_duration,
                        'result_type': type(processed_result).__name__
                    }
                    
                    self.query_logger.info(f"✓ Query {query_name} completed in {query_duration:.2f}s - {record_count} records")
                    
                except Exception as query_error:
                    query_duration = time.time() - query_start
                    error_msg = f"Query execution failed for {query_name}: {str(query_error)}"
                    self.query_logger.error(error_msg)
                    self.query_logger.debug(f"Query error details: {type(query_error).__name__}: {str(query_error)}")
                    self.query_logger.debug(f"Failed SQL: {sql}")
                    
                    # Store empty result to prevent downstream errors
                    query_results[query_name] = []
                    self.processing_metrics['data_processed']['query_results'][query_name] = {
                        'record_count': 0,
                        'execution_time': query_duration,
                        'result_type': 'error',
                        'error': str(query_error)
                    }
                    
                    # Continue with other queries instead of failing completely
                    continue
            
            # Process mappings
            self.query_logger.info(f"--- Processing {len(mappings)} mappings ---")
            for mapping_name, mapping_value in mappings.items():
                self.query_logger.debug(f"Mapping '{mapping_name}': {type(mapping_value)} = {str(mapping_value)[:100]}")
                query_results[mapping_name] = mapping_value
            
            # Final summary
            self.processing_metrics['data_processed']['total_records'] = total_records
            self.query_logger.info(f"=== QUERY EXECUTION COMPLETE ===")
            self.query_logger.info(f"Total queries executed: {self.processing_metrics['queries_executed']}")
            self.query_logger.info(f"Total records retrieved: {total_records}")
            self.query_logger.info(f"Query results keys: {list(query_results.keys())}")
            
            return query_results
            
        except Exception as e:
            error_msg = f"Critical error in query execution: {str(e)}"
            self.query_logger.error(error_msg)
            self.query_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            raise
        """Execute all queries defined in template"""
        query_results = {}
        
        try:
            queries = template_info.get('queries', {})
            
            # Handle both dict and list formats
            if isinstance(queries, dict):
                # Template uses dict format: {"query_name": {"sql": "...", ...}}
                for query_name, query_config in queries.items():
                    if isinstance(query_config, dict):
                        sql = query_config.get('sql')
                        parameters = query_config.get('parameters', [])
                        
                        if sql:
                            self.logger.info(f"Executing query: {query_name}")
                            
                            # Execute query
                            results = db_connector.execute_query(sql, parameters)
                            
                            # Apply transformations if specified
                            transformed_results = self._apply_transformations(
                                results, query_config.get('transformations', []), template_info
                            )
                            
                            query_results[query_name] = {
                                'data': transformed_results,
                                'config': query_config,
                                'row_count': len(transformed_results)
                            }
                            
                            self.logger.info(f"Query {query_name} returned {len(transformed_results)} rows")
            else:
                # Handle list format for backward compatibility
                for query in queries:
                    if isinstance(query, dict):
                        query_name = query.get('name')
                        sql = query.get('sql')
                        parameters = query.get('parameters', [])
                        
                        if query_name and sql:
                            self.logger.info(f"Executing query: {query_name}")
                            
                            # Execute query
                            results = db_connector.execute_query(sql, parameters)
                            
                            # Apply transformations if specified
                            transformed_results = self._apply_transformations(
                                results, query.get('transformations', []), template_info
                            )
                            
                            query_results[query_name] = {
                                'data': transformed_results,
                                'config': query,
                                'row_count': len(transformed_results)
                            }
                            
                            self.logger.info(f"Query {query_name} returned {len(transformed_results)} rows")
                    
        except Exception as e:
            self.logger.error(f"Error executing queries: {str(e)}")
            raise
            
        return query_results
        
    def _process_worksheet(self, sheet, template_info: Dict[str, Any], 
                          query_results: Dict[str, Any]):
        """Process a single worksheet"""
        try:
            self.logger.info(f"Processing worksheet: {sheet.title}")
            
            # Find and replace placeholders
            self._replace_placeholders(sheet, template_info, query_results)
            
            # Apply formatting if specified
            self._apply_worksheet_formatting(sheet, template_info)
            
        except Exception as e:
            self.logger.error(f"Error processing worksheet {sheet.title}: {str(e)}")
            raise
            
    def _replace_placeholders(self, sheet, template_info: Dict[str, Any], 
                             query_results: Dict[str, Any]):
        """Replace placeholders in worksheet with actual data"""
        try:
            # Get mappings from template
            mappings = template_info.get('mappings', {})
            
            # Iterate through all cells
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Find placeholders in cell value
                        placeholders = re.findall(r'\{\{([^}]+)\}\}', cell.value)
                        
                        for placeholder in placeholders:
                            replacement_value = self._get_placeholder_value(
                                placeholder, mappings, query_results
                            )
                            
                            if replacement_value is not None:
                                # Replace placeholder with actual value
                                cell.value = cell.value.replace(
                                    f'{{{{{placeholder}}}}}', str(replacement_value)
                                )
                                
        except Exception as e:
            self.logger.error(f"Error replacing placeholders: {str(e)}")
            raise
            
    def _get_placeholder_value(self, placeholder: str, mappings: Dict[str, Any], 
                              query_results: Dict[str, Any]) -> Any:
        """Get replacement value for a placeholder"""
        try:
            # Check if it's a special mapping (DATE, TIME, etc.)
            if placeholder in mappings:
                mapping_type = mappings[placeholder]
                
                if mapping_type == 'current_date':
                    return datetime.now().strftime('%Y-%m-%d')
                elif mapping_type == 'current_time':
                    return datetime.now().strftime('%H:%M:%S')
                elif mapping_type == 'current_timestamp':
                    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
            # Check if it's a query result
            if placeholder in query_results:
                query_result = query_results[placeholder]
                output_format = query_result['config'].get('output_format', 'table')
                
                if output_format == 'single_value':
                    # Return single value
                    data = query_result['data']
                    if data and len(data) > 0:
                        first_row = data[0]
                        if len(first_row) > 0:
                            return list(first_row.values())[0]
                            
                elif output_format == 'table':
                    # Return formatted table
                    return self._format_table_data(query_result['data'])
                    
                elif output_format == 'list':
                    # Return formatted list
                    return self._format_list_data(query_result['data'])
                    
            return None
            
        except Exception as e:
            self.logger.error(f"Error getting placeholder value for {placeholder}: {str(e)}")
            return None
            
    def _format_table_data(self, data: List[Dict[str, Any]]) -> str:
        """Format data as table string"""
        if not data:
            return "Tidak ada data"
            
        try:
            # Get column headers
            headers = list(data[0].keys())
            
            # Calculate column widths
            col_widths = {}
            for header in headers:
                col_widths[header] = len(str(header))
                
            for row in data:
                for header in headers:
                    value_len = len(str(row.get(header, '')))
                    if value_len > col_widths[header]:
                        col_widths[header] = value_len
                        
            # Build table string
            table_lines = []
            
            # Header line
            header_line = " | ".join([
                str(header).ljust(col_widths[header]) for header in headers
            ])
            table_lines.append(header_line)
            
            # Separator line
            separator_line = " | ".join([
                "-" * col_widths[header] for header in headers
            ])
            table_lines.append(separator_line)
            
            # Data lines
            for row in data:
                data_line = " | ".join([
                    str(row.get(header, '')).ljust(col_widths[header]) for header in headers
                ])
                table_lines.append(data_line)
                
            return "\n".join(table_lines)
            
        except Exception as e:
            self.logger.error(f"Error formatting table data: {str(e)}")
            return "Error formatting data"
            
    def _format_list_data(self, data: List[Dict[str, Any]]) -> str:
        """Format data as list string"""
        if not data:
            return "Tidak ada data"
            
        try:
            list_items = []
            for i, row in enumerate(data, 1):
                # Use first column value or all values
                if len(row) == 1:
                    value = list(row.values())[0]
                    list_items.append(f"{i}. {value}")
                else:
                    # Multiple columns - format as key: value pairs
                    row_items = []
                    for key, value in row.items():
                        row_items.append(f"{key}: {value}")
                    list_items.append(f"{i}. {', '.join(row_items)}")
                    
            return "\n".join(list_items)
            
        except Exception as e:
            self.logger.error(f"Error formatting list data: {str(e)}")
            return "Error formatting data"
            
    def _apply_transformations(self, data: List[Dict[str, Any]], 
                              transformations: List[Dict[str, Any]], 
                              template_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Apply data transformations"""
        if not transformations:
            return data
            
        try:
            transformed_data = data.copy()
            
            # Get global transformations from template
            global_transformations = template_info.get('transformations', [])
            all_transformations = transformations + global_transformations
            
            for transformation in all_transformations:
                transform_type = transformation.get('type')
                
                if transform_type == 'number_format':
                    transformed_data = self._apply_number_formatting(
                        transformed_data, transformation
                    )
                elif transform_type == 'date_format':
                    transformed_data = self._apply_date_formatting(
                        transformed_data, transformation
                    )
                elif transform_type == 'string_format':
                    transformed_data = self._apply_string_formatting(
                        transformed_data, transformation
                    )
                elif transform_type == 'filter':
                    transformed_data = self._apply_filter(
                        transformed_data, transformation
                    )
                elif transform_type == 'sort':
                    transformed_data = self._apply_sort(
                        transformed_data, transformation
                    )
                    
            return transformed_data
            
        except Exception as e:
            self.logger.error(f"Error applying transformations: {str(e)}")
            return data
            
    def _apply_number_formatting(self, data: List[Dict[str, Any]], 
                                transformation: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Apply number formatting transformation"""
        try:
            format_str = transformation.get('format', '#,##0')
            columns = transformation.get('columns', [])
            
            for row in data:
                for col_name, value in row.items():
                    if not columns or col_name in columns:
                        if isinstance(value, (int, float)):
                            # Apply basic number formatting
                            if format_str == '#,##0':
                                row[col_name] = f"{value:,}"
                            elif format_str == '#,##0.00':
                                row[col_name] = f"{value:,.2f}"
                                
            return data
            
        except Exception as e:
            self.logger.error(f"Error applying number formatting: {str(e)}")
            return data
            
    def _apply_date_formatting(self, data: List[Dict[str, Any]], 
                              transformation: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Apply date formatting transformation"""
        try:
            format_str = transformation.get('format', '%Y-%m-%d')
            columns = transformation.get('columns', [])
            
            for row in data:
                for col_name, value in row.items():
                    if not columns or col_name in columns:
                        if isinstance(value, (date, datetime)):
                            row[col_name] = value.strftime(format_str)
                            
            return data
            
        except Exception as e:
            self.logger.error(f"Error applying date formatting: {str(e)}")
            return data
            
    def _apply_string_formatting(self, data: List[Dict[str, Any]], 
                                transformation: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Apply string formatting transformation"""
        try:
            format_type = transformation.get('format_type', 'upper')
            columns = transformation.get('columns', [])
            
            for row in data:
                for col_name, value in row.items():
                    if not columns or col_name in columns:
                        if isinstance(value, str):
                            if format_type == 'upper':
                                row[col_name] = value.upper()
                            elif format_type == 'lower':
                                row[col_name] = value.lower()
                            elif format_type == 'title':
                                row[col_name] = value.title()
                                
            return data
            
        except Exception as e:
            self.logger.error(f"Error applying string formatting: {str(e)}")
            return data
            
    def _apply_filter(self, data: List[Dict[str, Any]], 
                     transformation: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Apply filter transformation"""
        try:
            column = transformation.get('column')
            operator = transformation.get('operator', '==')
            value = transformation.get('value')
            
            if not column:
                return data
                
            filtered_data = []
            for row in data:
                row_value = row.get(column)
                
                if operator == '==' and row_value == value:
                    filtered_data.append(row)
                elif operator == '!=' and row_value != value:
                    filtered_data.append(row)
                elif operator == '>' and row_value > value:
                    filtered_data.append(row)
                elif operator == '<' and row_value < value:
                    filtered_data.append(row)
                elif operator == 'contains' and isinstance(row_value, str) and value in row_value:
                    filtered_data.append(row)
                    
            return filtered_data
            
        except Exception as e:
            self.logger.error(f"Error applying filter: {str(e)}")
            return data
            
    def _apply_sort(self, data: List[Dict[str, Any]], 
                   transformation: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Apply sort transformation"""
        try:
            column = transformation.get('column')
            reverse = transformation.get('reverse', False)
            
            if not column:
                return data
                
            return sorted(data, key=lambda x: x.get(column, ''), reverse=reverse)
            
        except Exception as e:
            self.logger.error(f"Error applying sort: {str(e)}")
            return data
            
    def _process_worksheets_enhanced(self, workbook, query_results: Dict[str, Any], 
                                   template_info: Dict[str, Any]) -> None:
        """Process worksheets with comprehensive debug logging"""
        self.rendering_logger.info("=== WORKSHEET PROCESSING PHASE START ===")
        
        worksheets = template_info.get("worksheets", [])
        self.rendering_logger.info(f"Found {len(worksheets)} worksheets to process")
        
        if not worksheets:
            self.rendering_logger.warning("No worksheets defined in template")
            return
        
        try:
            for i, worksheet_config in enumerate(worksheets):
                worksheet_start = time.time()
                worksheet_name = worksheet_config.get("name", f"worksheet_{i}")
                
                self.rendering_logger.info(f"--- Processing Worksheet {i+1}/{len(worksheets)}: {worksheet_name} ---")
                
                # Validate worksheet configuration
                if not isinstance(worksheet_config, dict):
                    error_msg = f"Worksheet {i} configuration is not a dictionary: {type(worksheet_config)}"
                    self.rendering_logger.error(error_msg)
                    continue
                
                # Get worksheet from workbook
                try:
                    if worksheet_name in workbook.sheetnames:
                        worksheet = workbook[worksheet_name]
                        self.rendering_logger.debug(f"Found worksheet '{worksheet_name}' in workbook")
                    else:
                        self.rendering_logger.warning(f"Worksheet '{worksheet_name}' not found in workbook")
                        self.rendering_logger.debug(f"Available worksheets: {workbook.sheetnames}")
                        continue
                        
                except Exception as ws_error:
                    error_msg = f"Error accessing worksheet '{worksheet_name}': {str(ws_error)}"
                    self.rendering_logger.error(error_msg)
                    continue
                
                # Process worksheet data
                try:
                    # Get data source for this worksheet
                    data_source = worksheet_config.get("data_source", "")
                    self.rendering_logger.debug(f"Worksheet data source: '{data_source}'")
                    
                    if data_source and data_source in query_results:
                        worksheet_data = query_results[data_source]
                        self.rendering_logger.info(f"Using data from query '{data_source}': {len(worksheet_data) if isinstance(worksheet_data, list) else 1} records")
                    else:
                        worksheet_data = query_results
                        self.rendering_logger.debug(f"Using all query results as data source")
                    
                    # Scan for placeholders in worksheet
                    placeholders_found = self._scan_worksheet_placeholders(worksheet, worksheet_name)
                    
                    # Process each cell with placeholders
                    cells_processed = 0
                    cells_with_errors = 0
                    
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                original_value = cell.value
                                
                                # Check if cell contains placeholders
                                if any(placeholder in original_value for placeholder in placeholders_found):
                                    try:
                                        # Process cell with template engine
                                        processed_value = self._process_cell_value(
                                            original_value, worksheet_data, cell.coordinate
                                        )
                                        
                                        if processed_value != original_value:
                                            cell.value = processed_value
                                            cells_processed += 1
                                            
                                            self.rendering_logger.debug(
                                                f"Cell {cell.coordinate}: '{original_value[:50]}...' -> '{str(processed_value)[:50]}...'"
                                            )
                                        
                                    except Exception as cell_error:
                                        cells_with_errors += 1
                                        error_msg = f"Error processing cell {cell.coordinate}: {str(cell_error)}"
                                        self.rendering_logger.error(error_msg)
                                        self.rendering_logger.debug(f"Cell original value: {original_value}")
                    
                    worksheet_duration = time.time() - worksheet_start
                    
                    # Update metrics
                    self.processing_metrics['worksheets_processed'] += 1
                    self.processing_metrics['data_processed']['worksheet_details'][worksheet_name] = {
                        'placeholders_found': len(placeholders_found),
                        'cells_processed': cells_processed,
                        'cells_with_errors': cells_with_errors,
                        'processing_time': worksheet_duration,
                        'data_source': data_source
                    }
                    
                    self.rendering_logger.info(
                        f"✓ Worksheet '{worksheet_name}' processed in {worksheet_duration:.2f}s"
                    )
                    self.rendering_logger.info(
                        f"  - Placeholders found: {len(placeholders_found)}"
                    )
                    self.rendering_logger.info(
                        f"  - Cells processed: {cells_processed}"
                    )
                    if cells_with_errors > 0:
                        self.rendering_logger.warning(
                            f"  - Cells with errors: {cells_with_errors}"
                        )
                    
                except Exception as process_error:
                    worksheet_duration = time.time() - worksheet_start
                    error_msg = f"Error processing worksheet '{worksheet_name}': {str(process_error)}"
                    self.rendering_logger.error(error_msg)
                    self.rendering_logger.debug(f"Processing error details: {type(process_error).__name__}: {str(process_error)}")
                    
                    # Still update metrics for failed worksheet
                    self.processing_metrics['data_processed']['worksheet_details'][worksheet_name] = {
                        'processing_time': worksheet_duration,
                        'error': str(process_error)
                    }
            
            # Final summary
            self.rendering_logger.info(f"=== WORKSHEET PROCESSING COMPLETE ===")
            self.rendering_logger.info(f"Worksheets processed: {self.processing_metrics['worksheets_processed']}")
            
        except Exception as e:
            error_msg = f"Critical error in worksheet processing: {str(e)}"
            self.rendering_logger.error(error_msg)
            self.rendering_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            raise
    
    def _scan_worksheet_placeholders(self, worksheet, worksheet_name: str) -> List[str]:
        """Scan worksheet for placeholders and log findings"""
        placeholders = set()
        
        self.rendering_logger.debug(f"Scanning worksheet '{worksheet_name}' for placeholders...")
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Look for various placeholder patterns
                    import re
                    
                    # Common placeholder patterns
                    patterns = [
                        r'\{\{([^}]+)\}\}',  # {{placeholder}}
                        r'\{([^}]+)\}',      # {placeholder}
                        r'\$\{([^}]+)\}',    # ${placeholder}
                        r'%([^%]+)%',        # %placeholder%
                    ]
                    
                    for pattern in patterns:
                        matches = re.findall(pattern, cell.value)
                        for match in matches:
                            placeholders.add(match.strip())
        
        placeholder_list = list(placeholders)
        self.rendering_logger.debug(f"Found {len(placeholder_list)} unique placeholders in '{worksheet_name}': {placeholder_list}")
        
        return placeholder_list
    
    def _process_cell_value(self, cell_value: str, data: Any, cell_coordinate: str) -> Any:
        """Process individual cell value with placeholder replacement"""
        try:
            # This would integrate with your template engine
            # For now, implementing basic placeholder replacement
            
            if not isinstance(cell_value, str):
                return cell_value
            
            # Log the processing attempt
            self.rendering_logger.debug(f"Processing cell {cell_coordinate} with value: {cell_value[:100]}")
            
            # Here you would call your actual template engine
            # For example: return self.template_engine.render(cell_value, data)
            
            # Placeholder for actual implementation
            processed_value = cell_value  # Replace with actual template processing
            
            return processed_value
            
        except Exception as e:
            self.rendering_logger.error(f"Error processing cell {cell_coordinate}: {str(e)}")
            return cell_value  # Return original value on error
            
    def _apply_worksheet_formatting(self, sheet, template_info: Dict[str, Any]):
        """Apply additional formatting to worksheet"""
        try:
            # Apply any global formatting rules
            formatting_rules = template_info.get('formatting', {})
            
            # Auto-adjust column widths
            if formatting_rules.get('auto_width', True):
                self._auto_adjust_column_widths(sheet)
                
            # Apply borders
            if formatting_rules.get('borders', False):
                self._apply_borders(sheet)
                
        except Exception as e:
            self.logger.error(f"Error applying worksheet formatting: {str(e)}")
            
    def _generate_and_verify_output_enhanced(self, workbook, output_path: str) -> str:
        """Generate output file with comprehensive verification and logging"""
        self.main_logger.info("=== OUTPUT GENERATION & VERIFICATION PHASE START ===")
        
        try:
            generation_start = time.time()
            
            # Pre-save verification
            self.main_logger.info("--- Pre-save Verification ---")
            pre_save_issues = self._verify_workbook_content(workbook, "pre-save")
            
            # Save the workbook
            self.main_logger.info(f"Saving workbook to: {output_path}")
            
            try:
                workbook.save(output_path)
                save_duration = time.time() - generation_start
                self.main_logger.info(f"✓ Workbook saved successfully in {save_duration:.2f}s")
                
                # Verify file was created and is accessible
                import os
                if os.path.exists(output_path):
                    file_size = os.path.getsize(output_path)
                    self.main_logger.info(f"✓ Output file exists: {file_size:,} bytes")
                else:
                    raise FileNotFoundError(f"Output file was not created: {output_path}")
                
            except Exception as save_error:
                error_msg = f"Failed to save workbook: {str(save_error)}"
                self.main_logger.error(error_msg)
                raise
            
            # Post-save verification by re-opening the file
            self.main_logger.info("--- Post-save Verification ---")
            try:
                from openpyxl import load_workbook
                verification_workbook = load_workbook(output_path)
                
                post_save_issues = self._verify_workbook_content(verification_workbook, "post-save")
                
                # Compare pre and post save issues
                if len(post_save_issues) > len(pre_save_issues):
                    self.main_logger.warning("More issues found after saving than before!")
                
                verification_workbook.close()
                
            except Exception as verify_error:
                self.main_logger.warning(f"Could not perform post-save verification: {str(verify_error)}")
                post_save_issues = []
            
            # Generate verification report
            total_duration = time.time() - generation_start
            verification_report = self._generate_verification_report(
                output_path, pre_save_issues, post_save_issues, total_duration
            )
            
            # Update metrics
            self.processing_metrics['output_generated'] = True
            self.processing_metrics['output_path'] = output_path
            self.processing_metrics['verification'] = {
                'pre_save_issues': len(pre_save_issues),
                'post_save_issues': len(post_save_issues),
                'total_duration': total_duration
            }
            
            self.main_logger.info("=== OUTPUT GENERATION & VERIFICATION COMPLETE ===")
            self.main_logger.info(f"Final output: {output_path}")
            self.main_logger.info(f"Total generation time: {total_duration:.2f}s")
            
            if pre_save_issues or post_save_issues:
                self.main_logger.warning(f"Verification found {len(pre_save_issues + post_save_issues)} total issues")
            else:
                self.main_logger.info("✓ No verification issues found")
            
            return output_path
            
        except Exception as e:
            error_msg = f"Critical error in output generation: {str(e)}"
            self.main_logger.error(error_msg)
            self.main_logger.debug(f"Error details: {type(e).__name__}: {str(e)}")
            raise
    
    def _verify_workbook_content(self, workbook, stage: str) -> List[Dict[str, Any]]:
        """Verify workbook content for unrendered placeholders and issues"""
        self.main_logger.debug(f"--- Content Verification ({stage}) ---")
        
        issues = []
        total_cells_checked = 0
        
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self.main_logger.debug(f"Verifying sheet: {sheet_name}")
                
                sheet_issues = 0
                
                for row in sheet.iter_rows():
                    for cell in row:
                        total_cells_checked += 1
                        
                        if cell.value and isinstance(cell.value, str):
                            cell_value = str(cell.value)
                            
                            # Check for various placeholder patterns that might be unrendered
                            placeholder_patterns = [
                                (r'\{\{[^}]+\}\}', 'Double brace placeholder'),
                                (r'\{[^}]+\}', 'Single brace placeholder'),
                                (r'\$\{[^}]+\}', 'Dollar brace placeholder'),
                                (r'%[^%]+%', 'Percent placeholder'),
                                (r'#[A-Z_]+#', 'Hash placeholder'),
                                (r'@[A-Z_]+@', 'At placeholder'),
                            ]
                            
                            for pattern, description in placeholder_patterns:
                                import re
                                matches = re.findall(pattern, cell_value)
                                
                                if matches:
                                    for match in matches:
                                        issue = {
                                            'type': 'unrendered_placeholder',
                                            'sheet': sheet_name,
                                            'cell': cell.coordinate,
                                            'pattern_type': description,
                                            'placeholder': match,
                                            'cell_value': cell_value[:100],
                                            'stage': stage
                                        }
                                        issues.append(issue)
                                        sheet_issues += 1
                            
                            # Check for error indicators
                            error_indicators = ['#ERROR#', '#NULL#', '#MISSING#', 'ERROR:', 'FAILED:']
                            for indicator in error_indicators:
                                if indicator in cell_value.upper():
                                    issue = {
                                        'type': 'error_indicator',
                                        'sheet': sheet_name,
                                        'cell': cell.coordinate,
                                        'indicator': indicator,
                                        'cell_value': cell_value[:100],
                                        'stage': stage
                                    }
                                    issues.append(issue)
                                    sheet_issues += 1
                
                if sheet_issues > 0:
                    self.main_logger.warning(f"Sheet '{sheet_name}': {sheet_issues} issues found")
                else:
                    self.main_logger.debug(f"Sheet '{sheet_name}': No issues found")
            
            self.main_logger.info(f"Content verification ({stage}): {len(issues)} issues in {total_cells_checked:,} cells")
            
            # Log sample issues for debugging
            if issues:
                self.main_logger.debug("Sample issues found:")
                for issue in issues[:5]:  # Show first 5 issues
                    self.main_logger.debug(f"  - {issue['type']} in {issue['sheet']}!{issue['cell']}: {issue.get('placeholder', issue.get('indicator', 'N/A'))}")
            
            return issues
            
        except Exception as e:
            self.main_logger.error(f"Error during content verification: {str(e)}")
            return []
    
    def _generate_verification_report(self, output_path: str, pre_save_issues: List[Dict], 
                                    post_save_issues: List[Dict], duration: float) -> Dict[str, Any]:
        """Generate comprehensive verification report"""
        
        report = {
            'output_file': output_path,
            'generation_duration': duration,
            'verification_timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
            'pre_save_issues': {
                'count': len(pre_save_issues),
                'by_type': {},
                'by_sheet': {}
            },
            'post_save_issues': {
                'count': len(post_save_issues),
                'by_type': {},
                'by_sheet': {}
            },
            'summary': {
                'total_issues': len(pre_save_issues) + len(post_save_issues),
                'issues_resolved': max(0, len(pre_save_issues) - len(post_save_issues)),
                'new_issues': max(0, len(post_save_issues) - len(pre_save_issues)),
                'status': 'PASS' if len(post_save_issues) == 0 else 'ISSUES_FOUND'
            }
        }
        
        # Categorize issues
        for issues, stage in [(pre_save_issues, 'pre_save'), (post_save_issues, 'post_save')]:
            stage_data = report[f'{stage}_issues']
            
            for issue in issues:
                # By type
                issue_type = issue.get('type', 'unknown')
                stage_data['by_type'][issue_type] = stage_data['by_type'].get(issue_type, 0) + 1
                
                # By sheet
                sheet = issue.get('sheet', 'unknown')
                stage_data['by_sheet'][sheet] = stage_data['by_sheet'].get(sheet, 0) + 1
        
        # Log summary
        self.main_logger.info("--- Verification Report Summary ---")
        self.main_logger.info(f"Status: {report['summary']['status']}")
        self.main_logger.info(f"Total issues: {report['summary']['total_issues']}")
        
        if report['summary']['issues_resolved'] > 0:
            self.main_logger.info(f"Issues resolved: {report['summary']['issues_resolved']}")
        
        if report['summary']['new_issues'] > 0:
            self.main_logger.warning(f"New issues: {report['summary']['new_issues']}")
        
        return report
            
    def _auto_adjust_column_widths(self, sheet):
        """Auto-adjust column widths based on content"""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                            
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.error(f"Error auto-adjusting column widths: {str(e)}")
            
    def _apply_borders(self, sheet):
        """Apply borders to data range"""
        try:
            # Find data range
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row > 1 and max_col > 1:
                # Apply thin borders to data range
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        sheet.cell(row=row, column=col).border = thin_border
                        
        except Exception as e:
            self.logger.error(f"Error applying borders: {str(e)}")