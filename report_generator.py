"""
Main Report Generator untuk sistem laporan Excel yang fleksibel
Mengorkestrasi template processing, data retrieval, dan Excel output
"""

import os
import logging
import time
from datetime import datetime, date
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from template_processor import TemplateProcessor
from formula_engine import FormulaEngine
from firebird_connector import FirebirdConnector

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('report_generator_debug.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    Main class untuk generate laporan Excel dari template dan database Firebird
    """
    
    def __init__(self, template_path: str, formula_path: str, db_config: Dict[str, str]):
        """
        Inisialisasi report generator
        
        :param template_path: Path ke file template Excel
        :param formula_path: Path ke file formula definition
        :param db_config: Konfigurasi database Firebird
        """
        start_time = time.time()
        logger.info("=== INITIALIZING REPORT GENERATOR ===")
        logger.debug(f"Template path: {template_path}")
        logger.debug(f"Formula path: {formula_path}")
        logger.debug(f"Database config: {db_config}")
        
        self.template_path = template_path
        self.formula_path = formula_path
        self.db_config = db_config
        
        # Validate file existence
        if not os.path.exists(template_path):
            logger.error(f"Template file not found: {template_path}")
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        if not os.path.exists(formula_path):
            logger.error(f"Formula file not found: {formula_path}")
            raise FileNotFoundError(f"Formula file not found: {formula_path}")
        
        logger.debug(f"Template file size: {os.path.getsize(template_path)} bytes")
        logger.debug(f"Formula file size: {os.path.getsize(formula_path)} bytes")
        
        try:
            # Initialize components
            logger.debug("Initializing database connector...")
            self.db_connector = FirebirdConnector(**db_config)
            logger.debug("Database connector initialized successfully")
            
            logger.debug("Initializing template processor...")
            self.template_processor = TemplateProcessor(template_path, formula_path)
            logger.debug("Template processor initialized successfully")
            
            logger.debug("Initializing formula engine...")
            self.formula_engine = FormulaEngine(formula_path, self.db_connector)
            logger.debug("Formula engine initialized successfully")
            
        except Exception as e:
            logger.error(f"Failed to initialize components: {e}")
            raise
        
        # Working data
        self.workbook = None
        self.variables = {}
        self.repeating_data = {}
        
        init_time = time.time() - start_time
        logger.info(f"Report Generator initialized successfully in {init_time:.3f} seconds")
        logger.debug(f"Template: {template_path}")
        logger.debug(f"Formula: {formula_path}")
        logger.debug(f"Database: {db_config.get('database', 'Unknown')}")
    
    def generate_report(self, parameters: Dict[str, Any], output_path: str) -> bool:
        """
        Generate laporan Excel lengkap
        
        :param parameters: Parameter untuk generate laporan (tanggal, estate, dll)
        :param output_path: Path output file Excel
        :return: True jika berhasil, False jika gagal
        """
        total_start_time = time.time()
        logger.info("=== STARTING REPORT GENERATION ===")
        logger.debug(f"Parameters: {parameters}")
        logger.debug(f"Output path: {output_path}")
        
        try:
            # Step 1: Load template
            step_start = time.time()
            logger.info("Step 1: Loading Excel template...")
            self._load_template()
            step_time = time.time() - step_start
            logger.debug(f"Template loading completed in {step_time:.3f} seconds")
            
            # Step 2: Execute queries dan ambil data
            step_start = time.time()
            logger.info("Step 2: Executing database queries...")
            query_results = self.formula_engine.execute_data_queries(parameters)
            step_time = time.time() - step_start
            logger.debug(f"Query execution completed in {step_time:.3f} seconds")
            logger.debug(f"Query results keys: {list(query_results.keys()) if query_results else 'None'}")
            
            # Step 3: Process variables
            step_start = time.time()
            logger.info("Step 3: Processing variables...")
            self.variables = self.formula_engine.process_variables(query_results, parameters)
            step_time = time.time() - step_start
            logger.debug(f"Variable processing completed in {step_time:.3f} seconds")
            logger.debug(f"Total variables processed: {len(self.variables)}")
            logger.debug(f"Variable names: {list(self.variables.keys())}")
            
            # Step 4: Get repeating data
            step_start = time.time()
            logger.info("Step 4: Getting repeating data...")
            self.repeating_data = self.formula_engine.get_repeating_data(parameters)
            step_time = time.time() - step_start
            logger.debug(f"Repeating data retrieval completed in {step_time:.3f} seconds")
            logger.debug(f"Repeating data sheets: {list(self.repeating_data.keys()) if self.repeating_data else 'None'}")
            
            # Step 5: Process semua sheets
            step_start = time.time()
            logger.info("Step 5: Processing Excel sheets...")
            total_sheets = len(self.workbook.sheetnames)
            logger.debug(f"Total sheets to process: {total_sheets}")
            
            for i, sheet_name in enumerate(self.workbook.sheetnames, 1):
                sheet_start = time.time()
                logger.info(f"  Processing sheet {i}/{total_sheets}: {sheet_name}")
                self._process_sheet(sheet_name)
                sheet_time = time.time() - sheet_start
                logger.debug(f"  Sheet '{sheet_name}' processed in {sheet_time:.3f} seconds")
            
            step_time = time.time() - step_start
            logger.debug(f"All sheets processing completed in {step_time:.3f} seconds")
            
            # Step 6: Save output
            step_start = time.time()
            logger.info("Step 6: Saving output file...")
            self._save_output(output_path)
            step_time = time.time() - step_start
            logger.debug(f"Output saving completed in {step_time:.3f} seconds")
            
            total_time = time.time() - total_start_time
            logger.info(f"=== REPORT GENERATION COMPLETED SUCCESSFULLY ===")
            logger.info(f"Total generation time: {total_time:.3f} seconds")
            logger.info(f"Output file: {output_path}")
            
            # Log final file info
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                logger.debug(f"Output file size: {file_size} bytes")
            
            return True
            
        except Exception as e:
            total_time = time.time() - total_start_time
            logger.error(f"=== REPORT GENERATION FAILED ===")
            logger.error(f"Error after {total_time:.3f} seconds: {e}")
            logger.error("Full traceback:", exc_info=True)
            return False
        finally:
            # Cleanup
            if self.db_connector:
                logger.debug("Closing database connection...")
                self.db_connector.close()
                logger.debug("Database connection closed")
    
    def _load_template(self):
        """Load template Excel file"""
        logger.debug("Loading template file...")
        
        if not os.path.exists(self.template_path):
            logger.error(f"Template file not found: {self.template_path}")
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        try:
            self.workbook = load_workbook(self.template_path)
            sheet_count = len(self.workbook.sheetnames)
            logger.info(f"Template loaded successfully: {sheet_count} sheets")
            logger.debug(f"Sheet names: {self.workbook.sheetnames}")
            
            # Log sheet details
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                max_row = sheet.max_row
                max_col = sheet.max_column
                logger.debug(f"  Sheet '{sheet_name}': {max_row} rows x {max_col} columns")
                
        except Exception as e:
            logger.error(f"Failed to load template: {e}")
            raise
    
    def _process_sheet(self, sheet_name: str):
        """
        Process satu sheet dalam workbook
        
        :param sheet_name: Nama sheet yang akan diproses
        """
        logger.debug(f"Starting to process sheet: {sheet_name}")
        worksheet = self.workbook[sheet_name]
        
        try:
            # Get placeholders untuk sheet ini
            logger.debug(f"Getting placeholders for sheet: {sheet_name}")
            placeholders = self.template_processor.get_placeholders_for_sheet(sheet_name)
            placeholder_count = len(placeholders) if placeholders else 0
            logger.debug(f"Found {placeholder_count} placeholders in sheet '{sheet_name}'")
            
            if placeholders:
                placeholder_types = {}
                for p in placeholders:
                    p_type = p.get('type', 'unknown')
                    placeholder_types[p_type] = placeholder_types.get(p_type, 0) + 1
                logger.debug(f"Placeholder types in '{sheet_name}': {placeholder_types}")
            
            # Process static variables terlebih dahulu
            logger.debug(f"Processing static variables for sheet: {sheet_name}")
            self._replace_static_variables(worksheet, placeholders)
            
            # Process repeating sections
            if sheet_name in self.repeating_data:
                logger.debug(f"Processing repeating sections for sheet: {sheet_name}")
                self._process_repeating_sections(worksheet, sheet_name)
            else:
                logger.debug(f"No repeating data found for sheet: {sheet_name}")
            
            # Apply formatting jika diperlukan
            logger.debug(f"Applying sheet formatting for: {sheet_name}")
            self._apply_sheet_formatting(worksheet, sheet_name)
            
            logger.debug(f"Sheet '{sheet_name}' processing completed successfully")
            
        except Exception as e:
            logger.error(f"Error processing sheet '{sheet_name}': {e}")
            raise
    
    def _replace_static_variables(self, worksheet: Worksheet, placeholders: List[Dict]):
        """
        Replace static variables dalam worksheet
        
        :param worksheet: Worksheet yang akan diproses
        :param placeholders: List placeholders yang ditemukan
        """
        if not placeholders:
            logger.debug("No placeholders to process")
            return
            
        variable_placeholders = [p for p in placeholders if p.get('type') == 'variable']
        logger.debug(f"Processing {len(variable_placeholders)} variable placeholders")
        
        successful_replacements = 0
        failed_replacements = 0
        
        for placeholder in variable_placeholders:
            try:
                var_name = placeholder.get('name', '')
                cell_address = placeholder.get('cell', '')
                
                if not var_name or not cell_address:
                    logger.warning(f"Invalid placeholder data: {placeholder}")
                    failed_replacements += 1
                    continue
                
                # Get value dari variables
                if var_name in self.variables:
                    value = self.variables[var_name]
                    successful_replacements += 1
                    logger.debug(f"Variable '{var_name}' -> {cell_address}: {value}")
                else:
                    value = f"{{VAR_NOT_FOUND: {var_name}}}"
                    failed_replacements += 1
                    logger.warning(f"Variable '{var_name}' not found in variables")
                
                # Set value ke cell
                worksheet[cell_address] = value
                
            except Exception as e:
                failed_replacements += 1
                logger.error(f"Error replacing variable placeholder {placeholder}: {e}")
        
        logger.info(f"Variable replacement completed: {successful_replacements} successful, {failed_replacements} failed")
    
    def _process_repeating_sections(self, worksheet: Worksheet, sheet_name: str):
        """
        Process repeating sections dalam worksheet
        
        :param worksheet: Worksheet yang akan diproses
        :param sheet_name: Nama sheet
        """
        sheet_repeating_data = self.repeating_data.get(sheet_name, {})
        logger.debug(f"Processing {len(sheet_repeating_data)} repeating sections for sheet '{sheet_name}'")
        
        for section_name, section_data in sheet_repeating_data.items():
            section_start = time.time()
            logger.info(f"    Processing repeating section: {section_name}")
            logger.debug(f"Section data type: {type(section_data)}")
            logger.debug(f"Section data length: {len(section_data) if hasattr(section_data, '__len__') else 'N/A'}")
            
            try:
                # Get section configuration
                section_config = self._get_section_config(sheet_name, section_name)
                if not section_config:
                    logger.warning(f"No configuration found for section '{section_name}' in sheet '{sheet_name}'")
                    continue
                
                logger.debug(f"Section config: {section_config}")
                
                # Process section
                self._process_single_repeating_section(
                    worksheet, section_config, section_data
                )
                
                section_time = time.time() - section_start
                logger.debug(f"Section '{section_name}' processed in {section_time:.3f} seconds")
                
            except Exception as e:
                logger.error(f"Error processing repeating section '{section_name}': {e}")
    
    def _get_section_config(self, sheet_name: str, section_name: str) -> Optional[Dict]:
        """Get configuration untuk repeating section"""
        try:
            config = self.formula_engine.formulas.get('repeating_sections', {}).get(sheet_name, {}).get(section_name)
            logger.debug(f"Retrieved config for section '{section_name}' in sheet '{sheet_name}': {config is not None}")
            return config
        except Exception as e:
            logger.error(f"Error getting section config for '{section_name}' in '{sheet_name}': {e}")
            return None
    
    def _process_single_repeating_section(self, worksheet: Worksheet, config: Dict, data: List[Dict]):
        """
        Process satu repeating section
        
        :param worksheet: Worksheet
        :param config: Konfigurasi section
        :param data: Data untuk section
        """
        start_row = config.get('start_row', 1)
        template_rows = config.get('template_rows', 1)
        columns_mapping = config.get('columns', {})
        
        logger.debug(f"Section config - start_row: {start_row}, template_rows: {template_rows}")
        logger.debug(f"Columns mapping: {list(columns_mapping.keys())}")
        
        if not data:
            logger.info(f"      No data for section, skipping...")
            return
        
        data_count = len(data)
        logger.info(f"      Processing {data_count} rows starting from row {start_row}")
        
        try:
            # Insert rows jika diperlukan
            rows_needed = data_count * template_rows
            current_rows = template_rows  # Asumsi template sudah ada 1 set rows
            
            if rows_needed > current_rows:
                rows_to_insert = rows_needed - current_rows
                logger.debug(f"Inserting {rows_to_insert} rows at position {start_row + current_rows}")
                worksheet.insert_rows(start_row + current_rows, rows_to_insert)
                
                # Copy formatting dari template rows
                logger.debug("Copying template formatting to new rows...")
                self._copy_template_formatting(worksheet, start_row, template_rows, rows_to_insert)
            
            # Fill data
            successful_cells = 0
            failed_cells = 0
            
            for i, row_data in enumerate(data):
                current_row = start_row + (i * template_rows)
                logger.debug(f"Processing data row {i+1}/{data_count} at Excel row {current_row}")
                
                for col_name, col_config in columns_mapping.items():
                    try:
                        col_letter = col_config.get('column', 'A')
                        field_name = col_config.get('field', col_name)
                        format_type = col_config.get('format', 'text')
                        
                        # Get value dari row data
                        value = row_data.get(field_name, '')
                        
                        # Format value
                        formatted_value = self._format_cell_value(value, format_type)
                        
                        # Set ke cell
                        cell_address = f"{col_letter}{current_row}"
                        worksheet[cell_address] = formatted_value
                        successful_cells += 1
                        
                        logger.debug(f"Cell {cell_address}: {field_name} = {formatted_value}")
                        
                        # Apply cell formatting jika ada
                        if 'cell_format' in col_config:
                            self._apply_cell_formatting(worksheet[cell_address], col_config['cell_format'])
                            
                    except Exception as e:
                        failed_cells += 1
                        logger.error(f"Error setting cell value for column '{col_name}': {e}")
            
            logger.info(f"      Section data filling completed: {successful_cells} successful, {failed_cells} failed")
            
        except Exception as e:
            logger.error(f"Error processing single repeating section: {e}")
            raise
    
    def _copy_template_formatting(self, worksheet: Worksheet, start_row: int, template_rows: int, rows_to_insert: int):
        """
        Copy formatting dari template rows ke rows yang baru diinsert
        
        :param worksheet: Worksheet
        :param start_row: Row awal template
        :param template_rows: Jumlah template rows
        :param rows_to_insert: Jumlah rows yang diinsert
        """
        try:
            # Get template range
            template_end_row = start_row + template_rows - 1
            
            # Copy formatting untuk setiap row yang diinsert
            for i in range(rows_to_insert):
                source_row = start_row + (i % template_rows)
                target_row = template_end_row + 1 + i
                
                # Copy formatting untuk setiap column yang ada data
                for col in range(1, worksheet.max_column + 1):
                    source_cell = worksheet.cell(row=source_row, column=col)
                    target_cell = worksheet.cell(row=target_row, column=col)
                    
                    # Copy formatting
                    if source_cell.font:
                        target_cell.font = Font(
                            name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            color=source_cell.font.color
                        )
                    
                    if source_cell.border:
                        target_cell.border = Border(
                            left=source_cell.border.left,
                            right=source_cell.border.right,
                            top=source_cell.border.top,
                            bottom=source_cell.border.bottom
                        )
                    
                    if source_cell.fill:
                        target_cell.fill = PatternFill(
                            fill_type=source_cell.fill.fill_type,
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color
                        )
                    
                    if source_cell.alignment:
                        target_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical,
                            wrap_text=source_cell.alignment.wrap_text
                        )
                    
                    # Copy number format
                    target_cell.number_format = source_cell.number_format
        
        except Exception as e:
            print(f"      Warning: Could not copy template formatting: {e}")
    
    def _format_cell_value(self, value: Any, format_type: str) -> Any:
        """
        Format value berdasarkan format type
        
        :param value: Value yang akan diformat
        :param format_type: Tipe format
        :return: Formatted value
        """
        if value is None:
            return ''
        
        if format_type == 'number':
            try:
                return float(value)
            except:
                return 0
        
        elif format_type == 'integer':
            try:
                return int(float(value))
            except:
                return 0
        
        elif format_type == 'currency':
            try:
                return float(value)
            except:
                return 0
        
        elif format_type == 'percentage':
            try:
                return float(value) / 100  # Excel akan format sebagai percentage
            except:
                return 0
        
        elif format_type == 'date':
            if isinstance(value, (date, datetime)):
                return value
            else:
                try:
                    return datetime.strptime(str(value), '%Y-%m-%d').date()
                except:
                    return str(value)
        
        else:  # text
            return str(value)
    
    def _apply_cell_formatting(self, cell, format_config: Dict):
        """
        Apply formatting ke cell berdasarkan konfigurasi
        
        :param cell: Cell yang akan diformat
        :param format_config: Konfigurasi formatting
        """
        try:
            # Font formatting
            if 'font' in format_config:
                font_config = format_config['font']
                cell.font = Font(
                    name=font_config.get('name', 'Calibri'),
                    size=font_config.get('size', 11),
                    bold=font_config.get('bold', False),
                    italic=font_config.get('italic', False)
                )
            
            # Border formatting
            if 'border' in format_config:
                border_config = format_config['border']
                side_style = border_config.get('style', 'thin')
                cell.border = Border(
                    left=Side(style=side_style),
                    right=Side(style=side_style),
                    top=Side(style=side_style),
                    bottom=Side(style=side_style)
                )
            
            # Fill formatting
            if 'fill' in format_config:
                fill_config = format_config['fill']
                cell.fill = PatternFill(
                    fill_type='solid',
                    start_color=fill_config.get('color', 'FFFFFF')
                )
            
            # Alignment
            if 'alignment' in format_config:
                align_config = format_config['alignment']
                cell.alignment = Alignment(
                    horizontal=align_config.get('horizontal', 'left'),
                    vertical=align_config.get('vertical', 'center'),
                    wrap_text=align_config.get('wrap_text', False)
                )
            
            # Number format
            if 'number_format' in format_config:
                cell.number_format = format_config['number_format']
        
        except Exception as e:
            print(f"      Warning: Could not apply cell formatting: {e}")
    
    def _apply_sheet_formatting(self, worksheet: Worksheet, sheet_name: str):
        """
        Apply formatting khusus untuk sheet
        
        :param worksheet: Worksheet
        :param sheet_name: Nama sheet
        """
        try:
            # Get sheet formatting config jika ada
            sheet_config = self.formula_engine.formulas.get('sheet_formatting', {}).get(sheet_name, {})
            
            # Apply column widths
            if 'column_widths' in sheet_config:
                for col_letter, width in sheet_config['column_widths'].items():
                    worksheet.column_dimensions[col_letter].width = width
            
            # Apply row heights
            if 'row_heights' in sheet_config:
                for row_num, height in sheet_config['row_heights'].items():
                    worksheet.row_dimensions[int(row_num)].height = height
            
            # Apply print settings
            if 'print_settings' in sheet_config:
                print_config = sheet_config['print_settings']
                
                if 'orientation' in print_config:
                    worksheet.page_setup.orientation = print_config['orientation']
                
                if 'paper_size' in print_config:
                    worksheet.page_setup.paperSize = print_config['paper_size']
                
                if 'margins' in print_config:
                    margins = print_config['margins']
                    worksheet.page_margins.left = margins.get('left', 0.7)
                    worksheet.page_margins.right = margins.get('right', 0.7)
                    worksheet.page_margins.top = margins.get('top', 0.75)
                    worksheet.page_margins.bottom = margins.get('bottom', 0.75)
        
        except Exception as e:
            print(f"      Warning: Could not apply sheet formatting: {e}")
    
    def _save_output(self, output_path: str):
        """
        Save workbook ke output path
        
        :param output_path: Path output file
        """
        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Save workbook
        self.workbook.save(output_path)
        print(f"Report saved to: {output_path}")
    
    def validate_template(self) -> Dict[str, Any]:
        """
        Validasi template dan formula configuration
        
        :return: Dictionary berisi hasil validasi
        """
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'info': {}
        }
        
        try:
            # Validate template file
            if not os.path.exists(self.template_path):
                validation_result['errors'].append(f"Template file not found: {self.template_path}")
                validation_result['valid'] = False
            
            # Validate formula file
            if not os.path.exists(self.formula_path):
                validation_result['errors'].append(f"Formula file not found: {self.formula_path}")
                validation_result['valid'] = False
            
            if validation_result['valid']:
                # Load dan validate template
                temp_processor = TemplateProcessor(self.template_path)
                placeholders = temp_processor.get_all_placeholders()
                
                validation_result['info']['total_placeholders'] = len(placeholders)
                validation_result['info']['sheets'] = list(temp_processor.workbook.sheetnames)
                
                # Validate formula definitions
                formula_engine = FormulaEngine(self.formula_path, None)
                
                validation_result['info']['total_queries'] = len(formula_engine.formulas.get('queries', {}))
                validation_result['info']['total_variables'] = len(formula_engine.formulas.get('variables', {}))
                
                # Check if all placeholders have corresponding variables
                variable_names = set(formula_engine.formulas.get('variables', {}).keys())
                placeholder_names = set([p['name'] for p in placeholders if p['type'] == 'variable'])
                
                missing_variables = placeholder_names - variable_names
                if missing_variables:
                    validation_result['warnings'].extend([
                        f"Placeholder '{var}' tidak memiliki definisi variable" 
                        for var in missing_variables
                    ])
                
                unused_variables = variable_names - placeholder_names
                if unused_variables:
                    validation_result['warnings'].extend([
                        f"Variable '{var}' tidak digunakan dalam template" 
                        for var in unused_variables
                    ])
        
        except Exception as e:
            validation_result['errors'].append(f"Validation error: {e}")
            validation_result['valid'] = False
        
        return validation_result