"""
Script untuk membuat sample Excel template untuk FFB Analysis Report
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def create_sample_template():
    """Membuat sample Excel template"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet dan buat sheets baru
    wb.remove(wb.active)
    
    # Create Summary sheet
    summary_sheet = wb.create_sheet("Summary")
    create_summary_sheet(summary_sheet)
    
    # Create Detail sheet  
    detail_sheet = wb.create_sheet("Detail")
    create_detail_sheet(detail_sheet)
    
    # Save template
    template_path = r"d:\Gawean Rebinmas\Monitoring Database\Ifes Auto report\excel_method\sample_template.xlsx"
    wb.save(template_path)
    print(f"Sample template created: {template_path}")

def create_summary_sheet(ws):
    """Membuat Summary sheet dengan header dan placeholders"""
    
    # Header styling
    header_font = Font(name='Arial', size=14, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Fill colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    # Title
    ws['A1'] = '{report_title}'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    # Estate info
    ws['A3'] = 'Estate:'
    ws['B3'] = '{estate_name}'
    ws['A4'] = 'Periode:'
    ws['B4'] = '{report_period}'
    ws['A5'] = 'Tanggal Laporan:'
    ws['B5'] = '{generated_date}'
    
    # Summary section header
    ws['A7'] = 'RINGKASAN HARIAN'
    ws['A7'].font = subheader_font
    ws['A7'].fill = subheader_fill
    ws.merge_cells('A7:E7')
    
    # Table headers
    headers = ['Tanggal', 'Berat Bruto (Kg)', 'Berat Netto (Kg)', 'Total Nilai (Rp)', 'Jumlah Trip']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = subheader_font
        cell.fill = header_fill
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Template row untuk daily data (row 10 akan diisi oleh repeating section)
    ws['A10'] = 'Template Row'  # Akan diganti dengan data
    ws['B10'] = 0
    ws['C10'] = 0  
    ws['D10'] = 0
    ws['E10'] = 0
    
    # Apply formatting ke template row
    for col in range(1, 6):
        cell = ws.cell(row=10, column=col)
        cell.border = thin_border
        if col == 1:
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(horizontal='right')
    
    # Summary totals section
    ws['A15'] = 'TOTAL KESELURUHAN'
    ws['A15'].font = subheader_font
    ws['A15'].fill = subheader_fill
    ws.merge_cells('A15:B15')
    
    ws['A17'] = 'Total Berat Bruto:'
    ws['B17'] = '{total_berat_bruto}'
    ws['C17'] = 'Kg'
    
    ws['A18'] = 'Total Berat Netto:'
    ws['B18'] = '{total_berat_netto}'
    ws['C18'] = 'Kg'
    
    ws['A19'] = 'Total Potongan:'
    ws['B19'] = '{total_potongan}'
    ws['C19'] = 'Kg'
    
    ws['A20'] = 'Total Nilai:'
    ws['B20'] = '{total_nilai}'
    ws['C20'] = 'Rp'
    
    ws['A21'] = 'Jumlah Transaksi:'
    ws['B21'] = '{jumlah_transaksi}'
    ws['C21'] = 'Trip'
    
    ws['A22'] = 'Rata-rata Harga:'
    ws['B22'] = '{rata_harga}'
    ws['C22'] = 'Rp/Kg'
    
    ws['A23'] = 'Efisiensi Berat:'
    ws['B23'] = '{efisiensi_berat}'
    ws['C23'] = '%'
    
    ws['A24'] = 'Status Laporan:'
    ws['B24'] = '{status_laporan}'
    
    # Apply formatting ke summary section
    for row in range(17, 25):
        ws[f'A{row}'].font = normal_font
        ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12

def create_detail_sheet(ws):
    """Membuat Detail sheet dengan tabel transaksi"""
    
    # Fonts
    header_font = Font(name='Arial', size=14, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Fill colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    # Title
    ws['A1'] = 'DETAIL TRANSAKSI FFB'
    ws['A1'].font = header_font
    ws.merge_cells('A1:H1')
    
    # Estate info
    ws['A3'] = 'Estate: {estate_name}'
    ws['A4'] = 'Periode: {report_period}'
    
    # Detail transaction table header
    ws['A6'] = 'DAFTAR TRANSAKSI'
    ws['A6'].font = subheader_font
    ws['A6'].fill = subheader_fill
    ws.merge_cells('A6:H6')
    
    # Table headers
    headers = ['No', 'Tanggal', 'Jenis Buah', 'Berat Bruto', 'Berat Netto', 'Potongan', 'Harga/Kg', 'Total Harga']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Template row untuk transaction data (row 8 akan diisi oleh repeating section)
    template_data = [1, 'Template Date', 'Template Fruit', 0, 0, 0, 0, 0]
    for col, value in enumerate(template_data, 1):
        cell = ws.cell(row=8, column=col, value=value)
        cell.border = thin_border
        if col == 1 or col == 2:
            cell.alignment = Alignment(horizontal='center')
        elif col == 3:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = Alignment(horizontal='right')
    
    # Summary by fruit type section
    ws['A23'] = 'RINGKASAN PER JENIS BUAH'
    ws['A23'].font = subheader_font
    ws['A23'].fill = subheader_fill
    ws.merge_cells('A23:E23')
    
    # Fruit summary headers
    fruit_headers = ['Jenis Buah', 'Total Bruto', 'Total Netto', 'Total Nilai', 'Rata-rata Harga']
    for col, header in enumerate(fruit_headers, 1):
        cell = ws.cell(row=24, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Template row untuk fruit summary (row 25 akan diisi oleh repeating section)
    fruit_template = ['Template Fruit', 0, 0, 0, 0]
    for col, value in enumerate(fruit_template, 1):
        cell = ws.cell(row=25, column=col, value=value)
        cell.border = thin_border
        if col == 1:
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(name='Arial', size=10, bold=True)
        else:
            cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

if __name__ == "__main__":
    create_sample_template()